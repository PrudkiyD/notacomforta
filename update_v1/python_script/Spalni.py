import requests
import openpyxl
from bs4 import BeautifulSoup
import csv
import time
import re
import os
import json
from main.models import Seria, Stinka, Stinka_img
from pars.models import File


with open('/home/ay507291/notacomforta.pl.ua/www/pars/src.json', 'r') as f:
    file = json.load(f)

path_dirver = file['src'][0]['path_dirver']
accept = file['src'][1]['accept']
user_agent = file['src'][2]['user-agent']

HEADERS = {
    'accept': accept,
    'user-agent': user_agent
}

HOST = 'https://komfortmebli.com.ua'

path = os.getcwd()


def num_check(text):
    num = ''
    for t in text:
        if t.isdigit():
            num += t
    return num


def test(modul_cards, product_cards):
    pass
    for m in modul_cards:
        print('--->', m)
        for p in product_cards:
            if m['id'] == p['modul_id']:
                print('    ', p )


def UpdateData(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['name']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['name']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['name'] == el.pars_name:
                        new_db.append(p['name'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(price=p['price'], pars_category='spalni')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['name'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          spalni=True, name=p['name'], pars_name=p['name'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'], pars_category='spalni')
                    new_db.append(p['name'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass

    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(spalni=True):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def UpdateDataNeman(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(manufacturer_id=manu).filter(pars_name = m['prom']):
            modul_id = Seria.objects.get(pars_name = m['prom']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['prom'] == el.pars_name:
                        new_db.append(p['prom'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(price=p['price'], pars_category='spalni')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          spalni=True, name=p['name'], name_ru=p['name_ru'],\
                                            pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], description_ru=p['des_ru'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass

    '''#Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(spalni=True):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()'''


def UpdateDataSvitMebliv(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['prom']).id

            print('-->', m['name'])
            for p in product_cards:
                if p['modul_id'] == m['id']:
                    if Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id):
                        new_db.append(p['prom'])
                        print('old:', p['name'])
                        Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id).update(price=p['price'], pars_category='spalni')
                    else:
                        print('none:', p['name'])
                        new_db.append(p['prom'])
                        Stinka.objects.create(seria_id=modul_id, manufacturer_id=manu,\
                                          spalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          spalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
    
    #Видаляємо товар
    '''for m in Stinka.objects.filter(manufacturer_id=manu).filter(spalni=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()'''


def UpdateDataLion(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['prom']).id

            print('-->', m['name'])
            for p in product_cards:
                if p['modul_id'] == m['id']:
                    if Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id):
                        new_db.append(p['prom'])
                        print('old:', p['name'])
                        Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id).update(pars_category='spalni')
                    else:
                        print('none:', p['name'])
                        new_db.append(p['prom'])
                        Stinka.objects.create(seria_id=modul_id, manufacturer_id=manu,\
                                          spalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          spalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
    
    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(spalni=True):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def UpdateDataEverest(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['name']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['name']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['name'] == el.pars_name:
                        new_db.append(p['name'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(pars_category='spalni')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['name'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          spalni=True, name=p['name'], pars_name=p['name'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['name'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass

    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(spalni=True):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def UpdateDataKM(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(manufacturer_id=manu).filter(pars_name = m['prom']):
            modul_id = Seria.objects.get(pars_name = m['prom']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['prom'] == el.pars_name:
                        new_db.append(p['prom'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(price=p['price'], pars_category='spalni')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          spalni=True, name=p['name'],\
                                            pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass

    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(spalni=True):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()



def get_spalni_products_comfortmebli():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=22).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    
    modul_cards = []
    product_cards = []
    img_cards = []
    modul_id = 0
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            modul_id += 1

            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][9].value).split(';')
            img_list.insert(0, str(sheet[r][7].value)) 

            des = str(sheet[r][10].value)


            product_cards.append({
                'modul_id':modul_id,
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
                'price':price
            })

            modul_cards.append({
                'prom':prom,
                'id':modul_id,
                'name':name,
            })

            print(prom, '-->', name, '->', price)

            for img in img_list:
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })

                except Exception as ex:
                    print(f'///-------{ex}---------///')

            print('-------------------------')

    test(modul_cards, product_cards)
    UpdateDataKM(modul_cards, product_cards, img_cards, manu=1)


def get_spalni_products_everest():
    source = requests.get('https://everestm.com.ua/krovaty/gotovyye-nabory-dlya-spalni/',headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')
    item = soup.find('ul', class_='pagination').find_all('a')
    pag_list = ['https://everestm.com.ua/krovaty/gotovyye-nabory-dlya-spalni/', ]

    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0
    img_cards = []
    link_list = []

    # Забераємо пагенацію

    for i in item:
        if i.get('href') not in pag_list:
            pag_list.append(i.get('href'))
    time.sleep(2)

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        item = soup.find('div', class_='products-block').find_all('a', class_='product-name')
        for i in item:
            if i.get('href') not in link_list:
                link_list.append(i.get('href'))

        time.sleep(2)

    # Забераємо інформацію про товар

    for i in link_list:
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        name = soup.find('h1').get_text()
        desc_text = soup.find('div', class_='nav-desc').get_text()
        char = soup.find('div', class_='nav-characteristic').get_text()
        desc_text = desc_text + char

        try:
            price = soup.find('span', class_='autocalc-product-special').get_text()

        except:
            price = soup.find('span', class_='autocalc-product-price').get_text()

        modul_cards.append({
            'id': modul_id,
            'name': name
        })

        product_cards.append({
            'modul_id': modul_id,
            'id': prod_id,
            'name': name,
            'w': '',
            'des': desc_text,
            'price': price
        })
        
        try:
            img = soup.find('a', class_='thumbnail').get('href')
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name.replace('%', ''), "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': prod_id,
                'img': img_name.replace('%', '')
            })
        except:
            pass
        
        modul_id += 1
        prod_id += 1
    
    test(modul_cards, product_cards)
    UpdateDataEverest(modul_cards, product_cards, img_cards, manu=9)


def get_spalni_products_matrolux():
    source = requests.get('https://matroluxe.ua/ua/mebel-dlya-spalni').text
    soup = BeautifulSoup(source, 'html.parser')
    time.sleep(1)
    
    pag_list = ['https://matroluxe.ua/ua/mebel-dlya-spalni',]
    prod_link = []
    pagination = soup.find('div', class_='pagpages clearfix').find_all('a')

    modul_cards = []
    product_cards = []
    img_cards=[]
    modul_id = 0
    prod_id = 0

    #Забераємо пагенацію

    for p in pagination:
        pag = p.get('href')
        if pag not in pag_list:
            pag_list.append(pag)
            print('------->', pag)
            

    #Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url).text
        soup = BeautifulSoup(source, 'html.parser')

        links = soup.find_all('div', class_='product-layout product-grid col-lg-3 col-md-4 col-sm-6 col-xs-12')

        for l in links:
            link = l.find('a').get('href')

            if link not in prod_link:
                prod_link.append({
                    'num' : len(prod_link),
                    'url': link
                })
                print(len(prod_link) - 1, ':', link)

        time.sleep(1)

    #Забераємо інформацію про товар

    for url in prod_link:

        source = requests.get(url['url']).text
        soup = BeautifulSoup(source, 'html.parser')

        img_link = soup.find('div', class_='images_box').find_all('div', class_='item')
        name = soup.find('h1').get_text()
        price = soup.find('span', class_='autocalc-product-price').get_text()
        price = price[:-4]
        desc_text = soup.find('div', 'tab-content').find('div', class_='tab-pane active').get_text()

        modul_cards.append({
            'id': modul_id,
            'name': name
        })

        product_cards.append({
            'modul_id': modul_id,
            'id': prod_id,
            'name': name,
            'w': '',
            'des': desc_text,
            'price': price
        })


        for i in img_link:
            img = i.find('a').get('href')
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': url['num'],
                'img': img_name
            })
            
        modul_id += 1
        prod_id += 1
        time.sleep(1)

    test(modul_cards, product_cards)
    UpdateData(modul_cards, product_cards, img_cards, manu=3)


def get_spalni_products_neman():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=9).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet_1 = book.worksheets[0]

    modul_cards = []
    product_cards = []
    img_cards = []
    modul_id = 0
    prod_id = 0

    for r in range(sheet_1.max_row):
        r += 1

        cell = str(sheet_1[r][14].value).split('/')


        if str(cell[len(cell)-1]) == 'Комплекти спален':
            prom = str(sheet_1[r][0].value)
            name = str(sheet_1[r][5].value)
            des = str(sheet_1[r][37].value)
            name_ru = str(sheet_1[r][6].value)
            des_ru = str(sheet_1[r][38].value)
            price = str(sheet_1[r][9].value)[:-3]
            img_list = str(sheet_1[r][15].value).split(';')
            img_list = [line.rstrip() for line in img_list]

            print(name)

            modul_cards.append({
                'id':modul_id,
                'prom':prom,
                'name':name
            })

            product_cards.append({
                'modul_id':modul_id,
                'id':prod_id,
                'prom':prom,
                'name':name,
                'name_ru':name_ru,
                'w':'',
                'des':des,
                'des_ru':des_ru,
                'price':price
            })

            for i in img_list:
                try:
                    img = i
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_name = str(img_name).replace('%', '')
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })
                except:
                    pass

            prod_id += 1
            modul_id += 1

    UpdateDataNeman(modul_cards, product_cards, img_cards, manu=10)


def get_spalni_products_lion():
    modul_cards = []
    product_cards = []
    img_cards = []
    modul_id = 0
    prod_id = 0
    link_list = ['https://lion-mebli.com.ua/spalni/',]


    source = requests.get('https://lion-mebli.com.ua/spalni/',headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')

    #Пагінація
    pag_list = soup.find('ul', class_='pagination').find_all('a')
    for i in pag_list:
        
        link = i.get('href')

        if link not in link_list:
            link_list.append(link)



    #Збираємо товар
    prod_list=[]
    for url in link_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        items = soup.find_all('div', class_='category_product')

        for i in items:
            link = i.find('a').get('href')
            if link not in prod_list:
                prod_list.append(link)


    #Збираємо інформацію про товар
    for p in prod_list:
        source = requests.get(p,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        prod_id += 1
        prom = p
        name = soup.find('h1').getText()
        des = str(soup.find('div', class_='product_description'))
        img_list = soup.find('div', class_='product_small_nav').find_all('img')

        print(f"///---{name}---///")

        modul_cards.append({
            'id':prod_id,
            'prom':prom,
            'name':name,
        })

        product_cards.append({
            'modul_id':prod_id,
            'id':prod_id,
            'prom':prom,
            'name':name,
            'des':des,
            'price':''
        })

        for i in img_list:
            for i in img_list:
                try:
                    img = i.get('src')

                    img_cards.append({
                        'id': prod_id,
                        'img': img
                    })
                except:
                    pass
        
        time.sleep(1)

    #test(modul_cards, product_cards)
    UpdateDataLion(modul_cards, product_cards, img_cards, manu=6)


def get_spalni_products_svitmebliv():
    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0

    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=6).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[2]

    for r in range(sheet.max_row):
        r += 1


        if r == 100:
            break

        for c in range(sheet.max_column):
            cell = str(sheet[r][c].value)
            match_1 = re.search("спальня „", cell)
            match_2 = re.search('спальня "', cell)
            if match_1 or match_2:

                modul_id += 1
                prod_id += 1

                name = cell.replace('*','')
                column = c + 4
                row = r

                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom': prom,
                    'name': name
                })

                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'name': name,
                    'prom': prom,
                    'w': '',
                    'des': '',
                    'price': ''
                })

                print(f"Спальня: {name}")

                while True:
                    row += 1
                    try:
                        name = str(sheet[row][c].value)
                        price = int(sheet[row][column].value)
                        prod_id += 1
                        prom = name.replace(' ', '').lower()

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'name': name,
                            'prom': prom,
                            'w': '',
                            'des': '',
                            'price': price
                        })

                    except:
                        break
            else:
                pass

    sheet = book.worksheets[3]

    for r in range(sheet.max_row):
        r += 1


        if r == 100:
            break

        for c in range(sheet.max_column):
            cell = str(sheet[r][c].value)
            match_1 = re.search("спальня „", cell)
            match_2 = re.search('спальня "', cell)
            if match_1 or match_2:

                modul_id += 1
                prod_id += 1

                name = cell.replace('*','')
                column = c + 4
                row = r
                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })

                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'name': name,
                    'prom':prom,
                    'w': '',
                    'des': '',
                    'price': price
                })

                while True:
                    row += 1
                    try:
                        name = str(sheet[row][c].value)
                        price = int(sheet[row][column].value)
                        prom = name.replace(' ', '').lower()
                        
                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'name': name,
                            'prom':prom,
                            'w': '',
                            'des': '',
                            'price': price
                        })
                        prod_id += 1

                    except:
                        break
            else:
                pass

    #test(modul_cards, product_cards)
    UpdateDataSvitMebliv(modul_cards, product_cards, img_cards=[], manu=2)
   
        
        