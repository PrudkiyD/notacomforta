import requests
import xml.etree.ElementTree as ET
import openpyxl
from bs4 import BeautifulSoup
import csv
import time
import re
import os
import json
from pcstoly.models import Pcstoly, Pcstoly_size, Pcstoly_img
from pars.models import File


with open('/home/ay507291/notacomforta.pl.ua/www/pars/src.json', 'r') as f:
    file = json.load(f)

path_dirver = file['src'][0]['path_dirver']
accept = file['src'][1]['accept']
user_agent = file['src'][2]['user-agent']

HEADERS = {
    'accept': accept,
    'user-agent': user_agent
}


def num_check(text):
    num = ''
    for t in str(text):
        if t.isdigit():
            num += t
    return num



def get_pcstoly_kompanit():
    type_link = ['https://kompanit.com.ua/catalog/stoli-kompyuterni/c20',
                 'https://kompanit.com.ua/catalog/stoli-pismovi/c21',
                 'https://kompanit.com.ua/catalog/stoli-loft/c34'
                 ]
    
    pag_list = []
    prod_link = []
    cards = []
    size_cards = []
    img_cards = []

    for i in type_link:
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        pag_list.append(i)
        pagination = soup.find_all('a', class_='pagination__number')

        # Забераємо пагенацію

        for p in pagination:
            pag = p.get('href')
            if pag not in pag_list and pag:
                pag_list.append(pag)
        time.sleep(1)

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        links = soup.find_all('div', class_='gcell gcell--12 gcell--xs-6 gcell--md-4')

        for l in links:
            link = l.find('a').get('href')

            if link not in prod_link:
                prod_link.append({
                    'num': len(prod_link),
                    'url': link
                })

        time.sleep(1)

    # Забераємо інформацію про товар

    for url in prod_link:
        source = requests.get(url['url'],headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        name = soup.find('h1').get_text()
        color_desc = soup.find('div', class_='item-colors')
        desc_text = soup.find('div', class_='tabs _mb-sm').get_text()
        slick = soup.find_all('a', class_='slider__slide slick-slide')
        prop_list = soup.find('div', class_='tabs _mb-sm').find_all('li')
        size = ['', '']


        print(url['url'])
        print(name)

        for i in prop_list:
            val = i.get_text()

            w_match = re.search('Ширина', val)

            d_match = re.search('Глибина', val)

            if w_match:
                size[0] = num_check(val)

            if d_match:
                size[1] = num_check(val)

        cards.append({
            'id': url['num'],
            'name': name,
            'des': desc_text + str(color_desc),
        })

        size_cards.append({
            'id': url['num'],
            'w': size[0],
            'd': size[1],
            'price': '',
        })

        img_list = soup.find('div', class_='gcell gcell--12 gcell--def-6').find_all('a')

        for i in img_list:
            img = i.get('data-mfp-src')
            print(img)
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': url['num'],
                'img': img_name
            })

            print(img_name)

        print(size)
        print('-------------------------------')


    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Pcstoly.objects.filter(pars_name = product['name']):
                print('old', Pcstoly.objects.get(pars_name=product['name']).name)
                model_id =  Pcstoly.objects.get(pars_name=product['name']).id
                size_id = Pcstoly_size.objects.filter(pcstoly_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Pcstoly.objects.all().order_by('-id')[0].id + 1

                Pcstoly.objects.get_or_create(id=new_id, manufacturer_id=19, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.get_or_create(pcstoly_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Pcstoly_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_pcstoly_everest():
    source = requests.get('https://everestm.com.ua/stoli/kompjuternie-stoli/',headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')

    cards = []
    size_cards = []
    img_cards = []
    color_cards = []
    link_list = []
    prod_id = 0
    pag_list = ['https://everestm.com.ua/stoli/kompjuternie-stoli/', ]

    # Забераємо пагенацію
    try:
        item = soup.find('ul', class_='pagination').find_all('a')
        for i in item:
            if i.get('href') not in pag_list:
                print(i.get('href'))
                pag_list.append(i.get('href'))
        time.sleep(2)
    except:
        pass

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        item = soup.find('div', class_='products-block').find_all('a', class_='product-name')
        for i in item:
            if i.get('href') not in link_list:
                link_list.append(i.get('href'))
                print(len(link_list), i.get('href'))

        time.sleep(2)

    # Забераємо інформацію про товар

    for i in link_list:
        print('----------------------------------')
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
    

        name = soup.find('h1').getText()
        desc_text = soup.find('div', class_='nav-desc').getText()
        char = soup.find('div', class_='nav-characteristic').getText()
        prop_list = soup.find('div', class_='portrait').find('div', class_='col-sm-6').find_all('tr')
        size = ['', '']

        desc_text = desc_text + char

        try:
            price = soup.find('span', class_='autocalc-product-special').get_text()

        except:
            price = soup.find('span', class_='autocalc-product-price').get_text()


        print(name)
        print(price)

        for i in prop_list:
            val = i.get_text()

            w_match = re.search('Ширина', val)
            d_match = re.search('Глибина', val)

            if w_match:
                size[0] = num_check(val)

            if d_match:
                size[1] = num_check(val)

        print(size)

        cards.append({
            'id': prod_id,
            'name': name,
            'des': desc_text,
        })

        size_cards.append({
            'id': prod_id,
            'w': size[0],
            'd': size[1],
            'price': price,
        })

        try:
            img = soup.find('a', class_='thumbnail').get('href')
            print(img)
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name.replace('%', ''), "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': prod_id,
                'img': img_name.replace('%', '')
            })
        except:
            pass


        prod_id += 1


    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Pcstoly.objects.filter(pars_name = product['name']):
                print('old', Pcstoly.objects.get(pars_name=product['name']).name)
                '''model_id =  Pcstoly.objects.get(pars_name=product['name']).id
                size_id = Pcstoly_size.objects.filter(pcstoly_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.filter(id = size_id).update(price=s['price'])'''
                
            #Додаємо новий товар
            else:
                new_id = Pcstoly.objects.all().order_by('-id')[0].id + 1

                Pcstoly.objects.get_or_create(id=new_id, manufacturer_id=9, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.get_or_create(pcstoly_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Pcstoly_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_pcstoly_lion():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=7).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[2]
    cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)
        match = re.search('Столи геймерські', cell)

        if match:
            while True:
                r += 1
                name = str(sheet[r][1].value)

                match = re.search("Столи комп'ютерні, направляючі телескопічні", name)

                if match:
                    break

                elif name != 'None':

                    try:
                        name = " ".join(name.split())
                        size = str(sheet[r][4].value)
                        size = size.split('х')
                        try:
                            price = round(int(sheet[r][5].value))
                        except:
                            price = round(int(sheet[r][8].value))


                        prod_id += 1

                        cards.append({
                            'id': prod_id,
                            'name': name,
                            'des': '',
                        })

                        size_cards.append({
                            'id': prod_id,
                            'w': size[0],
                            'd': size[1],
                            'price': price,
                        })

                        print(name.replace('  ', ''), size, price)
                    except Exception as ex:
                        print(ex)
                        pass

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)
        match = re.search("Столи комп'ютерні, направляючі телескопічні", cell)

        if match:
            while True:
                r += 1
                name = str(sheet[r][1].value)

                match = re.search("1 категорія ЛДСП", name)

                if match:
                    break

                elif name != 'None':

                    try:
                        name = " ".join(name.split())
                        size = str(sheet[r][4].value)
                        size = size.split('х')
                        try:
                            price = round(int(sheet[r][5].value))
                        except:
                            price = round(int(sheet[r][8].value))

                        prod_id += 1

                        cards.append({
                            'id': prod_id,
                            'name': name,
                            'des': '',
                        })

                        size_cards.append({
                            'id': prod_id,
                            'w': size[0],
                            'd': size[1],
                            'price': price,
                        })

                        print(name.replace('  ', ''), size, price)
                    except Exception as ex:
                        print(ex)
                        pass

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Pcstoly.objects.filter(pars_name = product['name']):
                print('old', Pcstoly.objects.get(pars_name=product['name']).name)
                model_id =  Pcstoly.objects.get(pars_name=product['name']).id
                size_id = Pcstoly_size.objects.filter(pcstoly_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Pcstoly.objects.all().order_by('-id')[0].id + 1

                Pcstoly.objects.get_or_create(id=new_id, manufacturer_id=6, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.get_or_create(pcstoly_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_pcstoly_svit_mebliv():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=12).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][0].value)
        match = re.search('Стіл письмовий', cell)

        if match:
            try:
                name = str(sheet[r][0].value)
                price = round(int(sheet[r][1].value) * 1.35)

                cards.append({
                    'id': prod_id,
                    'name': name,
                    'des': '',
                })

                size_cards.append({
                    'id': prod_id,
                    'w': '',
                    'd': '',
                    'price': price,
                })

                print(prod_id, name, price)
                prod_id += 1

            except:
                pass

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Pcstoly.objects.filter(pars_name = product['name']):
                print('old', Pcstoly.objects.get(pars_name=product['name']).name)
                model_id =  Pcstoly.objects.get(pars_name=product['name']).id
                size_id = Pcstoly_size.objects.filter(pcstoly_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Pcstoly.objects.all().order_by('-id')[0].id + 1

                Pcstoly.objects.get_or_create(id=new_id, manufacturer_id=2, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.get_or_create(pcstoly_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_pcstoly_neman():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=9).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    size_cards = []
    img_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        size = []
        cell = str(sheet[r][3].value)
        match = re.search('письмовий', cell)
        match2 = re.search('Письмовий', cell)

        if match or match2:
            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][5].value)
            des = str(sheet[r][39].value)
            name_ru = str(sheet[r][6].value)
            des_ru = str(sheet[r][40].value)
            size = str(sheet[r][70].value).split('x')
            img_list = str(sheet[r][15].value).split(';')
            img_list = [line.rstrip() for line in img_list]
            price = str(sheet[r][9].value)[:-3]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'des': des,
                'name_ru': name_ru,
                'des_ru': des_ru,
            })

            print(prod_id, name)
            print(size)

            print(price)
            
            try:
                size_cards.append({
                    'id': prod_id,
                    'option': '',
                    'gear': '',
                    'w': size[1],
                    'h': '',
                    'd': size[0],
                    'price': price,
                })
            except:
                size_cards.append({
                    'id': prod_id,
                    'option': '',
                    'gear': '',
                    'w': size[0],
                    'h': '',
                    'd': '',
                    'price': price,
                })

            for i in img_list:
                try:
                    img = i
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_name = str(img_name).replace('%', '')
                    img_bytes = requests.get(img, headers=HEADERS).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })
                    print(img_name)
                except:
                    pass
            
            print('-------------------------------')

    #Додаємо записи в базу
    new_db = []
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Pcstoly.objects.filter(pars_name = product['prom']):
                print('old', Pcstoly.objects.get(pars_name=product['prom']).name)
                model_id =  Pcstoly.objects.get(pars_name=product['prom']).id
                size_id = Pcstoly_size.objects.filter(pcstoly_id=model_id).values()[0]['id']
                new_db.append(product['prom'])

                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Pcstoly.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Pcstoly.objects.get_or_create(id=new_id, manufacturer_id=10, name=product['name'], name_ru=product['name_ru'],\
                    pars_name=product['prom'], description=product['des'], description_ru=product['des_ru'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.get_or_create(pcstoly_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Pcstoly_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for m in Pcstoly.objects.filter(manufacturer_id=10):
        print(m)
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def get_pcstoly_mixmebli():
    img_cards = []
    cards = []
    size_cards = []
    url = 'https://baustoff.com.ua/data/yml_mixmebli.xml'
    response = requests.get(url)
    categoryId = ['101',]


    print('response ', response.status_code)

    if response.status_code == 200:
        root = ET.fromstring(response.content)

        for offer in root.find('.//shop/offers').iter('offer'):
            if offer.find('categoryId').text in categoryId and offer.attrib.get('available') == 'true':
                
                offer_id = offer.get('id')
                print(f"Offer ID: {offer_id}")

                try:
                    description = offer.find('description').text
                except:
                    description = ''
                    
                params_html = "<table border='1'>"

                width = ''
                depth = ''
                height =''

                for param in offer.iter('param'):
                    if param.attrib.get('name') == 'Ширина':
                        width = param.text

                    if param.attrib.get('name') == 'Глибина':
                        depth = param.text
                    
                    if param.attrib.get('name') == 'Висота':
                        height = param.text

                    params_html += f"<tr><td>{param.attrib.get('name')}</td><td>{param.text}</td></tr>"
                
                params_html += "</table><br>"


                complect = False
                if offer.find('categoryId').text == '102' or offer.find('categoryId').text == '107':
                    complect = True


                cards.append({
                    'id': offer_id,
                    'prom':offer_id,
                    'name': offer.find('name').text,
                    'des': description + params_html,
                    'complect':complect
                })

                size_cards.append({
                    'id': offer_id,
                    'w':width,
                    'd':depth,
                    'price':str(offer.find('price').text).replace('.00', '')
                })

                for i in offer.findall('picture'):
                    img = str(i.text)
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': offer_id,
                        'img': img_name
                    })
                
                print('------------------------------------')

    #Додаємо записи в базу
    new_db = []
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Pcstoly.objects.filter(pars_name = product['prom']):
                print('old', Pcstoly.objects.get(pars_name=product['prom']).name)
                model_id =  Pcstoly.objects.get(pars_name=product['prom']).id
                size_id = Pcstoly_size.objects.filter(pcstoly_id=model_id).values()[0]['id']
                new_db.append(product['prom'])

                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Pcstoly.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Pcstoly.objects.get_or_create(id=new_id, manufacturer_id=27, name=product['name'],\
                    pars_name=product['prom'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.get_or_create(pcstoly_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Pcstoly_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for m in Pcstoly.objects.filter(manufacturer_id=27):
        print(m)
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def get_pcstoly_comfortmebli():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=24).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    img_cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][8].value).split(';')
            img_list.insert(0, str(sheet[r][6].value)) 

            des = str(sheet[r][9].value)


            cards.append({
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
            })

            size_cards.append({
                'id': prod_id,
                'w':'',
                'h':'',
                'd':'',
                'option':'',
                'gear':'',
                'price':price
            })

            print(prom, '-->', name, '->', price)

            for img in img_list:
                print(f"img---{img}---///")
                
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })

                except Exception as ex:
                    print(f'///-------{ex}---------///')
            
            print('-------------------------')
    
    #Додаємо записи в базу
    new_db = []
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Pcstoly.objects.filter(pars_name = product['prom']):
                print('old', Pcstoly.objects.get(pars_name=product['prom']).name)
                model_id =  Pcstoly.objects.get(pars_name=product['prom']).id
                size_id = Pcstoly_size.objects.filter(pcstoly_id=model_id).values()[0]['id']
                new_db.append(product['prom'])

                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Pcstoly.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Pcstoly.objects.get_or_create(id=new_id, manufacturer_id=1, name=product['name'],\
                    pars_name=product['prom'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.get_or_create(pcstoly_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Pcstoly_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for m in Pcstoly.objects.filter(manufacturer_id=1):
        print(m)
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def get_pcstoly_matrolux():
    print('test')

    img_cards = []
    cards = []
    size_cards = []

    print('Start ...')
    req = requests.get('https://matroluxe.ua/index.php?route=extension/feed/ocext_feed_generator_google&token=4171&categoryview=1', headers=HEADERS)  
    src = req.text
    print('Get src ---///')
    soup = BeautifulSoup(src, 'xml')
    item = soup.find_all('entry')

    for i in item:
        try: 
            title = i.find('title').get_text() if i.find('title') else None
            product_type = i.find('product_type').get_text() if i.find('product_type') else None
            product_id = i.find('id').get_text() if i.find('id') else None
            price = str(i.find('price').get_text()).replace('.00 UAH', '') if i.find('price') else None
            description = str(i.find('description').get_text()).replace('характеристики новых шкафов: ', '') if i.find('description') else ' '
            additional_image_link = i.find('additional_image_link').get_text() if i.find('additional_image_link') else None
            image_link = i.find('image_link').get_text() if i.find('image_link') else None
            img_list = [additional_image_link, image_link]

            sale = False
            old_price = None

            if i.find('sale_price'):
                old_price = price
                sale = True
                price = str(i.find('sale_price').get_text()).replace('.00 UAH', '') if i.find('sale_price') else None
            
            
            
            if  product_type == "Корпусні меблі":
                if len(str(product_id).split('-')) == 1:
                
                    update = False

                    for cat_num in range(50):
                        if cat_num == 0:
                            cat = i.find(f'category').get_text() if i.find(f'category') else None
                        
                        else:
                            cat = i.find(f'category{cat_num}').get_text() if i.find(f'category{cat_num}') else None


                        if cat == "Комп'ютерні " or cat == "Офісні" or cat == "Пісьмові ":
                            update = True


                    if update:

                        print(title)
                        print('-'*50)

                        cards.append({
                            'prom':product_id,
                            'id': product_id,
                            'name': title,
                            'des':description,
                            'sale':sale
                        })

                        size_cards.append({
                            'id': product_id,
                            'old_price':old_price,
                            'price':price,
                            
                        })

                        
                        for img in img_list:
                            print(f"img---{img}---///")
                            try:
                                img_name = img.split('/')[len(img.split('/')) - 1]
                                img_bytes = requests.get(img, headers=HEADERS).content
                                with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                                    f.write(img_bytes)

                                img_cards.append({
                                    'id': product_id,
                                    'img': img_name
                                })

                            except Exception as ex:
                                print(f'///-------{ex}---------///')
                        
                        

                    

        except Exception as ex:
            print('-'*50)
            print(ex)
            print('-'*50)

    
    
    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Pcstoly.objects.filter(pars_name = product['prom']):
                print('old', Pcstoly.objects.get(pars_name=product['prom']).name)
                Pcstoly.objects.filter(pars_name = product['prom']).update(sale=product['sale'])
                model_id =  Pcstoly.objects.get(pars_name=product['prom']).id
                size_id = Pcstoly_size.objects.filter(pcstoly_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.filter(id = size_id).update(price=s['price'], old_price=s['old_price'])
                
            #Додаємо новий товар
            else:
                new_id = Pcstoly.objects.all().order_by('-id')[0].id + 1

                Pcstoly.objects.get_or_create(id=new_id, manufacturer_id=3, name=product['name'],\
                    pars_name=product['prom'], description=product['des'], sale=product['sale'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Pcstoly_size.objects.get_or_create(pcstoly_id=new_id,\
                            price=s['price'], old_price=s['old_price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Pcstoly_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')