import requests
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
import re
import openpyxl
import json
from shafi.models import Shafi, Shafi_img, Product_shafi
from pars.models import File


with open('/home/ay507291/notacomforta.pl.ua/www/pars//src.json', 'r') as f:
    file = json.load(f)

path_dirver = file['src'][0]['path_dirver']
accept = file['src'][1]['accept']
user_agent = file['src'][2]['user-agent']

HEADERS = {
    'accept': accept,
    'user-agent': user_agent
}


def num_check(text):
    num = ''
    for t in str(text):
        if t.isdigit():
            num += t
    return num


def get_prod_komfortmebli():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=18).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    img_cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][20].value).split(';')
            img_list.insert(1, str(sheet[r][18].value)) 

            
            des = (
                        f"<ul>\n"
                        f"  <li>Вартість шафи по розмірам: <br> {sheet[r][4].value}</li>\n"
                        f"  <li>Тип шафи: {sheet[r][5].value}</li>\n"
                        f"  <li>Форма шафи-купе: {sheet[r][6].value}</li>\n"
                        f"  <li>Кількість дверей, шт: {sheet[r][10].value}</li>\n"
                        f"  <li>Тип фасаду: {sheet[r][11].value}</li>\n"
                        f"  <li>Колір корпусу: {sheet[r][12].value}</li>\n"
                        f"  <li>Внутрішнє наповнення: {sheet[r][13].value}</li>\n"
                        f"  <li>Форма AL профілю: {sheet[r][14].value}</li>\n"
                        f"  <li>Колір AL профілю: {sheet[r][15].value}</li>\n"
                        f"  <li>Серія: {sheet[r][16].value}</li>\n"
                        f"  <li>Стиль: {sheet[r][17].value}</li>\n"
                        f"</ul>\n" 
                        f"<p>{sheet[r][21].value}</p>\n"
                    )


            cards.append({
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
            })

            size_cards.append({
                'id': prod_id,
                'w':sheet[r][7].value,
                'h':sheet[r][8].value,
                'd':sheet[r][9].value,
                'price':price
            })
    
            for img in img_list:
                print(f"img---{img}---///")
                try:
                    img_name = img.split('/')[len(img.split('/')) - 2] + img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })

                except Exception as ex:
                    print(f'///-------{ex}---------///')


            
            print(img_list)
            print('-------------------------')
            
            print(prom, '-->', name, '->', price)


    #Додаємо записи в базу
    new_db=[]
    for product in cards:
        print('------------------------------------------------')
        try:
            #Оновлюємо ціну
            if Shafi.objects.filter(pars_name = product['prom']):
                print('old', Shafi.objects.get(pars_name=product['prom']).shafi_name)
                model_id =  Shafi.objects.get(pars_name=product['prom']).id
                new_db.append(product['prom'])
                size_id = Product_shafi.objects.filter(shafi_name_id=model_id).values()[0]['id']
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.filter(id = size_id).update(price=s['price'])
                        break
                
            #Додаємо новий товар
            else:
                new_id = Shafi.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Shafi.objects.get_or_create(id=new_id, manufacturer_id=1, shafi_name=product['name'],\
                    pars_name=product['prom'], description=product['des'], categori='KYPE')
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.get_or_create(shafi_name_id=new_id,\
                             width=s['w'], height=s['h'], depth=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Shafi_img.objects.get_or_create(shafi_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for m in Shafi.objects.filter(manufacturer_id=1):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()
    


def get_link_matrolux():
    img_cards = []
    cards = []
    size_cards = []

    print('Start ...')
    req = requests.get('https://matroluxe.ua/index.php?route=extension/feed/ocext_feed_generator_google&token=4171&categoryview=1', headers=HEADERS)  
    src = req.text
    print('Get src ---///')
    soup = BeautifulSoup(src, 'xml')
    item = soup.find_all('entry')

    for i in item:
        try: 
            title = i.find('title').get_text() if i.find('title') else None
            product_type = i.find('product_type').get_text() if i.find('product_type') else None
            product_id = i.find('id').get_text() if i.find('id') else None
            price = str(i.find('price').get_text()).replace('.00 UAH', '') if i.find('price') else None
            description = str(i.find('description').get_text()).replace('характеристики новых шкафов: ', '') if i.find('description') else ' '
            additional_image_link = i.find('additional_image_link').get_text() if i.find('additional_image_link') else None
            image_link = i.find('image_link').get_text() if i.find('image_link') else None
            img_list = [additional_image_link, image_link]

            if i.find('sale_price'):
                price = str(i.find('sale_price').get_text()).replace('.00 UAH', '') if i.find('sale_price') else None
            
            
            
            if  product_type == "Шафи" or product_type == "Шафи-купе":
                if len(str(product_id).split('-')) == 1:
                    print(title)
                    print('-'*50)
                    categori='KYPE'
                    for cat_num in range(50):
                        if cat_num == 0:
                            cat = i.find(f'category').get_text() if i.find(f'category') else None
                        
                        else:
                            cat = i.find(f'category{cat_num}').get_text() if i.find(f'category{cat_num}') else None

                        if cat == 'Шафи-купе':
                            categori='KYPE'
                            print('KYPE')
                            break

                        if cat == 'Розпашні шафи':
                            categori='SHAFI'
                            print('SHAFI')
                            break

                        if cat == 'Шафа-пенал':
                            categori='PENAL'
                            print('PENAL')
                            break


                    cards.append({
                        'prom':product_id,
                        'id': product_id,
                        'name': title,
                        'des':description,
                        'categori':categori
                    })

                    size_cards.append({
                        'id': product_id,
                        'w':'',
                        'h':'',
                        'd':'',
                        'price':price
                    })


                    for img in img_list:
                        print(f"img---{img}---///")
                        try:
                            img_name = img.split('/')[len(img.split('/')) - 1]
                            img_bytes = requests.get(img, headers=HEADERS).content
                            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                                f.write(img_bytes)

                            img_cards.append({
                                'id': product_id,
                                'img': img_name
                            })

                        except Exception as ex:
                            print(f'///-------{ex}---------///')

                    

        except Exception as ex:
            print('-'*50)
            print(ex)
            print('-'*50)

    

    #Додаємо записи в базу
    new_db=[]
    for product in cards:
        print('------------------------------------------------')
        try:
            #Оновлюємо ціну
            if Shafi.objects.filter(pars_name = product['prom']):
                print('old', Shafi.objects.get(pars_name=product['prom']).shafi_name)
                model_id =  Shafi.objects.get(pars_name=product['prom']).id
                new_db.append(product['prom'])
                size_id = Product_shafi.objects.filter(shafi_name_id=model_id).values()[0]['id']
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.filter(id = size_id).update(price=s['price'])
                        break
                
            #Додаємо новий товар
            else:
                new_id = Shafi.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Shafi.objects.get_or_create(id=new_id, manufacturer_id=3, shafi_name=product['name'],\
                    pars_name=product['prom'], description=product['des'], categori=product['categori'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.get_or_create(shafi_name_id=new_id,\
                             width=s['w'], height=s['h'], depth=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Shafi_img.objects.get_or_create(shafi_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for m in Shafi.objects.filter(manufacturer_id=3):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def get_products_fenix():
    print(File.objects.get(id=3).file_up.url)
    id = 0
    cards = []
    size_cards = []
    img_cards = []
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=3).file_up.url

    book = openpyxl.load_workbook(filename=path)
    sheet_name ={'name': "Стандарт", 'link': 'http://f-mebel.ua/shkafyi-kupe/seriya-standart/'},\
                {'name': "Комфорт", 'link': 'http://f-mebel.ua/shkafyi-kupe/seriya-komfort/'},\
                {'name': "Ультра", 'link': 'http://f-mebel.ua/shkafyi-kupe/seriya-ultra/'},\
                {'name': "Люкс", 'link': 'http://f-mebel.ua/shkafyi-kupe/seriya-lyuks/'}

    for n in sheet_name:
        sheet = book[n['name']]
        row = 0
        name = sheet[1][3].value

        while True:
            row += 1
            try:
                int(sheet[row][0].value)
                id += 1
                price = round(sheet[row][5].value * sheet[row][3].value + sheet[row][4].value)
                width = sheet[row][0].value
                height = sheet[row][2].value
                depth = sheet[row][1].value
                name_product = name + " " + str(sheet[row][0].value) + '*' + str(sheet[row][2].value)+ '*' + str(sheet[row][1].value)

                cards.append({
                    'id': id,
                    'name': name_product
                })

                size_cards.append({
                    'id': id,
                    'w': num_check(width),
                    'h': num_check(height),
                    'd': num_check(depth),
                    'price': num_check(price)
                })

                print(f"{id}--->{name_product}")
            except:
                if sheet[row][0].value == "Дополнительная  комплектация":
                    break
                else:
                    pass

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Shafi.objects.filter(pars_name = product['name']):
                print('old', Shafi.objects.get(pars_name=product['name']).shafi_name)
                model_id =  Shafi.objects.get(pars_name=product['name']).id
                size_id = Product_shafi.objects.filter(shafi_name_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.filter(id = size_id).update(price=s['price'])
                        break
                
            #Додаємо новий товар
            else:
                new_id = Shafi.objects.all().order_by('-id')[0].id + 1

                Shafi.objects.get_or_create(id=new_id, manufacturer_id=4, shafi_name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.get_or_create(shafi_name_id=new_id,\
                             width=s['w'], height=s['h'], depth=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Shafi_img.objects.get_or_create(shafi_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_shafi_neman():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=9).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet_1 = book.worksheets[0]

    prod_id = 0
    cards = []
    size_cards = []
    img_cards = []

    for r in range(sheet_1.max_row):
        r += 1

        cell = str(sheet_1[r][3].value)
        match = re.search('Шафа', cell)
        match2 = re.search('шафа', cell)
        match3 = re.search('Стелаж', cell)
        match4 = re.search('стелаж', cell)
        match5 = re.search('Шкаф', cell)

        if match or match2 or match3 or match4 or match5:
            prod_id += 1
            prom = str(sheet_1[r][0].value)
            name = str(sheet_1[r][3].value)
            des = str(sheet_1[r][37].value)
            name_ru = str(sheet_1[r][5].value)
            des_ru = str(sheet_1[r][38].value)
            price = str(sheet_1[r][9].value)[:-3]
            img_list = str(sheet_1[r][15].value).split(';')
            img_list = [line.rstrip() for line in img_list]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'name_ru': name_ru,
                'des':des,
                'des_ru':des_ru,
                'type':'SHAFI'
            })

            size_cards.append({
                'id': prod_id,
                'w':'',
                'h':'',
                'd':'',
                'price':price
            })

            for i in img_list:
                try:
                    img = i
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_name = str(img_name).replace('%', '')
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })
                except:
                    pass

            print('--------------------------')
            print(prom)
            print(name, price)


    for r in range(sheet_1.max_row):
        r += 1

        cell = str(sheet_1[r][5].value)
        match = re.search('Пенал', cell)

        if match:
            prod_id += 1
            prom = str(sheet_1[r][0].value)
            name = str(sheet_1[r][5].value)
            des = str(sheet_1[r][37].value)
            name_ru = str(sheet_1[r][6].value)
            des_ru = str(sheet_1[r][38].value)
            price = str(sheet_1[r][9].value)[:-3]
            img_list = str(sheet_1[r][15].value).split(';')
            img_list = [line.rstrip() for line in img_list]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'name_ru': name_ru,
                'des':des,
                'des_ru':des_ru,
                'type':'PENAL'
            })

            size_cards.append({
                'id': prod_id,
                'w':'',
                'h':'',
                'd':'',
                'price':price
            })

            for i in img_list:
                try:
                    img = i
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_name = str(img_name).replace('%', '')
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })
                except:
                    pass

            print('--------------------------')
            print(prom)
            print(name, price)


    #Додаємо записи в базу
    new_db=[]
    for product in cards:
        print('------------------------------------------------')
        try:
            #Оновлюємо ціну
            if Shafi.objects.filter(pars_name = product['prom']).filter(manufacturer_id=10):
                print('old', Shafi.objects.get(pars_name=product['prom']).shafi_name)
                model_id =  Shafi.objects.get(pars_name=product['prom']).id
                new_db.append(product['prom'])
                size_id = Product_shafi.objects.filter(shafi_name_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.filter(id = size_id).update(price=s['price'])
                        break
                
            #Додаємо новий товар
            else:
                new_id = Shafi.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Shafi.objects.get_or_create(id=new_id, manufacturer_id=10,\
                    shafi_name=product['name'], shafi_name_ru=product['name_ru'],\
                        description=product['des'], description_ru=product['des_ru'],\
                            pars_name=product['prom'], categori=product['type'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.get_or_create(shafi_name_id=new_id,\
                             width=s['w'], height=s['h'], depth=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Shafi_img.objects.get_or_create(shafi_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for m in Shafi.objects.filter(manufacturer_id=10):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def get_shafi_mixmebli():
    img_cards = []
    cards = []
    size_cards = []
    url = 'https://baustoff.com.ua/data/yml_mixmebli.xml'
    response = requests.get(url)
    categoryId = ['73',]


    print('response ', response.status_code)

    if response.status_code == 200:
        root = ET.fromstring(response.content)

        for offer in root.find('.//shop/offers').iter('offer'):
            if offer.find('categoryId').text in categoryId and offer.attrib.get('available') == 'true':
                offer_id = offer.get('id')
                print(f"Offer ID: {offer_id}")

                description = offer.find('description').text
                params_html = "<table border='1'>"

                width = ''
                depth = ''
                height =''

                for param in offer.iter('param'):
                    if param.attrib.get('name') == 'Ширина':
                        width = param.text

                    if param.attrib.get('name') == 'Глибина':
                        depth = param.text
                    
                    if param.attrib.get('name') == 'Висота':
                        height = param.text

                    params_html += f"<tr><td>{param.attrib.get('name')}</td><td>{param.text}</td></tr>"
                
                params_html += "</table><br>"

                cards.append({
                    'id': offer_id,
                    'prom':offer_id,
                    'name': offer.find('name').text,
                    'des': description + params_html,
                })

                size_cards.append({
                    'id': offer_id,
                    'w':width,
                    'h':height,
                    'd':depth,
                    'price':str(offer.find('price').text).replace('.00', '')
                })

                for i in offer.findall('picture'):
                    img = str(i.text)
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': offer_id,
                        'img': img_name
                    })
                
                print('------------------------------------')

    #Додаємо записи в базу
    new_db=[]
    for product in cards:
        print('------------------------------------------------')
        try:
            #Оновлюємо ціну
            if Shafi.objects.filter(pars_name = product['prom']):
                print('old', Shafi.objects.get(pars_name=product['prom']).shafi_name)
                model_id =  Shafi.objects.get(pars_name=product['prom']).id
                new_db.append(product['prom'])
                size_id = Product_shafi.objects.filter(shafi_name_id=model_id).values()[0]['id']
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.filter(id = size_id).update(price=s['price'])
                        break
                
            #Додаємо новий товар
            else:
                new_id = Shafi.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Shafi.objects.get_or_create(id=new_id, manufacturer_id=27, shafi_name=product['name'],\
                    pars_name=product['prom'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Product_shafi.objects.get_or_create(shafi_name_id=new_id,\
                             width=s['w'], height=s['h'], depth=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Shafi_img.objects.get_or_create(shafi_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for m in Shafi.objects.filter(manufacturer_id=27):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()