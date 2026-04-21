import requests
import xml.etree.ElementTree as ET
import openpyxl
from bs4 import BeautifulSoup
import time
import re
import os
from googletrans import Translator
import json
from main.models import Seria, Stinka, Stinka_img
from pars.models import File


with open('/home/ay507291/notacomforta.pl.ua/www/pars/src.json', 'r') as f:
    file = json.load(f)

path_dirver = file['src'][0]['path_dirver']
accept = file['src'][1]['accept']
user_agent = file['src'][2]['user-agent']

HEADERS = {
    'accept': accept,
    'user-agent': user_agent
}


def num_check(text):
    num = ''
    for t in text:
        if t.isdigit():
            num += t
    return num


def trans(text):
    translator = Translator().translate(text=text, src='ru', dest='uk')
    return translator.text


def test(modul_cards, product_cards):
    for m in modul_cards:
        print(m['id'], m['name'])
        for p in product_cards:
            if m['id'] == p['modul_id']:
                print(f"modul: {p['modul_id']} id: {p['id']} name: {p['name']} price: {p['price']}")


def UpdateData(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['name']):
            modul_id = Seria.objects.get(pars_name = m['name']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['name'] == el.name:
                        new_db.append(p['name'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(price=p['price'], pars_category='vitalni')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['name'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['name'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['name'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
        
    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(vitalni=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def UpdateDataBMK(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']):
            modul_id = Seria.objects.get(pars_name = m['prom']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['prom'] == el.pars_name:
                        new_db.append(p['prom'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(price=p['price'], pars_category='vitalni')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
        
    '''#Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(vitalni=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()'''


def UpdateDataSvitMebliv(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['prom']).id

            print('-->', m['name'])
            for p in product_cards:
                if p['modul_id'] == m['id']:
                    if Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id):
                        new_db.append(p['prom'])
                        print('old:', p['name'])
                        Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id).update(price=p['price'], pars_category='vitalni')
                    else:
                        print('none:', p['name'])
                        new_db.append(p['prom'])
                        Stinka.objects.create(seria_id=modul_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
    
    #Видаляємо товар
    '''for m in Stinka.objects.filter(manufacturer_id=manu).filter(vitalni=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()'''


def UpdateDataLion(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['prom']).id

            print('-->', m['name'])
            for p in product_cards:
                if p['modul_id'] == m['id']:
                    if Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id):
                        new_db.append(p['prom'])
                        print('old:', p['name'])
                        Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id).update(pars_category='vitalni')
                    else:
                        print('none:', p['name'])
                        new_db.append(p['prom'])
                        Stinka.objects.create(seria_id=modul_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
    
    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(vitalni=True):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def UpdateDataEverest(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['name']):
            modul_id = Seria.objects.get(pars_name = m['name']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['name'] == el.name:
                        new_db.append(p['name'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(pars_category='vitalni')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['name'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['name'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['name'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
        
    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(vitalni=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def UpdateDataMixMebli(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['prom']).id

            print('-->', m['name'])
            for p in product_cards:
                if p['modul_id'] == m['id']:
                    if Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id):
                        new_db.append(p['prom'])
                        print('old:', p['name'])
                        Stinka.objects.filter(pars_name=p['prom'])\
                            .filter(seria_id=modul_id)\
                                .update(pars_category='vitalni',\
                                        description=p['des'],\
                                            price=p['price'],)
                    else:
                        print('none:', p['name'])
                        new_db.append(p['prom'])
                        Stinka.objects.create(seria_id=modul_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'], width=p['w'], depth=p['d'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
    
    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(vitalni=True):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def UpdateDataKM(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']):
            modul_id = Seria.objects.get(pars_name = m['prom']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['prom'] == el.pars_name:
                        new_db.append(p['prom'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(price=p['price'], pars_category='vitalni')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          vitalni=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], width=p['width'], depth=p['depth'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
        
    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(vitalni=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def get_vitalni_products_comfortmebli():

    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=20).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    
    modul_cards = []
    product_cards = []
    img_cards = []
    modul_id = 0
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            modul_id += 1

            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            width = str(sheet[r][5].value)
            depth = str(sheet[r][6].value)

            img_list =str(sheet[r][9].value).split(';')
            img_list.insert(0, str(sheet[r][7].value)) 

            des = str(sheet[r][10].value)


            product_cards.append({
                'modul_id':modul_id,
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
                'width':width,
                'depth':depth,
                'price':price
            })

            modul_cards.append({
                'prom':prom,
                'id':modul_id,
                'name':name,
            })

            print(prom, '-->', name, '->', price)

            for img in img_list:
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })

                except Exception as ex:
                    print(f'///-------{ex}---------///')

            print('-------------------------')


    test(modul_cards, product_cards)
    UpdateDataKM(modul_cards, product_cards, img_cards, manu=1)


def get_vitalni_products_lion():
    modul_cards = []
    product_cards = []
    img_cards = []
    modul_id = 0
    prod_id = 0
    link_list = ['https://lion-mebli.com.ua/living-room/',]


    source = requests.get('https://lion-mebli.com.ua/living-room/',headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')

    #Пагінація
    pag_list = soup.find('ul', class_='pagination').find_all('a')
    for i in pag_list:
        
        link = i.get('href')

        if link not in link_list:
            link_list.append(link)



    #Збираємо товар
    prod_list=[]
    for url in link_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        items = soup.find_all('div', class_='category_product')

        for i in items:
            link = i.find('a').get('href')
            if link not in prod_list:
                prod_list.append(link)


    #Збираємо інформацію про товар
    for p in prod_list:
        source = requests.get(p,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        prod_id += 1
        prom = p
        name = soup.find('h1').getText()
        des = str(soup.find('div', class_='product_description'))
        img_list = soup.find('div', class_='product_small_nav').find_all('img')

        print(f"///---{name}---///")

        modul_cards.append({
            'id':prod_id,
            'prom':prom,
            'name':name,
        })

        product_cards.append({
            'modul_id':prod_id,
            'id':prod_id,
            'prom':prom,
            'name':name,
            'des':des,
            'price':''
        })

        for i in img_list:
            for i in img_list:
                try:
                    img = i.get('src')

                    img_cards.append({
                        'id': prod_id,
                        'img': img
                    })
                except:
                    pass
        
        time.sleep(1)

    #test(modul_cards, product_cards)
    UpdateDataLion(modul_cards, product_cards, img_cards, manu=6)


def get_vitalni_products_svitmebliv():
    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0

    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=6).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    for i in range(sheet.max_row):
        i += 1
        try:
            name = sheet[i][0].value
            price = int(sheet[i][4].value)

            if name == 'ТВ-тумба Соло':
                pass
            else:
                prom = name.replace(' ', '').lower()
                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })

                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'w': '',
                    'des': '',
                    'price': price
                })

                modul_id += 1
                prod_id += 1
        except:
            pass

    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=6).file_up.url
    print("модульна система", path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[1]
    for c in range(sheet.max_column):
        for r in range(sheet.max_row):
            r += 1
            
            cell = str(sheet[r][c].value)
            match = re.search("модульна система", cell)
            if match:
                name = cell[16:]
                column = c + 4
                row = r
                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })

                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'w': '',
                    'des': '',
                    'price': ''
                })

                print('-->',name)

                while True:
                    row += 1

                    try:
                        
                        
                        name = str(sheet[row][c].value)
                        print(name)
                        price = int(sheet[row][column].value)
                        print(price)

                        prod_id += 1
                        prom = name.replace(' ', '').lower()

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': '',
                            'des': '',
                            'price': price
                        })

                        
                        

                    except Exception as ex:
                        print("---///----", ex)
                        break

                modul_id += 1
                prod_id += 1
            else:
                pass

    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=6).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[2]
    for r in range(sheet.max_row):
        r += 1
        for c in range(sheet.max_column):
            cell = str(sheet[r][c].value)
            match = re.search("модульна система", cell)
            if match != None:
                name = cell[16:]
                column = c + 3
                row = r
                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })

                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'w': '',
                    'des': '',
                    'price': ''
                })

                while True:
                    row += 1
                    try:
                        name = str(sheet[row][c].value)
                        price = int(sheet[row][column].value)
                        prod_id += 1
                        prom = name.replace(' ', '').lower()

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': '',
                            'des': '',
                            'price': price
                        })
                        
                    except:
                        break

                modul_id += 1
                prod_id += 1
            else:
                pass

    test(modul_cards, product_cards)
    UpdateDataSvitMebliv(modul_cards, product_cards, img_cards = [], manu=2)


def get_vitalni_products_BMK():
    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0
    row = 0

    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=14).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    for i in range(sheet.max_row):
        i +=1
        cell = str(sheet[i][1].value)
        equal = re.search('ВІТАЛЬНІ', cell)
        if equal:
            while True:
                i +=1
                try:
                    name = sheet[i][2].value
                    price = int(sheet[i][7].value)
                    prom = str(sheet[i][3].value)

                    modul_cards.append({
                        'id': modul_id,
                        'prom':prom,
                        'name': name
                    })

                    product_cards.append({
                        'modul_id': modul_id,
                        'id': prod_id,
                        'prom':prom,
                        'name': name,
                        'w': '',
                        'des': '',
                        'price': price
                    })

                    prod_id += 1
                    modul_id += 1
                    #print(name, prom)
                except:
                    row = i
                    break

    for r in range(sheet.max_row):
        r += row
        cell = str(sheet[r][1].value)

        if cell =='СТОЛИ':
                    break

        elif str(sheet[r+1][1].value) == '1':
            modul_id += 1
            name = cell
            prom = cell.lower()
            prom = prom.replace(' ', '')

            modul_cards.append({
                'id': modul_id,
                'prom':prom,
                'name': name
            })

            print(modul_id, name, prom)

            while True:
                r += 1
                try:
                    prod_id += 1
                    name = sheet[r][2].value
                    price = int(sheet[r][7].value)
                    prom = str(sheet[r][3].value)
                    
                    product_cards.append({
                        'modul_id': modul_id,
                        'id': prod_id,
                        'prom':prom,
                        'name': name,
                        'w': '',
                        'des': '',
                        'price': price
                    })
                    print(modul_id, name, prom, price)
                except:
                    break


    UpdateDataBMK(modul_cards, product_cards, img_cards=[], manu=7)


def get_vitalni_products_gerbor():
    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0

    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=15).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    for i in range(sheet.max_row):
        i +=1
        cell = str(sheet[i][2].value)
        equal = re.search('Вітальні', cell)
        if equal:
            while True:
                i +=1
                try:
                    name = sheet[i][2].value
                    price = int(sheet[i][7].value)
                    prom = str(sheet[i][3].value)

                    modul_cards.append({
                        'id': modul_id,
                        'prom':prom,
                        'name': name
                    })

                    product_cards.append({
                        'modul_id': modul_id,
                        'id': prod_id,
                        'prom':prom,
                        'name': name,
                        'w': '',
                        'des': '',
                        'price': price
                    })

                    prod_id += 1
                    modul_id += 1
                    print(name, prom)
                except:
                    
                    break

    UpdateDataBMK(modul_cards, product_cards, img_cards=[], manu=8)


def get_vitalni_products_everest():
    source = requests.get('https://everestm.com.ua/gostinie/gotovi-nabori-dlja-gostinoj/',headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')
    item = soup.find('ul', class_='pagination').find_all('a')
    pag_list = ['https://everestm.com.ua/gostinie/gotovi-nabori-dlja-gostinoj/', ]

    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0
    img_cards = []
    link_list = []

    # Забераємо пагенацію

    for i in item:
        if i.get('href') not in pag_list:
            pag_list.append(i.get('href'))
    time.sleep(2)

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        item = soup.find('div', class_='products-block').find_all('a', class_='product-name')
        for i in item:
            if i.get('href') not in link_list:
                link_list.append(i.get('href'))

        time.sleep(2)

    # Забераємо інформацію про товар

    for i in link_list:
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        name = soup.find('h1').get_text()
        desc_text = soup.find('div', class_='nav-desc').get_text()
        char = soup.find('div', class_='nav-characteristic').get_text()
        desc_text = desc_text + char

        try:
            price = soup.find('span', class_='autocalc-product-special').get_text()

        except:
            price = soup.find('span', class_='autocalc-product-price').get_text()

        modul_cards.append({
            'id': modul_id,
            'name': name
        })

        product_cards.append({
            'modul_id': modul_id,
            'id': prod_id,
            'name': name,
            'w': '',
            'des': desc_text,
            'price': price
        })
        
        try:
            img = soup.find('a', class_='thumbnail').get('href')
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name.replace('%', ''), "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': prod_id,
                'img': img_name.replace('%', '')
            })
        except:
            pass
        
        modul_id += 1
        prod_id += 1
    
    test(modul_cards, product_cards)
    UpdateDataEverest(modul_cards, product_cards, img_cards, manu=9)


def get_vitalni_mixmebli():
    modul_cards = []
    product_cards = []
    img_cards = []
    url = 'https://baustoff.com.ua/data/yml_mixmebli.xml'
    response = requests.get(url)
    categoryId = ['105',]

    modul_cards.append({
            'id': '105',
            'prom':'105',
            'name': 'Сканді'
        })

    print('response ', response.status_code)

    if response.status_code == 200:
        root = ET.fromstring(response.content)

        for offer in root.find('.//shop/offers').iter('offer'):
            if offer.find('categoryId').text in categoryId and offer.attrib.get('available') == 'true':
                offer_id = offer.get('id')
                print(f"Offer ID: {offer_id}")

                description = offer.find('description').text
                params_html = "<table border='1'>"

                width = ''
                depth = ''
                for param in offer.iter('param'):
                    if param.attrib.get('name') == 'Ширина':
                        width = param.text
                    if param.attrib.get('name') == 'Довжина':
                        depth = param.text

                    params_html += f"<tr><td>{param.attrib.get('name')}</td><td>{param.text}</td></tr>"
                
                params_html += "</table><br>"

                product_cards.append({
                    'modul_id': '105',
                    'id': offer_id,
                    'prom': offer_id,
                    'name': offer.find('name').text,
                    'w': width,
                    'd': depth,
                    'des': description + params_html,
                    'price': str(offer.find('price').text).replace('.00', '')
                })

                for i in offer.findall('picture'):
                    img = str(i.text)
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': offer_id,
                        'img': img_name
                    })
                
                print('------------------------------------')
    
    UpdateDataMixMebli(modul_cards, product_cards, img_cards, manu=27)

    #test(modul_cards, product_cards)
        
