import requests
import openpyxl
from bs4 import BeautifulSoup
import csv
import time
import re
import os
import json
from komodytumby.models import Komodytumby, Komodytumby_img
from pars.models import File

with open('/home/ay507291/notacomforta.pl.ua/www/pars/src.json', 'r') as f:
    file = json.load(f)

path_dirver = file['src'][0]['path_dirver']
accept = file['src'][1]['accept']
user_agent = file['src'][2]['user-agent']

HEADERS = {
    'accept': accept,
    'user-agent': user_agent
}


def num_check(text):
    num = ''
    for t in str(text):
        if t.isdigit():
            num += t
    return num


def get_komody_tumby_matrolux():
    img_cards = []
    cards = []
    size_cards = []

    print('Start ...')
    req = requests.get('https://matroluxe.ua/index.php?route=extension/feed/ocext_feed_generator_google&token=4171&categoryview=1', headers=HEADERS)  
    src = req.text
    print('Get src ---///')
    soup = BeautifulSoup(src, 'xml')
    item = soup.find_all('entry')

    for i in item:
        try: 
            title = i.find('title').get_text() if i.find('title') else None
            product_type = i.find('product_type').get_text() if i.find('product_type') else None
            product_id = i.find('id').get_text() if i.find('id') else None
            price = str(i.find('price').get_text()).replace('.00 UAH', '') if i.find('price') else None
            description = str(i.find('description').get_text()).replace('характеристики новых шкафов: ', '') if i.find('description') else ' '
            additional_image_link = i.find('additional_image_link').get_text() if i.find('additional_image_link') else None
            image_link = i.find('image_link').get_text() if i.find('image_link') else None
            img_list = [additional_image_link, image_link]

            sale = False
            old_price = None

            if i.find('sale_price'):
                old_price = price
                sale = True
                price = str(i.find('sale_price').get_text()).replace('.00 UAH', '') if i.find('sale_price') else None
            
            
            
            if  product_type == "Корпусні меблі":
                if len(str(product_id).split('-')) == 1:
                
                    update = False
                    subcat = None

                    for cat_num in range(50):
                        if cat_num == 0:
                            cat = i.find(f'category').get_text() if i.find(f'category') else None
                        
                        else:
                            cat = i.find(f'category{cat_num}').get_text() if i.find(f'category{cat_num}') else None


                        if cat == 'Комоди':
                            update = True
                            subcat = 'COMO'


                        if cat == 'Тумби':
                            update = True
                            subcat = 'COMO'


                        if cat == 'Під телевізор':
                            update = True
                            subcat = 'TV'


                        if cat == 'Тумби для взуття':
                            update = True
                            subcat = 'SHOOZ'


                        if cat == 'Приліжкові тумби':
                            update = True
                            subcat = 'FORBED'


                    if update:

                        print(title)
                        print('-'*50)

                        cards.append({
                            'prom':product_id,
                            'id': product_id,
                            'name': title,
                            'des':description,
                            'subcat':subcat,
                            'old_price':old_price,
                            'price':price,
                            'sale':sale,
                        })

                        
                        for img in img_list:
                            print(f"img---{img}---///")
                            try:
                                img_name = img.split('/')[len(img.split('/')) - 1]
                                img_bytes = requests.get(img, headers=HEADERS).content
                                with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                                    f.write(img_bytes)

                                img_cards.append({
                                    'id': product_id,
                                    'img': img_name
                                })

                            except Exception as ex:
                                print(f'///-------{ex}---------///')
                        

                    

        except Exception as ex:
            print('-'*50)
            print(ex)
            print('-'*50)

    
    
    #Додаємо записи в базу
    for product in cards:
        print('********--------------------------------------', product['name'])
        product['name']
        try:
            #Оновлюємо ціну
            if Komodytumby.objects.filter(pars_name = product['prom']):
                print('old', Komodytumby.objects.get(pars_name=product['prom']).name)
                model_id =  Komodytumby.objects.get(pars_name=product['prom']).id
                Komodytumby.objects.filter(id = model_id).update(price=product['price'],\
                                                                  sale = product['sale'], old_price=product['old_price'])
                        
                
            #Додаємо новий товар
            else:
                new_id = Komodytumby.objects.all().order_by('-id')[0].id + 1

                Komodytumby.objects.get_or_create(id=new_id, manufacturer_id=3, name=product['name'],\
                    pars_name=product['prom'], description=product['des'],\
                        price=product['price'], sale = product['sale'], old_price=product['old_price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Komodytumby_img.objects.get_or_create(key_img_id=new_id, img=i['img'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')
    


def get_komody_tumby_everest():
    
    source = requests.get('https://everestm.com.ua/vse-komodi-i-tumbi/',headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')
    item = soup.find('ul', class_='pagination').find_all('a')
    pag_list = ['https://everestm.com.ua/vse-komodi-i-tumbi/', ]

    cards = []
    img_cards = []
    color_cards = []
    link_list = []
    prod_id = 0

    # Забераємо пагенацію

    for i in item:
        if i.get('href') not in pag_list:
            print(i.get('href'))
            pag_list.append(i.get('href'))
    time.sleep(2)

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        item = soup.find('div', class_='products-block').find_all('a', class_='product-name')
        for i in item:
            if i.get('href') not in link_list:
                link_list.append(i.get('href'))
                print(len(link_list), i.get('href'))

        time.sleep(2)

    # Забераємо інформацію про товар

    for i in link_list:
        print('----------------------------------')
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        name = soup.find('h1').get_text()
        desc_text = soup.find('div', class_='nav-desc').get_text()
        char = soup.find('div', class_='nav-characteristic').get_text()
        prop_list = soup.find('div', class_='portrait').find('div', class_='col-sm-6').find_all('tr')
        size = ['', '', '']
        prom = i

        desc_text = desc_text + char

        try:
            price = soup.find('span', class_='autocalc-product-special').get_text()

        except:
            price = soup.find('span', class_='autocalc-product-price').get_text()


        print(name)
        print(price)

        for i in prop_list:
            val = i.get_text()

            w_match = re.search('Ширина', val)
            h_match = re.search('Висота', val)
            d_match = re.search('Глибина', val)

            if w_match:
                size[0] = num_check(val)

            if h_match:
                size[1] = num_check(val)

            if d_match:
                size[2] = num_check(val)
        

        cards.append({
            'id': prod_id,
            'prom': prom,
            'name': name,
            'des': desc_text,
            'w': size[0],
            'h': size[1],
            'd': size[2],
            'price': price,
        })

        try:
            img = soup.find('a', class_='thumbnail').get('href')
            print(img)
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name.replace('%', ''), "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': prod_id,
                'img': img_name.replace('%', '')
            })
        except:
            pass

        prod_id += 1

        time.sleep(1)

    #Додаємо записи в базу
    new_db = []
    for product in cards:
        
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Komodytumby.objects.filter(pars_name = product['prom']):
                print('old', Komodytumby.objects.get(pars_name=product['prom']).name)
                new_db.append(product['prom'])
                #model_id =  Komodytumby.objects.get(pars_name=product['name']).id
                #Komodytumby.objects.filter(id = model_id).update(price=product['price'])
                        
                
            #Додаємо новий товар
            else:
                new_id = Komodytumby.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Komodytumby.objects.get_or_create(id=new_id, manufacturer_id=9, name=product['name'],\
                    pars_name=product['prom'], description=product['des'],\
                        width=product['w'], deep=product['d'], price='0')

                for i in img_cards:
                    if i['id'] == product['id']:
                        Komodytumby_img.objects.get_or_create(key_img_id=new_id, img=i['img'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')
    
    #Видаляємо товар
    for m in Komodytumby.objects.filter(manufacturer_id=9):
        print(m)
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def get_komody_tumby_arbordrev():

    pag_list = ['https://arbordrev.com.ua/tumbikomodi', ]
    links_list = []

    cards = []
    img_cards = []
    color_cards = []
    link_list = []
    prod_id = 0
    size = ['', '', '']

    # Забераємо пагенацію

    for num in range(100):
        try:
            source = requests.get(pag_list[num],headers=HEADERS).text
            soup = BeautifulSoup(source, 'html.parser')
            pagination = soup.find('div', class_='wd-loop-footer products-footer').find('a').get('href')

            if pagination not in pag_list:
                pag_list.append(pagination)

        except:
            break

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        links = soup.find_all('h3', class_='wd-entities-title')

        for i in links:
            link = i.find('a').get('href')
            if link not in links_list:
                links_list.append({
                    'num': len(links_list),
                    'url': link
                })
                print(len(links_list), link)

        time.sleep(1)

    # Забераємо інформацію про товар

    for url in links_list:
        print('--------------------------------------------')
        source = requests.get(url['url'],headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        name = soup.find('h1').get_text()
        price = soup.find('div', class_='col-lg-6 col-12 col-md-6 wd-price-outside summary entry-summary').find('bdi').getText()
        desc_text = soup.find('div', class_='wc-tab-inner').get_text()
        chek_list = soup.find('table', class_='variations').find('tbody').find_all('tr')
        img_list = soup.find('div', class_='product-image-wrap').find('a').get('href')
        prop_list = soup.find('div', class_='woocommerce-Tabs-panel woocommerce-Tabs-panel--additional_information panel entry-content wc-tab').find_all('tr')
        list = []
        material = []

        print(url['num'], name, price)
        
        img_cards.append({
            'id': url['num'],
            'img': img_list
        })


        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Колір', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    color = a.find('img').get('src')
                    color_cards.append({
                        'id': url['num'],
                        'img': color
                    })

        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Порода дерева', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    active = a.get('aria-checked')

                    if active == 'true':
                        material.append(a.get_text())


        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Порода дерева', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    active = a.get_text()

                    if active not in material:
                        material.append(active)

        print(material)

        for i in prop_list:
            val = i.get_text()

            w_match = re.search('Ширина', val)
            h_match = re.search('Висота', val)
            d_match = re.search('Глибина', val)

            if w_match:
                size[0] = num_check(val)

            if h_match:
                size[1] = num_check(val)

            if d_match:
                size[2] = num_check(val)

        print(size)

        cards.append({
            'id': url['num'],
            'name': name,
            'des': desc_text,
            'w': size[0],
            'h': size[1],
            'd': size[2],
            'price': price,
        })

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Komodytumby.objects.filter(pars_name = product['name']):
                print('old', Komodytumby.objects.get(pars_name=product['name']).name)
                model_id =  Komodytumby.objects.get(pars_name=product['name']).id
                Komodytumby.objects.filter(id = model_id).update(price=product['price'])
                        
                
            #Додаємо новий товар
            else:
                new_id = Komodytumby.objects.all().order_by('-id')[0].id + 1

                Komodytumby.objects.get_or_create(id=new_id, manufacturer_id=14, name=product['name'],\
                    pars_name=product['name'], description=product['des'],\
                        width=product['w'], deep=product['d'], price=product['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Komodytumby_img.objects.get_or_create(key_img_id=new_id, img=i['img'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_komody_tumby_lion():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=7).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[2]
    cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = sheet[r][1].value

        match = re.search('Комоди,   направляючі телескопічні ', str(cell))

        if match:
            r = r + 2
            while True:
                r += 1
                cell = sheet[r][1].value

                if cell:
                    name = str(sheet[r][1].value)
                    price = round(int(sheet[r][5].value))
                    size = str(sheet[r][4].value)
                    size = size.split('х')

                    cards.append({
                        'id': prod_id,
                        'name': name,
                        'des': '',
                        'w': size[0],
                        'h': size[2],
                        'd': size[1],
                        'price': price,
                    })

                    print(prod_id, name, price, size)
                    prod_id += 1

                else:
                    break

    print('-------------------------------------------------------')

    for r in range(sheet.max_row):
        r += 1
        cell = sheet[r][1].value

        
        match = re.search('Тумби до ліжок,   направляючі телескопічні ', str(cell))

        if match:
            r = r + 2
            while True:
                r += 1
                cell = sheet[r][1].value

                if cell == "Офісні меблі, столи геймерські, комп'ютерні":
                    break

                else:
                    name = str(sheet[r][1].value)
                    price = round(int(sheet[r][5].value))
                    size = str(sheet[r][4].value)
                    size = size.split('х')

                    cards.append({
                        'id': prod_id,
                        'name': name,
                        'des': '',
                        'w': size[0],
                        'h': size[2],
                        'd': size[1],
                        'price': price,
                    })

                    print(prod_id, name, price, size)
                    prod_id += 1

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        

        try:
            #Оновлюємо ціну
            if Komodytumby.objects.filter(pars_name = product['name']):
                print('old', Komodytumby.objects.get(pars_name=product['name']).name)
                model_id =  Komodytumby.objects.get(pars_name=product['name']).id
                Komodytumby.objects.filter(id = model_id).update(price=product['price'])
                        
                
            #Додаємо новий товар
            else:
                new_id = Komodytumby.objects.all().order_by('-id')[0].id + 1

                Komodytumby.objects.get_or_create(id=new_id, manufacturer_id=6, name=product['name'],\
                    pars_name=product['name'], description=product['des'],\
                        width=product['w'], deep=product['d'], price=product['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_komody_tumby_svitmebliv():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=6).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[2]

    cards = []
    size_cards = []
    prod_id = 0

    for c in range(sheet.max_column):
        for r in range(sheet.max_row):
            r += 1
            cell = sheet[r][c].value
            match = re.search('комоди', str(cell))

            if match:
                while True:
                    r += 1
                    name = str(sheet[r][6].value)
                    price = str(sheet[r][10].value)
                    match = re.search('комод', name)

                    if match:
                        cards.append({
                            'id': prod_id,
                            'name': name,
                            'des': '',
                            'w': '',
                            'h': '',
                            'd': '',
                            'price': price,
                        })

                        print(prod_id, name, price)
                        prod_id += 1

                    else:
                        break

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Komodytumby.objects.filter(pars_name = product['name']):
                print('old', Komodytumby.objects.get(pars_name=product['name']).name)
                model_id =  Komodytumby.objects.get(pars_name=product['name']).id
                Komodytumby.objects.filter(id = model_id).update(price=product['price'])
                        
                
            #Додаємо новий товар
            else:
                new_id = Komodytumby.objects.all().order_by('-id')[0].id + 1

                Komodytumby.objects.get_or_create(id=new_id, manufacturer_id=2, name=product['name'],\
                    pars_name=product['name'], description=product['des'],\
                        width=product['w'], deep=product['d'], price=product['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_komody_tumby_olimp():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=8).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    prod_id = 0
    cards = []

    for r in range(sheet.max_row):

        r += 1
        cell = str(sheet[r][0].value)
        match = re.search('Тумби і комоди', cell)

        if match:

            while True:
                r += 1

                name = str(sheet[r][0].value)
                match = re.search('Столи та столики', name)


                if match:
                    break

                else:
                    try:

                        price = round(int(sheet[r][3].value) * 1.4)

                        cards.append({
                            'id': prod_id,
                            'name': name,
                            'des': '',
                            'w': '',
                            'h': '',
                            'd': '',
                            'price': price,
                        })

                        print(prod_id, name, price)
                        prod_id += 1

                    except:
                        pass

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Komodytumby.objects.filter(pars_name = product['name']):
                print('old', Komodytumby.objects.get(pars_name=product['name']).name)
                model_id =  Komodytumby.objects.get(pars_name=product['name']).id
                Komodytumby.objects.filter(id = model_id).update(price=product['price'])
                        
                
            #Додаємо новий товар
            else:
                new_id = Komodytumby.objects.all().order_by('-id')[0].id + 1

                Komodytumby.objects.get_or_create(id=new_id, manufacturer_id=13, name=product['name'],\
                    pars_name=product['name'], description=product['des'],\
                        width=product['w'], deep=product['d'], price=product['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_komody_tumby_kompanit():
    type_link = ['https://kompanit.com.ua/catalog/tumbi-pid-vzuttya/c24',
                 'https://kompanit.com.ua/catalog/tumbi-pid-tv/c1',
                 'https://kompanit.com.ua/catalog/tumbi-prilizhkovi/c18',
                 'https://kompanit.com.ua/catalog/komodi/c5',
                 'https://kompanit.com.ua/catalog/tryumo/c23'
                 ]

    pag_list = []
    prod_link = []
    cards = []
    img_cards = []


    for i in type_link:
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        pag_list.append(i)
        pagination = soup.find_all('a', class_='pagination__number')

        # Забераємо пагенацію

        for p in pagination:
            pag = p.get('href')
            if pag not in pag_list and pag:
                pag_list.append(pag)
        time.sleep(1)

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        links = soup.find_all('div', class_='gcell gcell--12 gcell--xs-6 gcell--md-4')

        for l in links:
            link = l.find('a').get('href')

            if link not in prod_link:
                prod_link.append({
                    'num': len(prod_link),
                    'url': link
                })

        time.sleep(1)

    # Забераємо інформацію про товар

    for url in prod_link:
        source = requests.get(url=url['url'],headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        name = soup.find('h1').get_text()
        color_desc = soup.find('div', class_='item-colors')
        desc_text = soup.find('div', class_='tabs _mb-sm').get_text()
        prop_list = soup.find('div', class_='tabs _mb-sm').find_all('li')
        size = ['', '', '']



        print(url['url'])
        print(name)

        for i in prop_list:
            val = i.get_text()

            w_match = re.search('Ширина', val)
            h_match = re.search('Висота', val)
            d_match = re.search('Глибина', val)

            if w_match:
                size[0] = num_check(val)

            if h_match:
                size[1] = num_check(val)

            if d_match:
                size[2] = num_check(val)

        cards.append({
            'id': url['num'],
            'name': name,
            'des': desc_text + str(color_desc),
            'w': size[0],
            'h': size[1],
            'd': size[2],
            'price': '',
        })

        img_list = soup.find('div', class_='gcell gcell--12 gcell--def-6').find_all('a')

        for i in img_list:
            img = i.get('data-mfp-src')
            print(img)
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': url['num'],
                'img': img_name
            })
        
        print(size)

        print('-------------------------------')


    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Komodytumby.objects.filter(pars_name = product['name']):
                print('old', Komodytumby.objects.get(pars_name=product['name']).name)
                model_id =  Komodytumby.objects.get(pars_name=product['name']).id
                #Komodytumby.objects.filter(id = model_id).update(price=product['price'])
                        
                
            #Додаємо новий товар
            else:
                new_id = Komodytumby.objects.all().order_by('-id')[0].id + 1

                Komodytumby.objects.get_or_create(id=new_id, manufacturer_id=19, name=product['name'],\
                    pars_name=product['name'], description=product['des'],\
                        width=product['w'], deep=product['d'], price=product['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Komodytumby_img.objects.get_or_create(key_img_id=new_id, img=i['img'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_komody_tumby_neman():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=9).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet_1 = book.worksheets[0]

    prod_id = 0
    cards = []
    img_cards = []

    for r in range(sheet_1.max_row):
        r += 1

        cell = str(sheet_1[r][3].value)
        match = re.search('Комод', cell)
        match2 = re.search('Тумба', cell)

        if match or match2:
            prod_id += 1
            prom = str(sheet_1[r][0].value)
            name = str(sheet_1[r][3].value)
            des = str(sheet_1[r][37].value)
            name_ru = str(sheet_1[r][4].value)
            des_ru = str(sheet_1[r][38].value)
            price = str(sheet_1[r][9].value)[:-3]
            img_list = str(sheet_1[r][15].value).split(';')
            img_list = [line.rstrip() for line in img_list]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'name_ru': name_ru,
                'des': des,
                'des_ru': des_ru,
                'w': '',
                'h': '',
                'd': '',
                'price': price,
            })



            for i in img_list:
                try:
                    img = i
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_name = str(img_name).replace('%', '')
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })
                except:
                    pass

            print('--------------------------')
            print(prom)
            print(name, price)

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Komodytumby.objects.filter(pars_name = product['prom']).filter(manufacturer_id=10):
                print('old', Komodytumby.objects.get(pars_name=product['prom']).name)
                model_id =  Komodytumby.objects.get(pars_name=product['prom']).id
                Komodytumby.objects.filter(id = model_id).update(price=product['price'])
                        
                
            #Додаємо новий товар
            else:
                new_id = Komodytumby.objects.all().order_by('-id')[0].id + 1

                Komodytumby.objects.get_or_create(id=new_id, manufacturer_id=10,\
                    name=product['name'], name_ru = product['name_ru'],\
                        description=product['des'], description_ru=product['des_ru'], pars_name=product['prom'],\
                            width=product['w'], deep=product['d'], price=product['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Komodytumby_img.objects.get_or_create(key_img_id=new_id, img=i['img'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_komody_tumby_comfortmebli():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=25).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    img_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][8].value).split(';')
            img_list.insert(0, str(sheet[r][6].value)) 

            des = str(sheet[r][9].value)


            cards.append({
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
                'price':price
            })

            print(prom, '-->', name, '->', price)

            for img in img_list:
                print(f"img---{img}---///")
                
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })

                except Exception as ex:
                    print(f'///-------{ex}---------///')
            
            print('-------------------------')

    #Додаємо записи в базу
    new_db = []
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Komodytumby.objects.filter(pars_name = product['prom']).filter(manufacturer_id=1):
                print('old', Komodytumby.objects.get(pars_name=product['prom']).name)
                new_db.append(product['prom'])
                model_id =  Komodytumby.objects.get(pars_name=product['prom']).id
                Komodytumby.objects.filter(id = model_id).update(price=product['price'])
                        
                
            #Додаємо новий товар
            else:
                new_id = Komodytumby.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Komodytumby.objects.get_or_create(id=new_id, manufacturer_id=1,\
                    name=product['name'],\
                        description=product['des'], pars_name=product['prom'],\
                            price=product['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Komodytumby_img.objects.get_or_create(key_img_id=new_id, img=i['img'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')
        
    #Видаляємо товар
    for m in Komodytumby.objects.filter(manufacturer_id=1):
        print(m)
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()