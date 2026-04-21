import requests
import openpyxl
from bs4 import BeautifulSoup
import csv
import time
import re
import os
import json
from main.models import Seria, Stinka, Stinka_img
from pars.models import File


with open('/home/ay507291/notacomforta.pl.ua/www/pars/src.json', 'r') as f:
    file = json.load(f)

path_dirver = file['src'][0]['path_dirver']
accept = file['src'][1]['accept']
user_agent = file['src'][2]['user-agent']

HEADERS = {
    'accept': accept,
    'user-agent': user_agent
}


path = os.getcwd()


def num_check(text):
    num = ''
    for t in text:
        if t.isdigit():
            num += t
    return num



def test(modul_cards, product_cards):
    pass
    for m in modul_cards:
        print(m['id'], m['name'])
        for p in product_cards:
            if m['id'] == p['modul_id']:
                print(f"modul: {p['modul_id']} id: {p['id']} name: {p['name']} price: {p['price']}")


def UpdateData(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['name']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['name']).id

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    if Stinka.objects.filter(pars_name=p['name']).filter(seria_id=modul_id):
                        new_db.append(p['name'])
                        print('old:', p['name'])
                        Stinka.objects.filter(pars_name=p['name']).filter(seria_id=modul_id).update(price=p['price'], pars_category='peredpokoi')
                    else:
                        print('none:', p['name'])
                        new_db.append(p['name'])
                        new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                        Stinka.objects.create(id=new_sid, seria_id=modul_id,  manufacturer_id=manu,\
                                          peredpokoi=True, name=p['name'], pars_name=p['name'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                        
                        try:
                            for i in img_cards:
                                if i['id'] == p['id']:
                                    Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                        except:
                            pass

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['name'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          peredpokoi=True, name=p['name'], pars_name=p['name'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['name'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass

    '''#Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(peredpokoi=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()'''


def UpdateDataBMK(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']):
            modul_id = Seria.objects.get(pars_name = m['prom']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['prom'] == el.pars_name:
                        new_db.append(p['prom'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(price=p['price'], pars_category='peredpokoi')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          peredpokoi=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
        
    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(pars_category='peredpokoi'):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def UpdateDataKompanit(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']):
            modul_id = Seria.objects.get(pars_name = m['prom']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['prom'] == el.pars_name:
                        new_db.append(p['prom'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(pars_category='peredpokoi')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          peredpokoi=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
        
    '''#Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(peredpokoi=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()'''


def UpdateDataSvitMebliv(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['prom']).id

            #print('-->', m['name'])
            for p in product_cards:
                if p['modul_id'] == m['id']:
                    if Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id):
                        new_db.append(p['prom'])
                        #print('old:', p['name'])
                        Stinka.objects.filter(pars_name=p['prom']).filter(seria_id=modul_id).update(price=p['price'], pars_category='peredpokoi')
                    else:
                        #print('none:', p['name'])
                        new_db.append(p['prom'])
                        Stinka.objects.create(seria_id=modul_id, manufacturer_id=manu,\
                                          peredpokoi=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                        
        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          peredpokoi=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'], pars_category='peredpokoi')
                    new_db.append(p['prom'])
                    #print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
    
    #Видаляємо товар
    '''for m in Stinka.objects.filter(manufacturer_id=manu).filter(peredpokoi=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()'''


def UpdateDataEverest(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['name']).filter(manufacturer_id=manu):
            modul_id = Seria.objects.get(pars_name = m['name']).id

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    if Stinka.objects.filter(pars_name=p['name']).filter(seria_id=modul_id):
                        new_db.append(p['name'])
                        print('old:', p['name'])
                        Stinka.objects.filter(pars_name=p['name']).filter(seria_id=modul_id).update(pars_category='peredpokoi')
                    else:
                        print('none:', p['name'])
                        new_db.append(p['name'])
                        new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                        Stinka.objects.create(id=new_sid, seria_id=modul_id,  manufacturer_id=manu,\
                                          peredpokoi=True, name=p['name'], pars_name=p['name'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                        
                        try:
                            for i in img_cards:
                                if i['id'] == p['id']:
                                    Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                        except:
                            pass

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['name'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          peredpokoi=True, name=p['name'], pars_name=p['name'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['name'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass

    '''#Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(peredpokoi=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()'''


def UpdateDataKM(modul_cards, product_cards, img_cards, manu):
    #Додаємо записи в базу
    new_db=[]
    for m in modul_cards:

        #Оновлюємо ціну
        if Seria.objects.filter(pars_name = m['prom']):
            modul_id = Seria.objects.get(pars_name = m['prom']).id
            element = Stinka.objects.filter(seria_id = modul_id)

            for el in element:
                for p in product_cards:
                    if p['prom'] == el.pars_name:
                        new_db.append(p['prom'])
                        print('old:', Stinka.objects.get(id=el.id))
                        Stinka.objects.filter(id = el.id).update(price=p['price'], pars_category='peredpokoi')

        #Додаємо новий товар
        else:
            new_id = Seria.objects.all().order_by('-id')[0].id + 1
            Seria.objects.create(id=new_id, manufacturer_id=manu, seria_name=m['name'], pars_name=m['prom'])

            for p in product_cards:
                if p['modul_id'] == m['id']:
                    new_sid = Stinka.objects.all().order_by('-id')[0].id + 1
                    Stinka.objects.create(id=new_sid, seria_id=new_id, manufacturer_id=manu,\
                                          peredpokoi=True, name=p['name'], pars_name=p['prom'], categori = 'STINKA',\
                                            description=p['des'], price=p['price'])
                    new_db.append(p['prom'])
                    print('new:', p['name'])
                    
                    try:
                        for i in img_cards:
                            if i['id'] == p['id']:
                                Stinka_img.objects.get_or_create(key_img_id=new_sid, img=i['img'])
                    except:
                        pass
        
    #Видаляємо товар
    for m in Stinka.objects.filter(manufacturer_id=manu).filter(peredpokoi=True):
        if m.pars_name and m.pars_name not in new_db:
            print('delet:', m)
            m.delete()




def get_peredpokoi_products_comfortmebli():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=21).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    
    modul_cards = []
    product_cards = []
    img_cards = []
    modul_id = 0
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            modul_id += 1

            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][9].value).split(';')
            img_list.insert(0, str(sheet[r][7].value)) 

            des = str(sheet[r][10].value)


            product_cards.append({
                'modul_id':modul_id,
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
                'price':price
            })

            modul_cards.append({
                'prom':prom,
                'id':modul_id,
                'name':name,
            })

            print(prom, '-->', name, '->', price)

            for img in img_list:
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })

                except Exception as ex:
                    print(f'///-------{ex}---------///')

            print('-------------------------')

    test(modul_cards, product_cards)
    UpdateDataKM(modul_cards, product_cards, img_cards, manu=1)

'''
def get_peredpokoi_products_lion():
    book = openpyxl.load_workbook(filename="pars_file/ОПТ-Лион-Меблі-для-дому_-офісу-с.xlsx")
    sheet = book.worksheets[2]


    with open("static/file/peredpokoi_seria_comfortmebli.csv", newline='') as f:
        id_list = []
        r = csv.DictReader(f, delimiter=";")

        for i in r:
            id_list.append(int(i['id']))
            seria_id = sorted(id_list)[-1] + 1

    with open("static/file/peredpokoi_stinka_comfortmebli.csv", newline='') as f:
        id_list = []
        r = csv.DictReader(f, delimiter=";")

        for i in r:
            id_list.append(int(i['id']))
            stinka_id = sorted(id_list)[-1] + 1

    with open("static/file/peredpokoi_comody_tumby_comfortmebli.csv", newline='') as f:
        id_list = []
        r = csv.DictReader(f, delimiter=";")

        for i in r:
            id_list.append(int(i['id']))
            ct_id = sorted(id_list)[-1] + 1

    with open("static/file/peredpokoi_other_prod_comfortmebli.csv", newline='') as f:
        id_list = []
        r = csv.DictReader(f, delimiter=";")

        for i in r:
            id_list.append(int(i['id']))
            other_id = sorted(id_list)[-1] + 1

    seria_cards = []
    stinka = []
    comody_tumby = []
    other_prod = []

    for i in range(sheet.max_row):
        i += 1
        cell = str(sheet[i][1].value).replace(" ", "")
        des_list = []
        match = re.search("Передпокої", cell)

        if match:
            while True:
                try:
                    i += 1
                    name = str(sheet[i][1].value)
                    des = str(sheet[i][4].value)
                    price = round(int(sheet[i][5].value))
                    w = des.split('х')[0]
                    des_list.append(name + des + " " + str(price) + 'грн')
                    comod = re.search('Комод', name)
                    tymba = re.search('Тумба', name)

                    if comod or tymba:
                        comody_tumby.append({
                            'id': ct_id,
                            'seria_id': seria_id,
                            'name': name,
                            'des': des,
                            'w': w,
                            'price': price
                        })
                        print(seria_id, ct_id, name, des)
                        ct_id += 1

                    else:
                        other_prod.append({
                            'id': other_id,
                            'seria_id': seria_id,
                            'name': name,
                            'des': des,
                            'w': w,
                            'price': price
                        })
                        print(seria_id, other_id, name, des)
                        other_id += 1

                except Exception as ex:
                    # print(ex)
                    match = re.search("Дрібні меблі", name)
                    if match:
                        break

                    elif name == "None":
                        pass

                    else:
                        seria_id += 1
                        stinka_id += 1
                        seria_cards.append({
                            'id': seria_id,
                            'name': name
                        })

                        stinka.append({
                            'id': stinka_id,
                            'seria_id': seria_id,
                            'name': name,
                            'des': " ".join(des_list),
                            'w': '',
                            'price': ''
                        })
                        print(seria_id, stinka_id, name)

    with open("static/file/peredpokoi_seria_lion.csv", 'w', newline='') as csvfile:
        w = csv.writer(csvfile, delimiter=";")
        w.writerow(["id", "name"])

        for i in seria_cards:
            w.writerow([i['id'], i['name']])

    with open("static/file/peredpokoi_stinka_lion.csv", 'w', newline='') as csvfile:
        w = csv.writer(csvfile, delimiter=";")
        w.writerow(["id", "seria_id", "name", "des", "w", "price"])

        for i in stinka:
            w.writerow([i['id'], i['seria_id'], i['name'], i['des'], i['w'], i['price']])

    with open("static/file/peredpokoi_comody_tumby_lion.csv", 'w', newline='') as csvfile:
        w = csv.writer(csvfile, delimiter=";")
        w.writerow(["id", "seria_id", "name", "des", "w", "price"])

        for i in comody_tumby:
            w.writerow([i['id'], i['seria_id'], i['name'], i['des'], i['w'], i['price']])

    with open("static/file/peredpokoi_other_prod_lion.csv", 'w', newline='') as csvfile:
        w = csv.writer(csvfile, delimiter=";")
        w.writerow(["id", "seria_id", "name", "des", "w", "price"])

        for i in other_prod:
            w.writerow([i['id'], i['seria_id'], i['name'], i['des'], i['w'], i['price']])
 '''           

def get_peredpokoi_products_svitmebliv():
    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0

    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=6).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]

    
    for r in range(sheet.max_row):
        r += 1
        for c in range(sheet.max_column):
            cell = str(sheet[r][c].value)
            match = re.search("Прихожі", cell)

            if match:
                while True:
                    r += 1
                    try:
                        name = sheet[r][c].value
                        price = int(sheet[r][c + 4].value)
                        prom = name.replace(' ', '').lower()
                        modul_cards.append({
                            'id': modul_id,
                            'prom':prom,
                            'name': name
                        })

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': '',
                            'des': '',
                            'price': price
                        })

                        modul_id += 1
                        prod_id += 1

                        print(f'Прихожа --> {name}')

                    except:
                        break

    for r in range(sheet.max_row):
        r += 1
        for c in range(sheet.max_column):
            cell = str(sheet[r][c].value)
            match = re.search("Прихожа", cell)
            
            if match:
                
                modul_id += 1
                prod_id += 1

                name = cell[7:]
                
                print(name)
                
                column = c + 4
                row = r
                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })

                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'w': '',
                    'des': '',
                    'price': ''
                })


                while True:
                    row += 1
                    try:
                        name = str(sheet[row][c].value)
                        price = int(sheet[row][column].value)
                        prod_id += 1
                        prom = name.replace(' ', '').lower()

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': '',
                            'des': '',
                            'price': price
                        })
                        

                    except:
                        break
            else:
                pass
    
    for r in range(sheet.max_row):
        r += 1
        for c in range(sheet.max_column):
            cell = str(sheet[r][c].value)
            match = re.search("тумба під взуття", cell)
            
            if match:
                
                modul_id += 1
                prod_id += 1

                name = cell
                
                print(f"{modul_id} {name}")
                
                column = c + 4
                row = r
                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })

                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'w': '',
                    'des': '',
                    'price': ''
                })


                while True:
                    row += 1
                    try:
                        name = str(sheet[row][c].value)
                        price = int(sheet[row][column].value)
                        prod_id += 1
                        prom = name.replace(' ', '').lower()

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': '',
                            'des': '',
                            'price': price
                        })
                        
                        print(f"{modul_id} {prod_id} {name} {prom} {price}")
                    except:
                        break
            else:
                pass

    #test(modul_cards, product_cards)
    UpdateDataSvitMebliv(modul_cards, product_cards, img_cards=[], manu=2)
   

def get_peredpokoi_products_gerbor():
    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0
    row = 0

    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=15).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]

    for r in range(sheet.max_row):
            r += 1
            cell = str(sheet[r][2].value)

            if cell == 'ПЕРЕДПОКОЇ':
                row = r


    for r in range(sheet.max_row):
        r += row
        cell = str(sheet[r][2].value)

        if cell =='КУХНІ':
            break

        elif str(sheet[r+1][1].value) == '1':
            modul_id += 1
            name = str(sheet[r][2].value)
            prom = name.lower()
            prom = prom.replace(' ', '')

            modul_cards.append({
                'id': modul_id,
                'prom':prom,
                'name': name
            })

            print(modul_id, '--->', name, prom)

            while True:
                r += 1
                try:
                    prod_id += 1
                    name = sheet[r][2].value
                    price = int(sheet[r][5].value)
                    prom = str(sheet[r][3].value)
                    
                    product_cards.append({
                        'modul_id': modul_id,
                        'id': prod_id,
                        'prom':prom,
                        'name': name,
                        'w': '',
                        'des': '',
                        'price': price
                    })
                    print(modul_id, name, prom, price)
                except:
                    break

    UpdateDataBMK(modul_cards, product_cards, img_cards=[], manu=8)

    
def get_peredpokoi_products_everest():

    source = requests.get('https://everestm.com.ua/prihozhie/prihozhiye/',headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')
    pag_list = ['https://everestm.com.ua/prihozhie/prihozhiye/', ]

    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0
    img_cards = []
    link_list = []

    # Забераємо пагенацію
    try:
        item = soup.find('ul', class_='pagination').find_all('a')
        for i in item:
            if i.get('href') not in pag_list:
                pag_list.append(i.get('href'))
        time.sleep(2)
    except:
        pass

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        item = soup.find('div', class_='products-block').find_all('a', class_='product-name')
        for i in item:
            if i.get('href') not in link_list:
                link_list.append(i.get('href'))

        time.sleep(2)

    # Забераємо інформацію про товар

    for i in link_list:
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        name = soup.find('h1').get_text()
        desc_text = soup.find('div', class_='nav-desc').get_text()
        char = soup.find('div', class_='nav-characteristic').get_text()
        desc_text = desc_text + char

        try:
            price = soup.find('span', class_='autocalc-product-special').get_text()

        except:
            price = soup.find('span', class_='autocalc-product-price').get_text()

        modul_cards.append({
            'id': modul_id,
            'name': name
        })

        product_cards.append({
            'modul_id': modul_id,
            'id': prod_id,
            'name': name,
            'w': '',
            'des': desc_text,
            'price': price
        })
        
        try:
            img = soup.find('a', class_='thumbnail').get('href')
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name.replace('%', ''), "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': prod_id,
                'img': img_name.replace('%', '')
            })
        except:
            pass
        
        modul_id += 1
        prod_id += 1
    
    test(modul_cards, product_cards)
    UpdateDataEverest(modul_cards, product_cards, img_cards , manu=9)


def get_peredpokoi_products_kompanit():
    source = requests.get('https://kompanit.com.ua/catalog/peredpokiy/c13',headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')

    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0
    img_cards = []
    link_list = []
    pag_list = ['https://kompanit.com.ua/catalog/peredpokiy/c13', ]

    # Забераємо пагенацію
    try:
        item = soup.find('div', class_='pagination').find_all('a')
        for i in item:
            if i.get('href') not in pag_list and i.get('href'):
                print('-->', i.get('href'))
                pag_list.append(i.get('href'))
        time.sleep(2)
    except:
        pass

        for url in pag_list:
            source = requests.get(url,headers=HEADERS).text
            soup = BeautifulSoup(source, 'html.parser')
            item = soup.find_all('a', class_='product__name')
            for i in item:
                if i.get('href') not in link_list:
                    link_list.append(i.get('href'))
                    print(len(link_list), i.get('href'))

        time.sleep(1)

    # Забераємо інформацію про товар
    for i in link_list:
        print('----------------------------------')
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
    
        prom = i
        name = soup.find('h1').getText()
        desc_text = soup.find('div', class_='tabs _mb-sm').getText(strip=True)

        print(prom)
        print(name)


        modul_cards.append({
            'id': modul_id,
            'name': name,
            'prom':prom
        })

        product_cards.append({
            'modul_id': modul_id,
            'id': prod_id,
            'prom':prom,
            'name': name,
            'w': '',
            'des': desc_text,
            'price':''
        })

        try:
            img = soup.find('img', class_='js-lazyload slider__slide-img').get('data-zzload-source-img')
            print(img)
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name.replace('%', ''), "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': prod_id,
                'img': img_name.replace('%', '')
            })
        except:
            pass

        modul_id += 1
        prod_id += 1

    UpdateDataKompanit(modul_cards, product_cards, img_cards, manu=19)