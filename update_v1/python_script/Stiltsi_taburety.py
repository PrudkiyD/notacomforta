import requests
import xml.etree.ElementTree as ET
import openpyxl
from bs4 import BeautifulSoup
import csv
import time
import re
import os
import json
from stiltsi_taburety.models import Stiltsi_taburety, Stiltsi_taburety_size, Stiltsi_taburety_img
from pars.models import File


with open('/home/ay507291/notacomforta.pl.ua/www/pars/src.json', 'r') as f:
    file = json.load(f)

path_dirver = file['src'][0]['path_dirver']
accept = file['src'][1]['accept']
user_agent = file['src'][2]['user-agent']

HEADERS = {
    'accept': accept,
    'user-agent': user_agent
}


def num_check(text):
    num = ''
    for t in str(text):
        if t.isdigit():
            num += t
    return num


def name():
    for s in Stiltsi_taburety.objects.filter(manufacturer_id=5):
        print('---------------')
        print(s.name)
        print(s.pars_name)



def get_stiltsi_taburety_arbordrev():
    pag_list = ['https://arbordrev.com.ua/tables-and-chairs/armchairs', ]
    links_list = []

    cards = []
    size_cards = []
    img_cards = []
    color_cards = []
    link_list = []
    prod_id = 0

    # Забераємо пагенацію

    for num in range(100):
        try:
            source = requests.get(pag_list[num],headers=HEADERS).text
            soup = BeautifulSoup(source, 'html.parser')
            pagination = soup.find('div', class_='wd-loop-footer products-footer').find('a').get('href')

            if pagination not in pag_list:
                pag_list.append(pagination)

        except:
            break

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        links = soup.find_all('h3', class_='wd-entities-title')

        for i in links:
            link = i.find('a').get('href')
            if link not in links_list:
                links_list.append({
                    'num': len(links_list),
                    'url': link
                })
                print(len(links_list), link)

        time.sleep(1)

    # Забераємо інформацію про товар

    for url in links_list:
        print(url['url'])
        source = requests.get(url['url'],headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        name = soup.find('h1').get_text()
        price = str(soup.find('div', class_='col-lg-6 col-12 col-md-6 wd-price-outside summary entry-summary').find('bdi').getText()).replace('грн', '')
        desc_text = soup.find('div', class_='wc-tab-inner').get_text()
        chek_list = soup.find('table', class_='variations').find('tbody').find_all('tr')
        img_list = soup.find('div', class_='product-images-inner').find_all('img')
        char = soup.find('table', class_='woocommerce-product-attributes shop_attributes').get_text()

        desc_text = char + desc_text



        for img in img_list:
            img = img.get('data-large_image')
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img,headers=HEADERS).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': url['num'],
                'img': img
            })
            print(img)

        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Колір', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    color = a.find('img').get('src')
                    color_cards.append({
                        'id': url['num'],
                        'img': color
                    })

        print(url['num'], name, price)
        print(desc_text)

        cards.append({
            'id': url['num'],
            'name': name,
            'des': desc_text,
        })

        size_cards.append({
            'id': url['num'],
            'w': '',
            'd': '',
            'price': price,
        })


    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Stiltsi_taburety.objects.filter(pars_name = product['name']):
                print('old', Stiltsi_taburety.objects.get(pars_name=product['name']).name)
                model_id =  Stiltsi_taburety.objects.get(pars_name=product['name']).id
                size_id = Stiltsi_taburety_size.objects.filter(stiltsi_taburety_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.filter(id = size_id).update(price=s['price'])
                        break
                
            #Додаємо новий товар
            else:
                new_id = Stiltsi_taburety.objects.all().order_by('-id')[0].id + 1

                Stiltsi_taburety.objects.get_or_create(id=new_id, manufacturer_id=14, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.get_or_create(stiltsi_taburety_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Stiltsi_taburety_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_stiltsi_taburety_kompanit():
    type_link = ['https://kompanit.com.ua/catalog/tabureti/c9'
                 ]

    pag_list = []
    prod_link = []
    cards = []
    size_cards = []
    img_cards = []

    for i in type_link:
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        pag_list.append(i)
        pagination = soup.find_all('a', class_='pagination__number')

        # Забераємо пагенацію

        for p in pagination:
            pag = p.get('href')
            if pag not in pag_list and pag:
                pag_list.append(pag)
        time.sleep(1)

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        links = soup.find_all('div', class_='gcell gcell--12 gcell--xs-6 gcell--md-4')

        for l in links:
            link = l.find('a').get('href')

            if link not in prod_link:
                prod_link.append({
                    'num': len(prod_link),
                    'url': link
                })

        time.sleep(1)

    # Забераємо інформацію про товар

    for url in prod_link:
        source = requests.get(url['url'],headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        name = soup.find('h1').get_text()
        color_desc = soup.find('div', class_='item-colors')
        desc_text = soup.find('div', class_='tabs _mb-sm').get_text()
        slick = soup.find_all('a', class_='slider__slide slick-slide')
        prop_list = soup.find('div', class_='tabs _mb-sm').find_all('li')
        size = ['', '']


        print(url['url'])
        print(name)

        for i in prop_list:
            val = i.get_text()

            w_match = re.search('Ширина', val)

            d_match = re.search('Глибина', val)

            if w_match:
                size[0] = num_check(val)

            if d_match:
                size[1] = num_check(val)

        cards.append({
            'id': url['num'],
            'name': name,
            'des': desc_text + str(color_desc),
        })

        size_cards.append({
            'id': url['num'],
            'w': size[0],
            'd': size[1],
            'price': '',
        })

        img_list = soup.find('div', class_='gcell gcell--12 gcell--def-6').find_all('a')

        for i in img_list:
            img = i.get('data-mfp-src')
            print(img)
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': url['num'],
                'img': img_name
            })

            print(img_name)

        print(size)
        print('-------------------------------')

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Stiltsi_taburety.objects.filter(pars_name = product['name']):
                print('old', Stiltsi_taburety.objects.get(pars_name=product['name']).name)
                model_id =  Stiltsi_taburety.objects.get(pars_name=product['name']).id
                size_id = Stiltsi_taburety_size.objects.filter(stiltsi_taburety_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Stiltsi_taburety.objects.all().order_by('-id')[0].id + 1

                Stiltsi_taburety.objects.get_or_create(id=new_id, manufacturer_id=19, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.get_or_create(stiltsi_taburety_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Stiltsi_taburety_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_stiltsi_taburety_tenero():
    source = requests.get('https://tenero.in.ua/google_merchant_center.xml?hash_tag=f849745a8bde8e71cf3a96764a271ef1&product_ids=&label_ids=&export_lang=ru&group_ids=', headers=HEADERS)
    source.encoding = 'utf-8'

    soup = BeautifulSoup(source.text, 'xml')


    item = soup.find_all('item')
    prod_link = []
    cards = []
    img_cards = []
    size_cards = []

    for i in item:
        type = i.find('g:product_type').get_text()
        img = i.find('g:image_link').get_text()
        match = re.search('кухонные стулья и табуретки', type)
        match_2 = re.search('стулья ученические', type)
        match_3 = re.search('офисные стулья', type)
        match_4 = re.search('стулья и кресла барные', type)

        link = i.find('g:link').get_text()
        link = link[:21] + 'ua/' + link[21:]
        if match or match_2 or match_3 or match_4:
            prod_link.append({
                'num': len(prod_link),
                'url': link,
                'img': img
            })

    for url in prod_link:
        source = requests.get(url['url'], headers=HEADERS).content
        soup = BeautifulSoup(source, 'html.parser')

        print(url['url'])
        
        name = soup.find('h1').getText()
        desc_text = soup.find('div', class_='UserContent__root--vmWEL')
        price = soup.find('p', class_='Text__ui_text_size_xs--34h-R').find('span').getText()

        cards.append({
            'id': url['num'],
            'name': name,
            'des': str(desc_text),
        })

        size_cards.append({
            'id': url['num'],
            'option': '',
            'gear': '',
            'w': '',
            'h': '',
            'd': '',
            'price': price,
        })

        img = url['img']
        img_name = img.split('/')[len(img.split('/')) - 1]
        img_bytes = requests.get(img).content
        with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
            f.write(img_bytes)

        img_cards.append({
            'id': url['num'],
            'img': img_name
        })
        

        
        print(url['num'], name)
        print(price)
        print('-------------------------------------')

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Stiltsi_taburety.objects.filter(pars_name = product['name']):
                print('old', Stiltsi_taburety.objects.get(pars_name=product['name']).name)
                model_id =  Stiltsi_taburety.objects.get(pars_name=product['name']).id
                size_id = Stiltsi_taburety_size.objects.filter(stiltsi_taburety_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Stiltsi_taburety.objects.all().order_by('-id')[0].id + 1

                Stiltsi_taburety.objects.get_or_create(id=new_id, manufacturer_id=12, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.get_or_create(stiltsi_taburety_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Stiltsi_taburety_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_stiltsi_taburety_modul_lux():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=10).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    prod_id = 0
    cards = []
    size_cards = []

    for r in range(sheet.max_row):

        r += 1
        cell = str(sheet[r][0].value)
        match = re.search('Стільці', cell)

        if match:
            r = r+3
            while True:
                r += 1

                cell = str(sheet[r][0].value)
                name = str(sheet[r][1].value)
                size = str(sheet[r][2].value)
                try:
                    price = str(round(float(sheet[r][3].value)))
                except:
                    price = ''

                size = size.split('*')

                if price.isdigit() == False:
                    break

                else:

                    cards.append({
                        'id': prod_id,
                        'name': name,
                        'des': '',
                    })

                    size_cards.append({
                        'id': prod_id,
                        'w': size[0],
                        'h': size[1],
                        'd': size[2],
                        'price': price,
                    })

                    prod_id += 1

                    print(prod_id, name, size, price)

    
    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Stiltsi_taburety.objects.filter(pars_name =str(product['name']).replace(' ', '').lower()):
                print('old', Stiltsi_taburety.objects.get(pars_name=str(product['name']).replace(' ', '').lower()).name)
                model_id =  Stiltsi_taburety.objects.get(pars_name=str(product['name']).replace(' ', '').lower()).id
                size_id = Stiltsi_taburety_size.objects.filter(stiltsi_taburety_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Stiltsi_taburety.objects.all().order_by('-id')[0].id + 1

                Stiltsi_taburety.objects.get_or_create(id=new_id, manufacturer_id=22, name=product['name'],\
                    pars_name=str(product['name']).replace(' ', '').lower(), description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.get_or_create(stiltsi_taburety_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_stiltsi_taburety_richman():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=5).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book["Ціни"]

    prod_id = 0
    check_name = ''
    prom = ''
    cards = []
    size_cards = []

    for r in range(sheet.max_row):
        r += 1
        prod = re.search(str(sheet[r][1].value), 'Стілець')
        prod_2 = re.search(str(sheet[r][1].value), 'КаБаРе')
        name = sheet[r][2].value

        if sheet[r][4].value:
            mod = str(sheet[r][3].value) + str(sheet[r][4].value)
        else:
            mod = str(sheet[r][3].value)
            
        price = str(sheet[r][6].value)

        if prod and name or prod_2 and name:
            if check_name != name:
                check_name = name
                prod_id += 1
                prom = sheet[r][0].value
                
                #print(Stoly.objects.filter(pars_name = name, manufacturer_id=5).update(pars_name = prom))
                
                print(prod_id, ':', prom, '-->', name)

                cards.append({
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'des': '',
                })


            size_cards.append({
                'id': prod_id,
                'dop':mod,
                'w': 0,
                'h':0,
                'd': 0,
                'price': price,
            })
            print(' ', prod_id, '-->', mod, price)
    
              
    #Додаємо записи в базу
    new_db=[]
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Stiltsi_taburety.objects.filter(pars_name = product['prom']):
                print('old', Stiltsi_taburety.objects.get(pars_name=product['prom']).name)
                model_id =  Stiltsi_taburety.objects.get(pars_name=product['prom']).id
                new_db.append(product['prom'])
                index = 0

                for s in size_cards:
                    if s['id'] == product['id']:
                        print(s['price'])
                        size_id = Stiltsi_taburety_size.objects.filter(stiltsi_taburety_id=model_id).values()[index]['id']
                        Stiltsi_taburety_size.objects.filter(id = size_id).update(price=s['price'])
                        index += 1
                        
                
            #Додаємо новий товар
            else:
                new_id = Stiltsi_taburety.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Stiltsi_taburety.objects.get_or_create(id=new_id, manufacturer_id=5, name=product['name'],\
                    pars_name=product['prom'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.get_or_create(stiltsi_taburety_id=new_id,\
                            dop=s['dop'], w=s['w'], d=s['d'], price=s['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for m in Stiltsi_taburety.objects.filter(manufacturer_id=5):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def get_stiltsi_taburety_lion():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=7).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[1]
    cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)
        match = re.search('Табурети', cell)

        if match:
            while True:
                r += 1
                name = str(sheet[r][1].value)

                match = re.search("Пуфи", name)
                match_2 = re.search("1 категорія ЛДСП: венге магія, бетон, дуб молочний, дуб сонома, дуб сонома трюфель, німфея альба, антрацит, дуб аппалачі.", name)


                if match or match_2:
                    break

                elif name != 'None':

                    try:
                        name = " ".join(name.split())
                        size = str(sheet[r][3].value)
                        size = size.split('х')
                        try:
                            price = round(int(sheet[r][5].value))
                        except:
                            price = round(int(sheet[r][8].value))


                        prod_id += 1

                        cards.append({
                            'id': prod_id,
                            'name': name,
                            'des': '',
                        })

                        size_cards.append({
                            'id': prod_id,
                            'w': size[0],
                            'd': size[1],
                            'price': price,
                        })

                        print(prod_id, name.replace('  ', ''), size, price)
                    except Exception as ex:
                        print(ex)
                        break


    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Stiltsi_taburety.objects.filter(pars_name = product['name']):
                print('old', Stiltsi_taburety.objects.get(pars_name=product['name']).name)
                model_id =  Stiltsi_taburety.objects.get(pars_name=product['name']).id
                size_id = Stiltsi_taburety_size.objects.filter(stiltsi_taburety_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Stiltsi_taburety.objects.all().order_by('-id')[0].id + 1

                Stiltsi_taburety.objects.get_or_create(id=new_id, manufacturer_id=6, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.get_or_create(stiltsi_taburety_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_stiltsi_taburety_mixmebli():
    img_cards = []
    cards = []
    size_cards = []
    url = 'https://baustoff.com.ua/data/yml_mixmebli.xml'
    response = requests.get(url)
    categoryId = ['72','74', '76', '77', '78', '80', '81', '82', '83', '84', '85', '87', '88', '89']


    print('response ', response.status_code)

    if response.status_code == 200:
        root = ET.fromstring(response.content)

        for offer in root.find('.//shop/offers').iter('offer'):
            if offer.find('categoryId').text in categoryId and offer.attrib.get('available') == 'true':
                
                offer_id = offer.get('id')
                print(f"Offer ID: {offer_id}")

                try:
                    description = offer.find('description').text
                except:
                    description = ''
                    
                params_html = "<table border='1'>"

                width = ''
                depth = ''
                height =''

                for param in offer.iter('param'):
                    if param.attrib.get('name') == 'Ширина':
                        width = param.text

                    if param.attrib.get('name') == 'Глибина':
                        depth = param.text
                    
                    if param.attrib.get('name') == 'Висота':
                        height = param.text

                    params_html += f"<tr><td>{param.attrib.get('name')}</td><td>{param.text}</td></tr>"
                
                params_html += "</table><br>"

                cards.append({
                    'id': offer_id,
                    'prom':offer_id,
                    'name': offer.find('name').text,
                    'des': description + params_html,
                })

                size_cards.append({
                    'id': offer_id,
                    'w':width,
                    'd':depth,
                    'price':str(offer.find('price').text).replace('.00', '')
                })

                for i in offer.findall('picture'):
                    img = str(i.text)
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': offer_id,
                        'img': img_name
                    })
                
                print('------------------------------------')


    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Stiltsi_taburety.objects.filter(pars_name = product['prom']):
                print('old', Stiltsi_taburety.objects.get(pars_name=product['prom']).name)
                model_id =  Stiltsi_taburety.objects.get(pars_name=product['prom']).id
                size_id = Stiltsi_taburety_size.objects.filter(stiltsi_taburety_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Stiltsi_taburety.objects.all().order_by('-id')[0].id + 1

                Stiltsi_taburety.objects.get_or_create(id=new_id, manufacturer_id=27, name=product['name'],\
                    pars_name=product['prom'], description=product['des'],)
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Stiltsi_taburety_size.objects.get_or_create(stiltsi_taburety_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Stiltsi_taburety_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')
