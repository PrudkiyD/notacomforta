import requests
import xml.etree.ElementTree as ET
import openpyxl
from bs4 import BeautifulSoup
import csv
import time
import re
import os
import json
from lizhka.models import Bed, Bed_size, Bed_img
from pars.models import File
from django.db.models import Q


with open('/home/ay507291/notacomforta.pl.ua/www/pars/src.json', 'r') as f:
    file = json.load(f)


path_dirver = file['src'][0]['path_dirver']
accept = file['src'][1]['accept']
user_agent = file['src'][2]['user-agent']

HEADERS = {
    'accept': accept,
    'user-agent': user_agent
}

def num_check(text):
    num = ''
    for t in str(text):
        if t.isdigit():
            num += t
    return num


def get_lizhka_products_matrolux():
    img_cards = []
    cards = []
    size_cards = []

    print('Start ...')
    req = requests.get('https://matroluxe.ua/index.php?route=extension/feed/ocext_feed_generator_google&token=4171&categoryview=1', headers=HEADERS)  
    src = req.text
    print('Get src ---///')
    soup = BeautifulSoup(src, 'xml')
    item = soup.find_all('entry')

    for i in item:
        try: 
            title = i.find('title').get_text() if i.find('title') else None
            product_type = i.find('product_type').get_text() if i.find('product_type') else None
            product_id = i.find('id').get_text() if i.find('id') else None
            price = str(i.find('price').get_text()).replace('.00 UAH', '') if i.find('price') else None
            description = str(i.find('description').get_text()).replace('характеристики новых шкафов: ', '') if i.find('description') else ' '
            additional_image_link = i.find('additional_image_link').get_text() if i.find('additional_image_link') else None
            image_link = i.find('image_link').get_text() if i.find('image_link') else None
            img_list = [image_link, additional_image_link]
            brand = i.find('brand').get_text() if i.find('brand') else 29

            product_url = i.find('link').get_text() if i.find('brand') else 'link'

            sale = False
            old_price = None

            if i.find('sale_price'):
                sale = True
                old_price = price
                price = str(i.find('sale_price').get_text()).replace('.00 UAH', '') if i.find('sale_price') else None
            
            
            
            if  product_type == "Ліжка":
                if len(str(product_id).split('-')) == 1 or len(str(product_id).split('-')) == 2:
                    print(title)
                    print('-'*50)
                    for cat_num in range(50):
                        karkas = False
                        derevo = False
                        metal = False
                        myaki = False
                        dspmdf = False
                        box = False
                        up = False

                        if cat_num == 0:
                            cat = i.find(f'category').get_text() if i.find(f'category') else None
                        
                        else:
                            cat = i.find(f'category{cat_num}').get_text() if i.find(f'category{cat_num}') else None

                        if cat == 'Каркаси ':
                            karkas = True
                        
                        if cat == "Дерев'яні":
                            derevo = True

                        if cat == "Металеві":
                            metal = True

                        if cat == "Подіуми":
                            myaki = True

                        if cat == "З ДСП" or cat == "Із ДСП і МДФ":
                            dspmdf = True

                        if cat == "З шухлядками":
                            box = True

                        if up == "З підйомним механізмом":
                            box = True

                    manu = 29

                    if brand == "Matroluxe":
                        manu = 3
                    
                    if brand == "Sofyno":
                        manu = 29

                    cards.append({
                        'prom':product_id,
                        'id': product_id,
                        'name': title,
                        'des':description,
                        'munu':manu,
                        'sale':sale,
                        'url':product_url
                    })

                    size_cards.append({
                        'id': product_id,
                        'w':'',
                        'h':'',
                        'd':'',
                        'price':price,
                        'old_price':old_price,
                        'karkas':karkas,
                        'derevo':derevo,
                        'metal':metal,
                        'myaki':myaki,
                        'dspmdf':dspmdf,
                        'box':box,
                        'up':up

                    })

                    
                    for img in img_list:
                        print(f"img---{img}---///")
                        try:
                            img_name = img.split('/')[len(img.split('/')) - 1]
                            img_bytes = requests.get(img, headers=HEADERS).content
                            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                                f.write(img_bytes)

                            img_cards.append({
                                'id': product_id,
                                'img': img_name
                            })

                        except Exception as ex:
                            print(f'///-------{ex}---------///')

                            
                    

        except Exception as ex:
            print('-'*50)
            print(ex)
            print('-'*50)



    #Додаємо записи в базу
    new_db=[]
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name=product['prom'], manufacturer_id__in=[3, 29]).exists():
                print('old', Bed.objects.get(pars_name=product['prom']).name)
                model_id =  Bed.objects.get(pars_name=product['prom']).id
                Bed.objects.filter(id = model_id).update(sale = product['sale'], manufacturer_id=product['munu'])
                size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']

                new_db.append(product['prom'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.filter(id = size_id).update(price=s['price'], old_price=s['old_price'])
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])

                Bed.objects.get_or_create(id=new_id, manufacturer_id=product['munu'], name=product['name'],\
                    pars_name=product['prom'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'],\
                                karkas=s['karkas'], derevo=s['derevo'], metal=s['metal'],\
                                    myaki=s['myaki'], dspmdf=s['dspmdf'], box=s['box'],\
                                        up=s['up'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Bed_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')
    
    #Видаляємо товар
    for m in Bed.objects.filter(Q(manufacturer_id = 3) | Q(manufacturer_id = 29)):
        print(m)
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()



def get_lizhka_product_arbordrev():
    cards = []
    img_cards = []
    size_cards = []

    material = ''
    gear = ''
    size = ''
    
    pag_list = ['https://arbordrev.com.ua/beds',]
    links_list = []

    #Забераємо пагенацію

    for num in range(100):
        try:
            source = requests.get(pag_list[num],headers=HEADERS).text
            soup = BeautifulSoup(source, 'html.parser')
            pagination = soup.find('div', class_='wd-loop-footer products-footer').find('a').get('href')

            if pagination not in pag_list:
                pag_list.append(pagination)

        except:
            break

    #Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        links = soup.find_all('h3', class_='wd-entities-title')

        for i in links:
            link = i.find('a').get('href')
            if link not in links_list:
                links_list.append({
                    'num': len(links_list),
                    'url': link
                })
                print(len(links_list)-1, link)


    for url in links_list:
        material = ''
        gear = ''
        size = ''
        source = requests.get(url['url'], headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        name = soup.find('h1').get_text()
        price = soup.find('div', class_='col-lg-6 col-12 col-md-6 wd-price-outside summary entry-summary').find('bdi').getText()
        desc_text = soup.find('div', class_='wc-tab-inner').get_text()
        chek_list = soup.find('table', class_='variations').find('tbody').find_all('tr')
        img_list = soup.find('div', class_='product-images-inner').find_all('img')

        print(url['num'], name, url['url'])

        cards.append({
            'id': url['num'],
            'name': name,
            'des': desc_text,
        })

        

        for img in img_list:
            img = img.get('data-large_image')
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img,headers=HEADERS).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': url['num'],
                'img': img
            })


        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Порода дерева', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    active = a.get('aria-checked')

                    if active == 'true':
                        material = a.get_text()


        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Підйомний механізм', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    active = a.get('aria-checked')

                    if active == 'true':
                        gear = a.get_text()


        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Доступні розміри', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    active = a.get('aria-checked')

                    if active == 'true':
                        size = a.get_text()
                        size = size.split('х')

        try:
            size_cards.append({
                'id': url['num'],
                'option': material,
                'gear': gear,
                'w': size[0],
                'h': '',
                'd': size[1],
                'price': price,
            })
        
        except:
            size_cards.append({
                'id': url['num'],
                'option': material,
                'gear': gear,
                'w': size[0],
                'h': '',
                'd': '',
                'price': num_check(price),
            })



        print(material, gear, size, price)
                    
        print('-----------------------------')

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name = product['name']):
                print(Bed.objects.get(pars_name=product['name']).name)
                model_id =  Bed.objects.get(pars_name=product['name']).id
                size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']
                index = 0

                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])
                        index += 1
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1

                Bed.objects.get_or_create(id=new_id, manufacturer_id=14, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Bed_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')
    

def get_lizhka_product_everest():
    source = requests.get('https://everestm.com.ua/krovaty/', headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')
    pag_list = ['https://everestm.com.ua/krovaty/',]
    prod_link = []
    time.sleep(1)

    cards = []
    img_cards = []
    size_cards = []
    color_cards = []

    img_list = []
    size = []

    # Забераємо пагенацію

    pagination = soup.find('ul', class_='pagination').find_all('a')
    for p in pagination:
        pag = p.get('href')

        if pag not in pag_list:
            pag_list.append(pag)

        time.sleep(1)

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url, headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        links = soup.find_all('a', class_='product-name')

        for i in links:
            link = i.get('href')
            match = re.search('Ліжко', i.get_text())
            if link not in prod_link and match:
                prod_link.append({
                    'num': len(prod_link),
                    'url': link
                })

        time.sleep(1)

    # Забераємо інформацію про товар

    for url in prod_link:

        source = requests.get(url['url'], headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        name = soup.find('h1').get_text()
        price = soup.find('span', class_='autocalc-product-special').get_text()
        desc_text = soup.find('div', class_='nav-desc active').get_text()
        img_list = soup.find('div', class_='carousel slide').find_all('a')
        color_list = soup.find('div', class_='form-group').find_all('img')
        size_list = name.split(' ')
        size = size_list[len(size_list) - 1]
        size = size.split('х')

        print(url['num'], url['url'], name)
        print(size)

        for i in img_list:
            img = i.get('href')
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_name = str(img_name).replace('%', '')
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': url['num'],
                'img': img_name
            })


        for c in color_list:
            color = c.get('src')
            img_name = color.split('/')[len(color.split('/')) - 1]
            img_bytes = requests.get(color).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                f.write(img_bytes)

            color_cards.append({
                'id': url['num'],
                'img': img_name
            })


        cards.append({
            'id': url['num'],
            'name': name,
            'des': desc_text,
        })

        try:
            size_cards.append({
                'id': url['num'],
                'option': '',
                'gear': '',
                'w': size[0],
                'h': '',
                'd': size[1],
                'price': price,
            })

        except:
            size_cards.append({
                'id': url['num'],
                'option': '',
                'gear': '',
                'w': '',
                'h': '',
                'd': '',
                'price': price,
            })

        time.sleep(1)

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name = product['name']):
                print('old', Bed.objects.get(pars_name=product['name']).name)
                '''model_id =  Bed.objects.get(pars_name=product['name']).id
                size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])'''
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1

                Bed.objects.get_or_create(id=new_id, manufacturer_id=9, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price='0')

                for i in img_cards:
                    if i['id'] == product['id']:
                        Bed_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')
    

def get_lizhka_product_richman():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=5).file_up.url
    book = openpyxl.load_workbook(filename=path)
    sheet = book["Ціни"]

    prod_id = 0
    check_name = ''
    prom = ''
    cards = []
    size_cards = []

    for r in range(sheet.max_row):
        r += 1
        prod = re.search(str(sheet[r][1].value), 'Ліжко')
        name = sheet[r][2].value
        if sheet[r][4].value:
            mod = str(sheet[r][3].value) + str(sheet[r][4].value)
        else:
            mod = str(sheet[r][3].value)
            
        price = str(sheet[r][6].value)

        if prod and name:
            if check_name != name:
                check_name = name
                prod_id += 1
                prom = sheet[r][0].value
                
                #print(Bed.objects.filter(pars_name = name, manufacturer_id=5).update(pars_name = prom))
                cards.append({
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'des': '',
                })
                print(prod_id, ':', prom, '-->', name)


            size_cards.append({
                'id':prod_id,
                'option': '',
                'gear': '',
                'w': 0,
                'h': '',
                'd': 0,
                'price': price,
            })
            print(' ', prod_id, '-->', mod, price)


    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name = product['prom'], manufacturer_id=5):
                print('old', Bed.objects.get(pars_name=product['prom']).name)
                model_id =  Bed.objects.get(pars_name=product['prom']).id
                index = 0

                for s in size_cards:
                    if s['id'] == product['id']:
                        size_id = Bed_size.objects.filter(bed_id=model_id).values()[index]['id']
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])
                        index += 1
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1

                Bed.objects.get_or_create(id=new_id, manufacturer_id=5, name=product['name'],\
                    pars_name=product['prom'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')
    

def get_lizhka_product_svitmebliv():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=6).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[5]

    cards = []
    size_cards = []
    prod_id = 0

    for c in range(sheet.max_column):
        for r in range(sheet.max_row):
            r += 1
            cell = sheet[r][c].value
            match = re.search('М`яке ліжко', str(cell))

            if match:
                prod_id += 1
                print('-----------------////---------------------')

                name = f"{str(sheet[r][c].value)} {str(sheet[r+1][c].value)}"

                cards.append({
                    'id': prod_id,
                    'name': name,
                    'des': '',
                })

                
                print(name)
                row = r+1
                count = 0

                while True:
                    row += 1
                    match = re.search('М`яке ліжко', str(sheet[row][c].value))

                    if match or count == 100:
                        break

                    if sheet[row][c].value:

                        name = sheet[row][c].value
                        price = str(sheet[row][c+4].value)


                        size_cards.append({
                            'id': prod_id,
                            'option': name,
                            'gear': '',
                            'w': '',
                            'h': '',
                            'd': '',
                            'price': price,
                        })

                        print("-->", name, ':', price)
                    else:
                        count += 1
                
                print('-----------------////---------------------')

                
 

    sheet = book.worksheets[3]
    for r in range(sheet.max_row):
        r += 1
        cell = sheet[r][0].value
        match = re.search('Каркаси до ліжок', str(cell))

        if match:
            while True:
                r += 1
                if sheet[r][0].value:
                    prod_id += 1
                    name = str(sheet[r][0].value)
                    price = str(sheet[r][3].value)


                    cards.append({
                        'id': prod_id,
                        'name': name,
                        'des': '',
                    })

                    size_cards.append({
                        'id': prod_id,
                        'option': '',
                        'gear': '',
                        'w': '',
                        'h': '',
                        'd': '',
                        'price': price,
                    })
                    print(prod_id, name, price)
                
                else:
                    break


    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name = product['name']):
                print('old', Bed.objects.get(pars_name=product['name']).name)
                model_id =  Bed.objects.get(pars_name=product['name']).id
                size_id = Bed_size.objects.filter(bed_id=model_id)
                counter = 0

                for s in size_cards:
                    if s['id'] == product['id']:
                        print(s['price'])
                        Bed_size.objects.filter(id = size_id.values()[counter]['id']).update(price=s['price'])
                        counter += 1
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1

                Bed.objects.get_or_create(id=new_id, manufacturer_id=2, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_lizhka_product_lion():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=7).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[2]
    cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)
        match = re.search(' Ліжка 2х ярусні', cell)

        if match:
            while True:
                r += 1
                name = str(sheet[r][1].value)

                match = re.search('Стрази', name)

                if match:
                    break

                elif name != 'None':
                    try:
                        name = " ".join(name.split())
                        size = str(sheet[r][4].value)
                        size = size.split('х')
                        price = round(int(sheet[r][5].value))
                        prod_id += 1

                        cards.append({
                            'id': prod_id,
                            'name': name,
                            'des': '',
                        })

                        size_cards.append({
                            'id': prod_id,
                            'option': '',
                            'gear': '',
                            'w': size[0],
                            'h': size[2],
                            'd': size[1],
                            'price': price,
                        })

                        print(name.replace('  ', ''), size, price)
                    except:
                        pass

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)
        match = re.search('Ліжка двоспальні', cell)

        if match:
            while True:
                r += 1
                name = str(sheet[r][1].value)

                match = re.search('Комоди', name)

                if match:
                    break

                elif name != 'None':
                    try:
                        name = " ".join(name.split())
                        size = str(sheet[r][4].value)
                        size = size.split('х')
                        price = round(int(sheet[r][5].value))
                        prod_id += 1

                        cards.append({
                            'id': prod_id,
                            'name': name,
                            'des': '',
                        })

                        size_cards.append({
                            'id': prod_id,
                            'option': '',
                            'gear': '',
                            'w': size[0],
                            'h': size[2],
                            'd': size[1],
                            'price': price,
                        })

                        print(name.replace('  ', ''), size, price)
                    except:
                        pass

    #Додаємо записи в базу
    new_db = []
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name = product['name']):
                print('old', Bed.objects.get(pars_name=product['name']).name)
                model_id =  Bed.objects.get(pars_name=product['name']).id
                size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']
                new_db.append(product['name'])
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['name'])
                Bed.objects.get_or_create(id=new_id, manufacturer_id=6, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for m in Bed.objects.filter(manufacturer_id=6):
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()


def get_lizhka_product_olimp():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=8).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    size_cards = []
    prod_id = 0
    option = ''
    count = 0

    for r in range(sheet.max_row):

        r += 1
        cell = str(sheet[r][0].value)
        match = re.search('Тумби і комоди', cell)

        if match:
            break

        try:
            name = str(sheet[r][0].value)
            price = round(int(sheet[r][3].value) * 1.4)
            size = str(sheet[r][1].value)
            size = size.split('*')

            if name == 'None':
                if size[0].isdigit():

                    size_cards.append({
                        'id': prod_id,
                        'option': '',
                        'gear': '',
                        'w': size[1],
                        'h': '',
                        'd': size[0],
                        'price': price,
                    })

                    print(prod_id, size, price)

                else:
                    option = size[0]
                    option_num = count
                    option_price = price

                    while True:

                        price = round(int(sheet[option_num][3].value) * 1.4) + option_price
                        size = str(sheet[option_num][1].value)
                        size = size.split('*')

                        if size[0].isdigit():

                            size_cards.append({
                                'id': prod_id,
                                'option': option,
                                'gear': '',
                                'w': size[1],
                                'h': '',
                                'd': size[0],
                                'price': price,
                            })

                            print(prod_id, size, option, price)
                        else:
                            break

                        option_num += 1

            else:
                prod_id += 1

                count = r

                cards.append({
                    'id': prod_id,
                    'name': name,
                    'des': '',
                })

                size_cards.append({
                    'id': prod_id,
                    'option': '',
                    'gear': '',
                    'w': size[1],
                    'h': '',
                    'd': size[0],
                    'price': price,
                })

                print(prod_id, name, size, price)

        except:
            pass

    #Додаємо записи в базу
    new_db = []
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name = product['name']):
                print('old', Bed.objects.get(pars_name=product['name']).name)
                model_id =  Bed.objects.get(pars_name=product['name']).id
                index = 0
                new_db.append(product['name'])

                for s in size_cards:
                    if s['id'] == product['id']:
                        size_id = Bed_size.objects.filter(bed_id=model_id).values()[index]['id']
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])
                        index += 1
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['name'])
                Bed.objects.get_or_create(id=new_id, manufacturer_id=13, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'])

                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for i in Bed.objects.filter(manufacturer_id=13):
        if i.pars_name and i.pars_name not in new_db:
            Bed.objects.get(id=i.id).delete()
            print(i.name)


def get_lizhka_product_neman():
    
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=9).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    size_cards = []
    img_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        size = []
        cell = sheet[r][8].value
        match = re.search('Ліжко', str(sheet[r][3].value))
        match2 = re.search('Тахта', str(sheet[r][3].value))

        if match or match2:
            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][3].value)
            name_ru = str(sheet[r][4].value)
            des = str(sheet[r][39].value)
            des_ru = str(sheet[r][40].value)
            size = str(sheet[r][64].value).split('x')
            img_list = str(sheet[r][15].value).split(';')
            img_list = [line.rstrip() for line in img_list]
            price = str(sheet[r][9].value)[:-3]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'name_ru': name_ru,
                'des': des,
                'des_ru': des_ru,
            })

            print(prod_id, name)
            print(size)

            print(price)
            

            try:
                size_cards.append({
                    'id': prod_id,
                    'option': '',
                    'gear': '',
                    'w': size[1],
                    'h': '',
                    'd': size[0],
                    'price': price,
                })
            except:
                size_cards.append({
                    'id': prod_id,
                    'option': '',
                    'gear': '',
                    'w': size[0],
                    'h': '',
                    'd': '',
                    'price': price,
                })

            for i in img_list:
                try:
                    img = i
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_name = str(img_name).replace('%', '')
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })
                    print(img_name)
                except:
                    pass
            
            print('-------------------------------')
            


    #Додаємо записи в базу
    new_db =[]
    for product in cards:
        print('------------------------------------------------')
        product['name']
        
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(manufacturer_id=10).filter(pars_name = product['prom']):
                print('old', Bed.objects.get(pars_name=product['prom']).name)
                model_id =  Bed.objects.get(pars_name=product['prom']).id
                size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']
                new_db.append(product['prom'])

                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Bed.objects.get_or_create(id=new_id, manufacturer_id=10,\
                    name=product['name'], description=product['des'],\
                    name_ru=product['name_ru'], description_ru=product['des_ru'],\
                          pars_name=product['prom'],)
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Bed_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])


        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for i in Bed.objects.filter(manufacturer_id=10):
        if i.pars_name and i.pars_name not in new_db:
            Bed.objects.get(id=i.id).delete()
            print(i.name)


def get_lizhka_product_tenero():
    source = requests.get('https://tenero.in.ua/google_merchant_center.xml?hash_tag=f849745a8bde8e71cf3a96764a271ef1&product_ids=&label_ids=&export_lang=ru&group_ids=', headers=HEADERS)
    source.encoding = 'utf-8'

    soup = BeautifulSoup(source.text, 'xml')


    item = soup.find_all('item')
    prod_link = []
    cards = []
    img_cards = []
    size_cards = []

    for i in item:
        type = i.find('g:product_type').get_text()
        img = i.find('g:image_link').get_text()
        match = re.search('кровати', type)
        link = i.find('g:link').get_text()
        link = link[:21] + 'ua/' + link[21:]
        link = str(link).split('?')[0]
        if match:
            prod_link.append({
                'num': len(prod_link),
                'url': link,
                'img': img,
                'title':i.find('g:title').get_text(),
                'price':i.find('g:price').get_text(),
            })

    for url in prod_link:
        source = requests.get(url['url'], headers=HEADERS).content
        soup = BeautifulSoup(source, 'html.parser')

        
        prom = url['url']
        print('-->', url['url'])

        name = url['title']
        price = str(url['price']).replace('.00 UAH', '')
        desc_text = soup.find('div', class_='UserContent__root--ZhQBm') #Може змінитися клас але завжди буде UserContent__root


        
        cards.append({
            'id': url['num'],
            'prom':prom,
            'name': name,
            'des': str(desc_text),
        })

        print(str(desc_text))

        size_cards.append({
            'id': url['num'],
            'option': '',
            'gear': '',
            'w': '',
            'h': '',
            'd': '',
            'price': price,
        })

        img = url['img']
        img_name = img.split('/')[len(img.split('/')) - 1]
        img_bytes = requests.get(img).content
        with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
            f.write(img_bytes)

        img_cards.append({
            'id': url['num'],
            'img': img_name
        })

        
        print(url['num'], name, price)

        
        
        print('-------------------------------------')

    
    #Додаємо записи в базу
    new_db = []
    for product in cards:
        print('------------------------------------------------')
        product['name']

        #Оновлюємо ціну
        if Bed.objects.filter(pars_name = product['prom']):
            print('old', Bed.objects.get(pars_name=product['prom']).name)
            model_id = Bed.objects.get(pars_name=product['prom']).id
            size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']
            new_db.append(product['prom'])
            Bed.objects.filter(pars_name = product['prom']).update(description = product['des'])

            for s in size_cards:
                if s['id'] == product['id']:
                    Bed_size.objects.filter(id = size_id).update(price=s['price'])
                    break
            
        #Додаємо новий товар
        else:
            new_id = Bed.objects.all().order_by('-id')[0].id + 1
            new_db.append(product['prom'])
            Bed.objects.create(id=new_id, manufacturer_id=12, name=product['name'],\
                pars_name=product['prom'], description = product['des'])
            
            for s in size_cards:
                if s['id'] == product['id']:
                    Bed_size.objects.create(bed_id=new_id, option=s['option'],\
                        gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'])

            for i in img_cards:
                if i['id'] == product['id']:
                    Bed_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


            print('new', new_id, product['name'])
    
    #Видаляємо товар
    for i in Bed.objects.filter(manufacturer_id=12):
        if i.pars_name and i.pars_name not in new_db:
            Bed.objects.get(id=i.id).delete()
            print(i.name)


def get_lizhka_product_estella():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=2).file_up.url
    print(path)
    with open(path, 'r', ) as f:
        src = f.read()

    soup = BeautifulSoup(src, 'xml')
    item = soup.find_all('offer')
    
    cards = []
    img_cards = []
    size_cards = []
    prod_id = 0

    item_list=[]
    
    for i in item:
        cell = str(i.find('name').getText()).replace(' ', '').lower()

        if cell not in item_list:
            item_list.append(cell)
            print(cell)

    for i in item_list:
        print('-->', i)
        for o in item:
            cell = str(o.find('name').getText()).replace(' ', '').lower()
            if i == cell:
                try:
                    print('     ', o.find('name').getText(), o.find('model').getText(), o.find('price').getText())
                except:
                    pass


   
    '''#Додаємо записи в базу
    new_db = []
    for product in cards:
        print('------------------------------------------------')
        new_db.append(product['name'])
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name = product['name']):
                print('old', Bed.objects.get(pars_name=product['name']).name)
                model_id =  Bed.objects.get(pars_name=product['name']).id
                size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']

                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1

                Bed.objects.get_or_create(id=new_id, manufacturer_id=11, name=product['name'],\
                    pars_name=product['name'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Bed_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')



    #Видаляємо товар
    for i in Bed.objects.filter(manufacturer_id=11):
        if i.pars_name and i.pars_name not in new_db:
            Bed.objects.get(id=i.id).delete()
            print(i.name)'''


def get_lizhka_kompanit():
    source = requests.get('https://kompanit.com.ua/catalog/lizhka/c6',headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')

    cards = []
    size_cards = []
    img_cards = []
    color_cards = []
    link_list = []
    prod_id = 0
    pag_list = ['https://kompanit.com.ua/catalog/lizhka/c6', ]

    # Забераємо пагенацію
    try:
        item = soup.find('div', class_='pagination').find_all('a')
        for i in item:
            if i.get('href') not in pag_list and i.get('href'):
                print('-->', i.get('href'))
                pag_list.append(i.get('href'))
        time.sleep(2)
    except:
        pass

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        item = soup.find_all('a', class_='product__name')
        for i in item:
            if i.get('href') not in link_list:
                link_list.append(i.get('href'))
                print(len(link_list), i.get('href'))

        time.sleep(2)

    # Забераємо інформацію про товар

    for i in link_list:
        print('----------------------------------')
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
    
        prom = i
        name = soup.find('h1').getText()
        desc_text = soup.find('div', class_='tabs _mb-sm').getText(strip=True)
        prop_list = soup.find('div', class_='tabs _mb-sm').find_all('li')
        size = ['', '']


        print(prom)
        print(name)

        for i in prop_list:
            val = i.getText()

            w_match = re.search('Ширина', val)
            d_match = re.search('Глибина', val)

            if w_match:
                size[0] = num_check(val)

            if d_match:
                size[1] = num_check(val)

        print(size)

        cards.append({
            'id': prod_id,
            'prom':prom,
            'name': name,
            'des': desc_text,
        })

        size_cards.append({
            'id': prod_id,
            'w': size[0],
            'd': size[1],
            'price': '',
        })


        try:
            img = soup.find('img', class_='js-lazyload slider__slide-img').get('data-zzload-source-img')
            print(img)
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_bytes = requests.get(img).content
            with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name.replace('%', ''), "wb") as f:
                f.write(img_bytes)

            img_cards.append({
                'id': prod_id,
                'img': img_name.replace('%', '')
            })
        except:
            pass


        prod_id += 1

    #Додаємо записи в базу
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name = product['prom']).filter(manufacturer_id=19):
                print('old', Bed.objects.get(pars_name=product['prom']).name)
                '''model_id =  Bed.objects.get(pars_name=product['prom']).id
                Bed.objects.filter(id = model_id).update(sale = product['sale'])
                size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])'''
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1

                Bed.objects.get_or_create(id=new_id, manufacturer_id=19, name=product['name'],\
                    pars_name=product['prom'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id,\
                            w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Bed_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')


def get_lizhka_mixmebli():
    img_cards = []
    cards = []
    size_cards = []
    url = 'https://baustoff.com.ua/data/yml_mixmebli.xml'
    response = requests.get(url)
    categoryId = ['91', '92', '93', '94', '95', '96', '97', '98', '99', '100']


    print('response ', response.status_code)

    if response.status_code == 200:
        root = ET.fromstring(response.content)

        for offer in root.find('.//shop/offers').iter('offer'):
            if offer.find('categoryId').text in categoryId and offer.attrib.get('available') == 'true':
                
                offer_id = offer.get('id')
                print(f"Offer ID: {offer_id}")

                try:
                    description = offer.find('description').text
                except:
                    description = ''
                    
                params_html = "<table border='1'>"

                width = ''
                depth = ''
                height =''

                for param in offer.iter('param'):
                    if param.attrib.get('name') == 'Ширина':
                        width = param.text

                    if param.attrib.get('name') == 'Довжина':
                        depth = param.text
                    
                    if param.attrib.get('name') == 'Висота':
                        height = param.text

                    params_html += f"<tr><td>{param.attrib.get('name')}</td><td>{param.text}</td></tr>"
                
                params_html += "</table><br>"

                cards.append({
                    'id': offer_id,
                    'prom':offer_id,
                    'name': offer.find('name').text,
                    'des': description + params_html,
                })

                myaki = False
                if offer.find('categoryId').text == '92':
                    myaki = True

                size_cards.append({
                    'id': offer_id,
                    'w':width,
                    'd':depth,
                    'myaki':myaki,
                    'price':str(offer.find('price').text).replace('.00', '')
                })

                for i in offer.findall('picture'):
                    img = str(i.text)
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': offer_id,
                        'img': img_name
                    })
                
                print('------------------------------------')

    #Додаємо записи в базу
    new_db =[]
    for product in cards:
        print('------------------------------------------------')
        product['name']
        
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(manufacturer_id=27).filter(pars_name = product['prom']):
                print('old', Bed.objects.get(pars_name=product['prom']).name)
                model_id =  Bed.objects.get(pars_name=product['prom']).id
                size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']
                new_db.append(product['prom'])

                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])
                Bed.objects.get_or_create(id=new_id, manufacturer_id=27,\
                    name=product['name'], description=product['des'],\
                          pars_name=product['prom'],)
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id,\
                                                       w=s['w'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Bed_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])


        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')

    #Видаляємо товар
    for i in Bed.objects.filter(manufacturer_id=27):
        if i.pars_name and i.pars_name not in new_db:
            Bed.objects.get(id=i.id).delete()
            print(i.name)


def  get_lizhka_komfortmebli():
    path ="/home/ay507291/notacomforta.pl.ua/www"+File.objects.get(id=19).file_up.url
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    img_cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][8].value).split(';')
            img_list.insert(0, str(sheet[r][6].value)) 

            des = str(sheet[r][9].value)


            cards.append({
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
            })

            size_cards.append({
                'id': prod_id,
                'w':'',
                'h':'',
                'd':'',
                'option':'',
                'gear':'',
                'price':price
            })

            print(prom, '-->', name, '->', price)

            for img in img_list:
                print(f"img---{img}---///")
                
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })

                except Exception as ex:
                    print(f'///-------{ex}---------///')
            
            print('-------------------------')

    #Додаємо записи в базу
    new_db=[]
    for product in cards:
        print('------------------------------------------------')
        product['name']
        try:
            #Оновлюємо ціну
            if Bed.objects.filter(pars_name = product['prom']):
                print('old', Bed.objects.get(pars_name=product['prom']).name)
                model_id =  Bed.objects.get(pars_name=product['prom']).id
                Bed.objects.filter(id = model_id).update(sale = product['sale'])
                size_id = Bed_size.objects.filter(bed_id=model_id).values()[0]['id']

                new_db.append(product['prom'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.filter(id = size_id).update(price=s['price'])
                
            #Додаємо новий товар
            else:
                new_id = Bed.objects.all().order_by('-id')[0].id + 1
                new_db.append(product['prom'])

                Bed.objects.get_or_create(id=new_id, manufacturer_id=1, name=product['name'],\
                    pars_name=product['prom'], description=product['des'])
                
                for s in size_cards:
                    if s['id'] == product['id']:
                        Bed_size.objects.get_or_create(bed_id=new_id, option=s['option'],\
                            gear=s['gear'], w=s['w'], h=s['h'], d=s['d'], price=s['price'])

                for i in img_cards:
                    if i['id'] == product['id']:
                        Bed_img.objects.get_or_create(key_img_id=new_id, img=i['img'])


                print('new', new_id, product['name'])

        except Exception as ex:
            print('///----------ERORR--------///')
            print(ex)
            print('///----------ERORR--------///')
    
    #Видаляємо товар
    for m in Bed.objects.filter(manufacturer_id = 1):
        print(m)
        if m.pars_name not in new_db:
            print('delet:', m)
            m.delete()