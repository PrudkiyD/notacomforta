from catalog.models import Product, ProductImage, ProductPrice, Category, Subcategory
from .Tools import HEADERS, change_category, change_category_modul, file_path, product_images_path, num_check
import xml.etree.ElementTree as ET
from update.models import File
from bs4 import BeautifulSoup
import requests
import openpyxl
import re
import logging

logger = logging.getLogger(__name__)


def get_stoly_neman():
    path =File.objects.get(id=24).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    size_cards = []
    img_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        size = []
        cell = str(sheet[r][10].value)

        match = re.search('Обідні', cell)
        match2 = re.search('Розсувні|Столи-трансформери|Столи лофт', cell)
        

        if r > 1 and not match or r > 1 and not match2:
            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][3].value)
            des = str(sheet[r][45].value)
            name_ru = str(sheet[r][5].value)
            des_ru = str(sheet[r][47].value)
            size = [None,None,None]
            img_list = str(sheet[r][17].value).split(';')
            img_list = [line.rstrip() for line in img_list]
            price = str(sheet[r][11].value)[:-3]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'des': des,
                'name_ru': name_ru,
                'des_ru': des_ru,
            })

            logger.info(prom, name)
            #logger.info(size)
            #logger.info(price)
            
            try:
                size_cards.append({
                    'id': prod_id,
                    'option': None,
                    'gear': None,
                    'w': size[1],
                    'h': None,
                    'd': size[0],
                    'price': price,
                })
            except:
                size_cards.append({
                    'id': prod_id,
                    'option': None,
                    'gear': None,
                    'w': size[0],
                    'h': None,
                    'd': None,
                    'price': price,
                })

            for img in img_list:                
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_cards.append({
                        'id': prod_id,
                        'img': img_name,
                        'url': img
                    })

                    #logger.info(img_name)

                except Exception as ex:
                    logger.info(f'///{ex}')
            
            logger.info('-'*50)


    #Оновлюємо ціни
    #Додаємо записи в базу

    stock = []

    for item in cards:
        category = Category.objects.get(id=12)
        manufacturer = 10
        external_category = 'get_stoly_neman'

        change_category(manufacturer, 'stoly', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                            is_main=main_price,
                            setup=s['option']
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                product.save()

                logger.info('new', item['name'])

                

                stock.append(product.id)

        except Exception as ex:
            
            logger.info(ex)
    
    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info( product.name)


def get_stoly_matrolux():
    img_cards = []
    cards = []
    size_cards = []

    logger.info('Start ...')
    req = requests.get(File.objects.get(id=4).url, headers=HEADERS)  
    src = req.text
    logger.info('Get src ---///')
    soup = BeautifulSoup(src, 'xml')
    item = soup.find_all('entry')

    for i in item:
        try: 
            title = i.find('title').get_text() if i.find('title') else None
            product_type = i.find('product_type').get_text() if i.find('product_type') else None
            product_id = i.find('id').get_text() if i.find('id') else None
            price = str(i.find('price').get_text()).replace('.00 UAH', '') if i.find('price') else None
            description = str(i.find('description').get_text()).replace('характеристики новых шкафов: ', '') if i.find('description') else ' '
            additional_image_link = i.find('additional_image_link').get_text() if i.find('additional_image_link') else None
            image_link = i.find('image_link').get_text() if i.find('image_link') else None
            img_list = [additional_image_link, image_link]

            sale = False
            old_price = None

            if i.find('sale_price'):
                old_price = price
                sale = True
                price = str(i.find('sale_price').get_text()).replace('.00 UAH', '') if i.find('sale_price') else None
            
            
            
            if  product_type == "Корпусні меблі":
                if len(str(product_id).split('-')) == 1:
                
                    update = False
                    subcategory = []

                    for cat_num in range(50):
                        if cat_num == 0:
                            cat = i.find(f'category').get_text() if i.find(f'category') else None
                        
                        else:
                            cat = i.find(f'category{cat_num}').get_text() if i.find(f'category{cat_num}') else None


                        if cat == 'Журнальні':
                            update = True
                            subcategory.append(34)

                        if cat == 'Столи-трансформери':
                            update = True
                            subcategory.append(31)


                    if update:

                        logger.info(title)
                        logger.info('-'*50)

                        cards.append({
                            'prom':product_id,
                            'id': product_id,
                            'name': title,
                            'des':description,
                            'subcategory':subcategory
                        })

                        size_cards.append({
                            'id': product_id,
                            'old_price':old_price,
                            'price':price,
                            'sale':sale,
                        })

                        
                        for img in img_list:
                            try:
                                img_name = img.split('/')[len(img.split('/')) - 1]
                                img_cards.append({
                                    'id': product_id,
                                    'img': img_name,
                                    'url': img
                                })

                                

                            except Exception as ex:
                                logger.info(f'///-------{ex}---------///')
                        

                    

        except Exception as ex:
            logger.info('-'*50)
            logger.info(ex)
            logger.info('-'*50)

    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=12)
        manufacturer = 3
        external_category = 'stoly'

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.old_price = s['old_price']
                        curent_price.sale = s['sale']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                for sub in item['subcategory']:
                    subcategory = Subcategory.objects.get(id=sub)
                    product.subcategory.add(subcategory)

                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            old_price=s['old_price'],
                            is_main=main_price,
                            sale=s['sale']
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                            logger.info(f"Завантажено: {i['url']}")
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                

                logger.info('new', item['name'])

                

        except Exception as ex:
            
            logger.info(ex)


def get_stoly_modul_lux():
    path = File.objects.get(id=27).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    prod_id = 0
    cards = []
    size_cards = []

    for r in range(sheet.max_row):
            
            r += 1
            cell = str(sheet[r][0].value)
            match = re.search('СТОЛИ СЕРВІРУВАЛЬНІ', cell)

            if match:
                r = r+3
                while True:
                    r += 1

                    cell = str(sheet[r][0].value)
                    name = str(sheet[r][1].value)
                    size = str(sheet[r][2].value)
                    try:
                        price = str(round(float(sheet[r][3].value)))
                    except:
                        price = ''

                    size = size.split('*')
                    match = re.search('СТОЛИ ОБІДНІ (СЛОДСЬКА ПРОГРАМА)', cell)

                    if match or len(size) == 1:
                        break

                    else:

                        cards.append({
                            'id': prod_id,
                            'prom':str(name).replace(' ', '').lower(),
                            'name': name,
                            'des': '',
                        })

                        size_cards.append({
                            'id': prod_id,
                            'w': size[0],
                            'd': size[1],
                            'price': price,
                        })

                        prod_id += 1

                        logger.info(prod_id, name, size, price)

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][0].value)
        match = re.search('СТОЛИ ОБІДНІ', cell)

        if match:
            r = r + 3
            while True:
                r += 1

                cell = str(sheet[r][0].value)
                name = str(sheet[r][1].value)
                size = str(sheet[r][2].value)
                try:
                    price = str(round(float(sheet[r][3].value)))
                except:
                    price = ''

                size = size.split('*')
                match = re.search('СТОЛИ-ТРЮМО', cell)

                if match:
                    break

                else:

                    cards.append({
                        'id': prod_id,
                        'prom':str(name).replace(' ', '').lower(),
                        'name': name,
                        'des': '',
                    })

                    size_cards.append({
                        'id': prod_id,
                        'w': size[0],
                        'd': size[1],
                        'price': price,
                    })

                    prod_id += 1

                    logger.info(prod_id, name, size, price)
    

    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=12)
        manufacturer = 22
        external_category = 'stoly'

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],
                )

                product.category.add(category)

                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                
                logger.info('new', item['name'])

                

        except Exception as ex:
            
            logger.info(ex)


def get_stoly_jam():
    cards = []
    size_cards = []
    img_cards = []
    prod_link = []
    pag_list = [File.objects.get(id=28).url, ]
    size = [None, None]

    url = File.objects.get(id=28).url 

    resp = requests.get(url)
    html = resp.text

    match = re.search(r'defaultHash\s*=\s*"([^"]+)"', html)

    if match:
        default_hash = match.group(1)
        print("defaultHash:", default_hash)
    else:
        print("Хеш не знайдено")

    cookies = {
        "challenge_passed": default_hash
        }

    # Забераємо пагенацію
    logger.info("Забераємо пагенацію")
    try:
        for p in pag_list:
            source = requests.get(p,headers=HEADERS, cookies=cookies).text
            soup = BeautifulSoup(source, 'html.parser')

            item = soup.find('div', class_='pagination-container').find_all('a')

            for i in item:
                if i.get('href'):
                    pag = File.objects.get(id=29).url + str(i.get('href'))

                    if pag not in pag_list:
                        pag_list.append(pag)
                        logger.info(pag)
    except Exception as ex:
        logger.info(ex)


    # Забераємо посилання на товар
    logger.info("Забераємо посилання на товар")
    for url in pag_list:
        logger.info(url)
        source = requests.get(url,headers=HEADERS, cookies=cookies).text
        soup = BeautifulSoup(source, 'html.parser')
        links = soup.find_all('li', class_="catalog-grid__item")

        for l in links:
            link = File.objects.get(id=29).url + l.find('a').get('href')

            if link not in prod_link:
                logger.info(link)
                prod_link.append({
                    'num': len(prod_link),
                    'url': link
                })


    # Забераємо інформацію про товар
    logger.info("Забераємо інформацію про товар")
    for url in prod_link:
        size = [None, None]
        prom = url['url']
        logger.info('->',url['url'])
        source = requests.get(url['url'],headers=HEADERS, cookies=cookies).text
        soup = BeautifulSoup(source, 'html.parser')


        name = soup.find('h1').getText()
        price = str(soup.find('div', class_='product__column--right').find('div',  class_="product-price__item").getText(strip=True))
        desc_text = soup.find('div', class_='text').getText()

        logger.info(name)
        logger.info(price)


        prop_list = soup.find('div', class_='product__column--right').find_all('div',  class_="modification")

        for p in prop_list:
            val_text = p.getText(strip=True)
    
            match = re.search('Довжина', val_text, re.IGNORECASE)
            match_2 = re.search('Ширина', val_text, re.IGNORECASE)

            if match:
                size[1] = num_check(val_text)
                
            else:
                size[1] = None

            if match_2:
                size[0] = num_check(val_text)

            else:
                size[1] = None
    
        cards.append({
            'id': url['num'],
            'prom':prom,
            'name': name,
            'des': desc_text,
        })
    
        size_cards.append({
            'id': url['num'],
            'w': size[0],
            'd': size[1],
            'price': num_check(price),
        })

        logger.info(num_check(price))
        logger.info(size)

        img_list = soup.find('div', class_='product__column--left').find_all('li', class_='gallery__thumb')

        for i in img_list:
            try:
                img = File.objects.get(id=29).url + str(i.find('a').get('data-href'))
                img_name = img.split('/')[len(img.split('/')) - 1]
                img_cards.append({
                            'id': url['num'],
                            'img': img_name.replace('%', ''),
                            'url': img
                        })

            except Exception as ex:
                logger.info(f'///-------{ex}---------///')


    #Додаємо записи в базу

    stock = []

    Product.objects.filter(external_category = 'stoly', manufacturer_id = 21).delete()

    for item in cards:
        category = Category.objects.get(id=12)
        manufacturer = 21
        external_category = 'get_stoly_jam'

        change_category(manufacturer, 'stoly', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]
                        curent_price.price = s['price']
                        curent_price.width = s['w']
                        curent_price.depth = s['d']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)
                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS, cookies=cookies).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                            logger.info(f"Завантажено: {i['url']}")
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()


                logger.info('new', item['name'])

                

                stock.append(product.id)

        except Exception as ex:
            
            logger.info(ex)
    
    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)
    logger.info('Видаляємо товар якого немає в наявності')
    logger.info(stock)
    if stock:
        for product in products:
            if product.id not in stock:
                product.delete()

                logger.info( product.name)