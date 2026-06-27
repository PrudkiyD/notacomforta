from catalog.models import Product, ProductImage, ProductPrice, Category, Subcategory
from .Tools import HEADERS, change_category, change_category_modul, file_path, product_images_path, num_check
import xml.etree.ElementTree as ET
from update.models import File
from bs4 import BeautifulSoup
import random
import string
import requests
import openpyxl
import time
import re
import logging

logger = logging.getLogger(__name__)



def get_lizhka_komfortmebli():
    path = File.objects.get(id=10).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    img_cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][12].value).split(';')
            img_list.insert(0, str(sheet[r][10].value)) 

            des = str(sheet[r][7].value)


            cards.append({
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
            })

            size_cards.append({
                'id': prod_id,
                'w':str(sheet[r][4].value),
                'h':str(sheet[r][5].value),
                'd':str(sheet[r][6].value),
                'option':'',
                'gear':'',
                'price':price
            })

            logger.info(prom, '-->', name, '->', price)

            for img in img_list:
                logger.info(f"img---{img}---///")
                
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_cards.append({
                        'id': prod_id,
                        'img': ''.join(random.choices(string.ascii_letters, k=7))+img_name,
                        'url': img
                    })

                except Exception as ex:
                    logger.info(f'///-------{ex}---------///')
            
            logger.info('-------------------------')

    #Додаємо записи в базу

    stock = []

    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = 1
        external_category = 'get_lizhka_komfortmebli'

        change_category(manufacturer, 'lizhka', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]
                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                product.save()

                logger.info('new', item['name'])

                

                stock.append(product.id)

        except Exception as ex:
            
            logger.info("Помилка при оновленню товара: ", ex)

    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info('Видалино: ', product.name)


def get_lizhka_products_matrolux():
    img_cards = []
    cards = []
    size_cards = []

    logger.info('Start ...')
    req = requests.get(File.objects.get(id=4).url, headers=HEADERS)  
    src = req.text
    logger.info('Get src ---///')
    soup = BeautifulSoup(src, 'xml')
    item = soup.find_all('entry')

    for i in item:
        try: 
            title = i.find('title').get_text() if i.find('title') else None
            product_type = i.find('product_type').get_text() if i.find('product_type') else None
            product_id = i.find('id').get_text() if i.find('id') else None
            price = str(i.find('price').get_text()).replace('.00 UAH', '') if i.find('price') else None
            description = str(i.find('description').get_text()).replace('характеристики новых шкафов: ', '') if i.find('description') else ' '
            additional_image_link = i.find('additional_image_link').get_text() if i.find('additional_image_link') else None
            image_link = i.find('image_link').get_text() if i.find('image_link') else None
            img_list = [image_link, additional_image_link]
            brand = i.find('brand').get_text() if i.find('brand') else 29

            product_url = i.find('link').get_text() if i.find('brand') else 'link'

            sale = False
            old_price = None
            subcategory = []

            if i.find('sale_price'):
                sale = True
                old_price = price
                price = str(i.find('sale_price').get_text()).replace('.00 UAH', '') if i.find('sale_price') else None
            
            
            
            if  product_type == "Ліжка":
                if len(str(product_id).split('-')) == 1 or len(str(product_id).split('-')) == 2:
                    logger.info(title)
                    logger.info('-'*50)
                    for cat_num in range(50):

                        if cat_num == 0:
                            cat = i.find(f'category').get_text() if i.find(f'category') else None
                        
                        else:
                            cat = i.find(f'category{cat_num}').get_text() if i.find(f'category{cat_num}') else None

                        if cat == 'Каркаси ':
                            subcategory.append(16)
                        
                        if cat == "Дерев'яні":
                            subcategory.append(10)

                        if cat == "Металеві":
                            subcategory.append(11)

                        if cat == "Подіуми":
                            subcategory.append(12)

                        if cat == "З ДСП" or cat == "Із ДСП і МДФ":
                            subcategory.append(13)

                        if cat == "З шухлядками":
                            subcategory.append(14)

                        if cat == "З підйомним механізмом":
                            subcategory.append(15)

                    manu = 29

                    if brand == "Matroluxe":
                        manu = 3
                    
                    if brand == "Sofyno":
                        manu = 29

                    cards.append({
                        'prom':product_id,
                        'id': product_id,
                        'name': title,
                        'des':description,
                        'munu':manu,
                        'url':product_url
                    })

                    size_cards.append({
                        'id': product_id,
                        'w':'',
                        'h':'',
                        'd':'',
                        'price':price,
                        'old_price':old_price,
                        'sale':sale,
                        'subcategory': subcategory if subcategory else None
                    })

                    
                    for img in img_list:
                        try:
                            img_name = img.split('/')[len(img.split('/')) - 1]
                            img_cards.append({
                                'id': product_id,
                                'img': img_name,
                                'url': img
                            })

                        except Exception as ex:
                            logger.info(f'///-------{ex}---------///')

                            
                    

        except Exception as ex:
            logger.info('-'*50)
            logger.info(ex)
            logger.info('-'*50)

    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = item['munu']
        external_category = 'lizhka'
        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.old_price = s['old_price']
                        curent_price.sale = s['sale']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            old_price=s['old_price'],
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                            is_main=main_price,
                            sale=s['sale']
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                product.save()

                logger.info('new', item['name'])

                

        except Exception as ex:
            
            logger.info("Помилка при оновленню товара: ", ex)
            

def get_lizhka_product_arbordrev():
    cards = []
    img_cards = []
    size_cards = []

    material = ''
    gear = ''
    size = ''
    
    pag_list = [File.objects.get(id=11).url,]
    links_list = []

    #Забераємо пагенацію

    for num in range(100):
        try:
            source = requests.get(pag_list[num],headers=HEADERS).text
            soup = BeautifulSoup(source, 'html.parser')
            pagination = soup.find('div', class_='wd-loop-footer products-footer').find('a').get('href')

            if pagination not in pag_list:
                pag_list.append(pagination)

        except:
            break

    #Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        links = soup.find_all('h3', class_='wd-entities-title')

        for i in links:
            link = i.find('a').get('href')
            if link not in links_list:
                links_list.append({
                    'num': len(links_list),
                    'url': link
                })
                logger.info(len(links_list)-1, link)


    for url in links_list:
        material = ''
        gear = ''
        size = ''
        source = requests.get(url['url'], headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        name = soup.find('h1').get_text()
        price = soup.find('div', class_='col-lg-6 col-12 col-md-6 wd-price-outside summary entry-summary').find('bdi').getText()
        desc_text = soup.find('div', class_='wc-tab-inner').get_text()
        chek_list = soup.find('table', class_='variations').find('tbody').find_all('tr')
        img_list = soup.find('div', class_='product-images-inner').find_all('img')

        logger.info(url['num'], name, url['url'])

        cards.append({
            'id': url['num'],
            'name': name,
            'prom': name,
            'des': desc_text,
        })

        

        for img in img_list:
            img = img.get('data-large_image')
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_cards.append({
                                'id': url['num'],
                                'img': img_name,
                                'url': img
                            })


        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Порода дерева', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    active = a.get('aria-checked')

                    if active == 'true':
                        material = a.get_text()


        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Підйомний механізм', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    active = a.get('aria-checked')

                    if active == 'true':
                        gear = a.get_text()


        for i in chek_list:
            text = i.find('label').get_text()
            chek = re.search('Доступні розміри', text)

            if chek:
                arria = i.find_all('li')

                for a in arria:
                    active = a.get('aria-checked')

                    if active == 'true':
                        size = a.get_text()
                        size = size.split('х')

        try:
            size_cards.append({
                'id': url['num'],
                'option': material,
                'gear': gear,
                'w': size[0],
                'h': '',
                'd': size[1],
                'price': price,
            })
        
        except:
            size_cards.append({
                'id': url['num'],
                'option': material,
                'gear': gear,
                'w': size[0],
                'h': '',
                'd': '',
                'price': num_check(price),
            })



        logger.info(material, gear, size, price)
                    
        logger.info('-----------------------------')

    
    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = 14
        external_category = 'lizhka'
        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = num_check(s['price'])
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            info=s['option'],
                            price=num_check(s['price']),
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:

                        images_product = ProductImage.objects.create(
                            product=product,
                            image="/media/product_images/" + str(i['img']),
                            is_main=main_img
                        )

                        img_bytes = requests.get(i['url'], headers=HEADERS).content
                        with open(product_images_path + i['img'], "wb") as f:
                            f.write(img_bytes)

                        main_img = False
                        images_product.save()

                product.save()

                logger.info('new', item['name'])

        except Exception as ex:
            
            logger.info(ex)
            

def get_lizhka_product_everest():
    source = requests.get(File.objects.get(id=12).url, headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')
    pag_list = [File.objects.get(id=12).url,]
    prod_link = []
    time.sleep(1)

    cards = []
    img_cards = []
    size_cards = []

    img_list = []
    size = []

    # Забераємо пагенацію

    pagination = soup.find('ul', class_='pagination').find_all('a')
    for p in pagination:
        pag = p.get('href')

        if pag not in pag_list:
            pag_list.append(pag)

        time.sleep(1)

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url, headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        links = soup.find_all('a', class_='product-name')

        for i in links:
            link = i.get('href')
            match = re.search('Ліжко', i.get_text())
            if link not in prod_link and match:
                prod_link.append({
                    'num': len(prod_link),
                    'url': link
                })

        time.sleep(1)

    # Забераємо інформацію про товар

    for url in prod_link:

        source = requests.get(url['url'], headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')

        name = soup.find('h1').get_text()
        price = soup.find('span', class_='autocalc-product-special').get_text()
        desc_text = soup.find('div', class_='nav-desc active').get_text()
        img_list = soup.find('div', class_='carousel slide').find_all('a')
        size_list = name.split(' ')
        size = size_list[len(size_list) - 1]
        size = size.split('х')

        logger.info(url['num'], url['url'], name)
        logger.info(size)

        for i in img_list:
            img = i.get('href')
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_name = str(img_name).replace('%', '')
            img_cards.append({
                                'id': url['num'],
                                'img': img_name,
                                'url': img
                            })


        cards.append({
            'id': url['num'],
            'name': name,
            'des': desc_text,
        })

        try:
            size_cards.append({
                'id': url['num'],
                'option': '',
                'gear': '',
                'w': size[0],
                'h': '',
                'd': size[1],
                'price': price,
            })

        except:
            size_cards.append({
                'id': url['num'],
                'option': '',
                'gear': '',
                'w': '',
                'h': '',
                'd': '',
                'price': price,
            })

        time.sleep(1)
    
    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = 9
        external_category = 'lizhka'
        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = num_check(s['price'])
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=num_check(s['price']),
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:

                        images_product = ProductImage.objects.create(
                            product=product,
                            image="/media/product_images/" + str(i['img']),
                            is_main=main_img
                        )

                        img_bytes = requests.get(i['url'], headers=HEADERS).content
                        with open(product_images_path + i['img'], "wb") as f:
                            f.write(img_bytes)

                        main_img = False
                        images_product.save()

                product.save()

                logger.info('new', item['name'])

        except Exception as ex:
            
            logger.info(ex)


def get_lizhka_product_svitmebliv():
    logger.info('Ліжка ...')
    path = File.objects.get(id=13).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    
    cards = []
    size_cards = []
    prod_id = 0


    sheet = book.worksheets[4]
    for c in range(sheet.max_column):
        for r in range(sheet.max_row):
            r += 1
            cell = sheet[r][c].value
            match = re.search('М`яке ліжко', str(cell))

            if match:
                prod_id += 1

                name = f"{str(sheet[r][c].value)} {str(sheet[r+1][c].value)}"

                cards.append({
                    'prom':str(name).replace(' ', '').lower(),
                    'id': prod_id,
                    'name': name,
                    'des': '',
                })

                row = r+1
                count = 0

                while True:
                    row += 1
                    match = re.search('М`яке ліжко', str(sheet[row][c].value))

                    if match or count == 100:
                        break

                    if sheet[row][c].value:

                        name = sheet[row][c].value
                        price = str(sheet[row][c+4].value)

                        size_cards.append({
                            'prom':name,
                            'id': prod_id,
                            'info': name,
                            'gear': None,
                            'w': None,
                            'h': None,
                            'd': None,
                            'price': price,
                        })

                    else:
                        count += 1

    sheet = book.worksheets[3]
    for r in range(sheet.max_row):
        r += 1
        cell = sheet[r][0].value
        match = re.search('Каркаси до ліжок', str(cell))

        if match:
            while True:
                r += 1
                if sheet[r][0].value:
                    prod_id += 1
                    name = str(sheet[r][0].value)
                    price = str(sheet[r][3].value)


                    cards.append({
                        'prom':str(name).replace(' ', '').lower(),
                        'id': prod_id,
                        'name': name,
                        'des': '',
                    })

                    size_cards.append({
                        'id': prod_id,
                        'info': None,
                        'gear': None,
                        'w': None,
                        'h': None,
                        'd': None,
                        'price': price,
                    })
                
                else:
                    break


    #Додаємо записи в базу

    stock = []

    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = 2
        external_category = 'get_lizhka_product_svitmebliv'

        change_category(manufacturer, 'lizhka', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        try:
                            curent_price = price[index]
                            curent_price.price = s['price']
                            curent_price.save()

                            index += 1
                        
                        except:
                            prace_product = ProductPrice.objects.create(
                                product=product.first(),
                                price=s['price'],
                                info=s['info'],
                            )

                            prace_product.save()


                logger.info('old', product[0].name)

                

                stock.append(product[0].id)
            
            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                product.save()

                logger.info('new', item['name'])

                

                stock.append(product.id)
                
        except Exception as ex:
            
            logger.info("Помилка при оновленню товара: ", ex)

    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info('Видалино: ', product.name)


def get_lizhka_product_lion():
    pass
    '''
    path =File.objects.get().files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[2]
    cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)
        match = re.search(' Ліжка 2х ярусні', cell)

        if match:
            while True:
                r += 1
                name = str(sheet[r][1].value)

                match = re.search('Стрази', name)

                if match:
                    break

                elif name != 'None':
                    try:
                        name = " ".join(name.split())
                        size = str(sheet[r][4].value)
                        size = size.split('х')
                        price = round(int(sheet[r][5].value))
                        prod_id += 1

                        cards.append({
                            'id': prod_id,
                            'name': name,
                            'des': '',
                        })

                        size_cards.append({
                            'id': prod_id,
                            'option': '',
                            'gear': '',
                            'w': size[0],
                            'h': size[2],
                            'd': size[1],
                            'price': price,
                        })

                        logger.info(name.replace('  ', ''), size, price)
                    except:
                        pass

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)
        match = re.search('Ліжка двоспальні', cell)

        if match:
            while True:
                r += 1
                name = str(sheet[r][1].value)

                match = re.search('Комоди', name)

                if match:
                    break

                elif name != 'None':
                    try:
                        name = " ".join(name.split())
                        size = str(sheet[r][4].value)
                        size = size.split('х')
                        price = round(int(sheet[r][5].value))
                        prod_id += 1

                        cards.append({
                            'id': prod_id,
                            'name': name,
                            'des': '',
                        })

                        size_cards.append({
                            'id': prod_id,
                            'option': '',
                            'gear': '',
                            'w': size[0],
                            'h': size[2],
                            'd': size[1],
                            'price': price,
                        })

                        logger.info(name.replace('  ', ''), size, price)
                    except:
                        pass'
    '''


def get_lizhka_product_olimp():
    path = File.objects.get(id=14).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    size_cards = []
    prod_id = 0
    option = ''
    count = 0

    for r in range(sheet.max_row):

        r += 1
        cell = str(sheet[r][0].value)
        match = re.search('Тумби і комоди', cell)

        if match:
            break

        try:
            name = str(sheet[r][0].value)
            price = round(int(sheet[r][3].value) * 1.4)
            size = str(sheet[r][1].value)
            size = size.split('*')

            if name == 'None':
                if size[0].isdigit():

                    size_cards.append({
                        'id': prod_id,
                        'option': '',
                        'gear': '',
                        'w': size[1],
                        'h': '',
                        'd': size[0],
                        'price': price,
                    })

                    logger.info(prod_id, size, price)

                else:
                    option = size[0]
                    option_num = count
                    option_price = price

                    while True:

                        price = round(int(sheet[option_num][3].value) * 1.4) + option_price
                        size = str(sheet[option_num][1].value)
                        size = size.split('*')

                        if size[0].isdigit():

                            size_cards.append({
                                'id': prod_id,
                                'option': option,
                                'gear': '',
                                'w': size[1],
                                'h': '',
                                'd': size[0],
                                'price': price,
                            })

                            logger.info(prod_id, size, option, price)
                        else:
                            break

                        option_num += 1

            else:
                prod_id += 1

                count = r

                cards.append({
                    'prom': name,
                    'id': prod_id,
                    'name': name,
                    'des': '',
                })

                size_cards.append({
                    'id': prod_id,
                    'option': '',
                    'gear': '',
                    'w': size[1],
                    'h': '',
                    'd': size[0],
                    'price': price,
                })

                logger.info(prod_id, name, size, price)

        except:
            pass

    
    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = 13
        external_category = 'lizhka'
        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = num_check(s['price'])
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=num_check(s['price']),
                            is_main=main_price,
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                        )

                        prace_product.save()
                        
                        main_price = False

                product.save()

                logger.info('new', item['name'])

        except Exception as ex:
            
            logger.info(ex)


def get_lizhka_product_neman():
    path = File.objects.get(id=15).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    size_cards = []
    img_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        size = []
        cell = sheet[r][8].value
        match = re.search('Ліжко', str(sheet[r][3].value))
        match2 = re.search('Тахта', str(sheet[r][3].value))

        if match or match2:
            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][3].value)
            name_ru = str(sheet[r][4].value)
            des = str(sheet[r][39].value)
            des_ru = str(sheet[r][40].value)
            size = str(sheet[r][64].value).split('x')
            img_list = str(sheet[r][15].value).split(';')
            img_list = [line.rstrip() for line in img_list]
            price = str(sheet[r][9].value)[:-3]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'name_ru': name_ru,
                'des': des,
                'des_ru': des_ru,
            })

            logger.info(prod_id, name)
            logger.info(size)

            logger.info(price)
            

            try:
                size_cards.append({
                    'id': prod_id,
                    'option': '',
                    'gear': '',
                    'w': size[1],
                    'h': '',
                    'd': size[0],
                    'price': price,
                })
            except:
                size_cards.append({
                    'id': prod_id,
                    'option': '',
                    'gear': '',
                    'w': size[0],
                    'h': '',
                    'd': '',
                    'price': price,
                })

            for i in img_list:
                try:
                    img = i
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_name = str(img_name).replace('%', '')
                    
                    img_cards.append({
                        'id': prod_id,
                        'img': img_name,
                        'url': img
                    })
                    logger.info(img_name)
                except:
                    pass
            
            logger.info('-------------------------------')


def get_lizhka_product_tenero():
    source = requests.get(File.objects.get(id=16).url, headers=HEADERS)
    source.encoding = 'utf-8'

    soup = BeautifulSoup(source.text, 'xml')


    item = soup.find_all('item')
    prod_link = []
    cards = []
    img_cards = []
    size_cards = []

    for i in item:
        type = i.find('g:product_type').get_text()
        img = i.find('g:image_link').get_text()
        match = re.search('кровати', type)
        link = i.find('g:link').get_text()
        link = link[:21] + 'ua/' + link[21:]
        link = str(link).split('?')[0]
        if match:
            prod_link.append({
                'num': len(prod_link),
                'url': link,
                'img': img,
                'title':i.find('g:title').get_text(),
                'price':i.find('g:price').get_text(),
            })

    for url in prod_link:
        source = requests.get(url['url'], headers=HEADERS).content
        soup = BeautifulSoup(source, 'html.parser')
    
        prom = url['url']
        logger.info('-->', url['url'])

        name = soup.find('h1').get_text() if soup.find('h1') else url['title']
        price = str(url['price']).replace('.00 UAH', '')
        desc_text = soup.find("div", attrs={"data-qaid": "product_description"})

        logger.info(name)
        
        cards.append({
            'id': url['num'],
            'prom':prom,
            'name': name,
            'des': str(desc_text),
        })

        size_cards.append({
            'id': url['num'],
            'option': '',
            'gear': '',
            'w': '',
            'h': '',
            'd': '',
            'price': price,
        })

        img = url['img']
        img_name = img.split('/')[len(img.split('/')) - 1]
        img_cards.append({
                                'id': url['num'],
                                'img': img_name,
                                'url': img
                            })

        
        logger.info(url['num'], name, price)

        
        
        logger.info('-------------------------------------')
    
    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = 12
        external_category = 'lizhka'

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():

                product=product.first()

                product.description = item['des']
                product.save()

                price = ProductPrice.objects.filter(product=product)
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],
                    description = item['des']
                )

                product.category.add(category)

                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False
                
                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                            logger.info(f"Завантажено: {i['url']}")
                        except Exception as ex:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )
                            logger.info(ex)

                        main_img = False
                        images_product.save()

                
                logger.info('new', item['name'])

                

        except Exception as ex:
            logger.info("Помилка при оновленні:")
            logger.info(item['name'])
            logger.info(ex)
    
    

def get_lizhka_product_estella():
    # Не оновлюємо
    pass


def get_lizhka_kompanit():
    src_url = File.objects.get(id=17).url
    source = requests.get(src_url,headers=HEADERS).text
    soup = BeautifulSoup(source, 'html.parser')

    cards = []
    size_cards = []
    img_cards = []
    color_cards = []
    link_list = []
    prod_id = 0
    pag_list = [src_url, ]

    # Забераємо пагенацію
    try:
        item = soup.find('div', class_='pagination').find_all('a')
        for i in item:
            if i.get('href') not in pag_list and i.get('href'):
                logger.info('-->', i.get('href'))
                pag_list.append(i.get('href'))
        time.sleep(2)
    except:
        pass

    # Забераємо посилання на товар

    for url in pag_list:
        source = requests.get(url,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
        item = soup.find_all('a', class_='product__name')
        for i in item:
            if i.get('href') not in link_list:
                link_list.append(i.get('href'))
                logger.info(len(link_list), i.get('href'))

        time.sleep(2)

    # Забераємо інформацію про товар

    for i in link_list:
        logger.info('----------------------------------')
        source = requests.get(i,headers=HEADERS).text
        soup = BeautifulSoup(source, 'html.parser')
    
        prom = i
        name = soup.find('h1').getText()
        desc_text = soup.find('div', class_='tabs _mb-sm').getText(strip=True)
        prop_list = soup.find('div', class_='tabs _mb-sm').find_all('li')
        size = ['', '']


        logger.info(prom)
        logger.info(name)

        for i in prop_list:
            val = i.getText()

            w_match = re.search('Ширина', val)
            d_match = re.search('Глибина', val)

            if w_match:
                size[0] = num_check(val)

            if d_match:
                size[1] = num_check(val)

        logger.info(size)

        cards.append({
            'id': prod_id,
            'prom':prom,
            'name': name,
            'des': desc_text,
        })

        size_cards.append({
            'id': prod_id,
            'w': size[0],
            'd': size[1],
            'price': '',
        })


        try:
            img = soup.find('img', class_='js-lazyload slider__slide-img').get('data-zzload-source-img')
            logger.info(img)
            img_name = img.split('/')[len(img.split('/')) - 1]
            img_cards.append({
                                'id': prod_id,
                                'img': img_name,
                                'url': img
                            })
        except:
            pass


        prod_id += 1

    #Додаємо записи в базу
    
    stock = []

    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = 19
        external_category = 'get_lizhka_kompanit'

        change_category(manufacturer, 'lizhka', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                logger.info('old', product[0].name)

                

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=0,
                            width=s['w'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                product.save()

                logger.info('new', item['name'])

                

                stock.append(product.id)

        except Exception as ex:
            
            logger.info("Помилка при оновленню товара: ", ex)

    
    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info('Видалино: ', product.name)


def get_lizhka_mixmebli():
    img_cards = []
    cards = []
    size_cards = []
    url = File.objects.get(id=9).url
    response = requests.get(url)
    categoryId = ['91', '92', '93', '94', '95', '96', '97', '98', '99', '100']


    logger.info('response ', response.status_code)

    if response.status_code == 200:
        root = ET.fromstring(response.content)

        for offer in root.find('.//shop/offers').iter('offer'):
            if offer.find('categoryId').text in categoryId and offer.attrib.get('available') == 'true':
                
                offer_id = offer.get('id')
                logger.info(f"Offer ID: {offer_id}")

                try:
                    description = offer.find('description').text
                except:
                    description = ''

                logger.info(description)
                    
                params_html = "<table border='1'>"

                width = ''
                depth = ''
                height =''

                for param in offer.iter('param'):
                    if param.attrib.get('name') == 'Ширина':
                        width = param.text

                    if param.attrib.get('name') == 'Довжина':
                        depth = param.text
                    
                    if param.attrib.get('name') == 'Висота':
                        height = param.text

                    params_html += f"<tr><td>{param.attrib.get('name')}</td><td>{param.text}</td></tr>"
                
                params_html += "</table><br>"

                cards.append({
                    'id': offer_id,
                    'prom':offer_id,
                    'name': offer.find('name').text,
                    'des': description + params_html,
                })

                myaki = False
                if offer.find('categoryId').text == '92':
                    myaki = True

                size_cards.append({
                    'id': offer_id,
                    'w':width,
                    'd':depth,
                    'myaki':myaki,
                    'price': str(int(float(offer.find('price').text)))
                })

                for i in offer.findall('picture'):
                    img = str(i.text)
                    try:
                        img_name = img.split('/')[len(img.split('/')) - 1]
                        img_cards.append({
                            'id': offer_id,
                            'img': img_name,
                            'url': img
                        })

                    except Exception as ex:
                        pass
                
                logger.info(f"{offer.find('name').text} {str(offer.find('price').text).replace('.00', '')}")
                logger.info(50 * '-')

    
    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = 27
        external_category = 'lizhka'

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],
                    description=item['des'],
                )

                product.category.add(category)

                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False
                
                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                            logger.info(f"Завантажено: {i['url']}")
                        except Exception as ex:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )
                            logger.info(ex)

                        main_img = False
                        images_product.save()

                
                logger.info('new', item['name'])

                

        except Exception as ex:
            
            logger.info(ex)
    



def get_lizhka_yudin():
    path = File.objects.get(id=25).files
    
    logger.info(path)

    book = openpyxl.load_workbook(filename=path)
    sheet_name = ["ЛІЖКА",]
    


    prod_id = 0
    prom = ''
    cards = []
    size_cards = []
    cell = ''

    for s in sheet_name:
        sheet = book[s]

        for r in range(sheet.max_row):
            r += 1

            check = cell
            cell = sheet[r][1].value

            if check == "Назва" and cell != "Назва":

                if cell:
            
                    name = str(sheet[r][1].value).replace('\n', ' ').replace('  ', '')
                    price_list = [
                        sheet[r][2].value,
                        sheet[r][3].value,
                        sheet[r][4].value,
                        sheet[r][5].value,
                        sheet[r][6].value,
                    ]

                    prom = name.replace(' ', '').replace('\t', '').lower()

                    cards.append({
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                        })
                    
                    logger.info((
                        f'{prod_id}: {prom}\n'
                        f'Назва: {name}\n'
                    ))
                        
                    
                    tkan = 0
                    
                    for price in price_list:

                        if price:
                            size_cards.append({
                                'id': prod_id,
                                'info': f"Категорія тканини: {tkan}" ,
                                'price': num_check(price)
                            })

                            tkan += 1
                    

                    prod_id += 1
                    cell = "Назва"


    #Додаємо записи в базу

    stock = []

    for item in cards:
        category = Category.objects.get(id=8)
        manufacturer = 28
        external_category = 'get_lizhka_yudin'

        change_category(manufacturer, 'lizhka', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]
                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

                stock.append(product[0].id)

            #Додаємо новий товар
            else:


                product = Product.objects.create(
                    published=False,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                )

                product.category.add(category)
                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            is_main=main_price,
                            info=s['info']
                        )

                        prace_product.save()
                        
                        main_price = False

                product.save()

                logger.info('new', item['name'])

                

                stock.append(product.id)

        except Exception as ex:
            
            logger.info(ex)

    
    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info('Видалино: ', product.name)