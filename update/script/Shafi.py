from catalog.models import Product, ProductImage, ProductPrice, Category, Subcategory, Seria
from .Tools import HEADERS, change_category, change_category_modul, file_path, product_images_path, num_check
import xml.etree.ElementTree as ET
from update.models import File, History
from bs4 import BeautifulSoup
import requests
import openpyxl
import re


def get_prod_komfortmebli():
    path = File.objects.get(id=3).files
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    img_cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][20].value).split(';')
            img_list.insert(0, str(sheet[r][18].value)) 

            des = (
                        f"<ul>\n"
                        f"  <li>Вартість шафи по розмірам: <br> {sheet[r][4].value}</li>\n"
                        f"  <li>Тип шафи: {sheet[r][5].value}</li>\n"
                        f"  <li>Форма шафи-купе: {sheet[r][6].value}</li>\n"
                        f"  <li>Кількість дверей, шт: {sheet[r][10].value}</li>\n"
                        f"  <li>Тип фасаду: {sheet[r][11].value}</li>\n"
                        f"  <li>Колір корпусу: {sheet[r][12].value}</li>\n"
                        f"  <li>Внутрішнє наповнення: {sheet[r][13].value}</li>\n"
                        f"  <li>Форма AL профілю: {sheet[r][14].value}</li>\n"
                        f"  <li>Колір AL профілю: {sheet[r][15].value}</li>\n"
                        f"  <li>Серія: {sheet[r][16].value}</li>\n"
                        f"  <li>Стиль: {sheet[r][17].value}</li>\n"
                        f"</ul>\n" 
                        f"<p>{sheet[r][21].value}</p>\n"
                    )


            cards.append({
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
            })

            size_cards.append({
                'id': prod_id,
                'w':sheet[r][7].value,
                'h':sheet[r][8].value,
                'd':sheet[r][9].value,
                'price':price
            })

            for img in img_list:
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_cards.append({
                        'id': prod_id,
                        'img': img_name,
                        'url': img
                    })

                except Exception as ex:
                    print(f'///-------{ex}---------///')


            print(prod_id,'-->', prom, ':', name, ':', price)
            print('-------------------------')


    #Додаємо записи в базу
    stock = []

    for item in cards:
        category = Category.objects.get(id=3)
        manufacturer = 1
        external_category = 'get_prod_komfortmebli'

        change_category(manufacturer, 'shafi', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]
                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                print('old', product[0].name)

                history = History.objects.create(
                            name=f"Оновлено товар",
                            description=product[0].name
                        )
                history.save()

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                product.save()

                print('new', item['name'])

                history = History.objects.create(
                            name=f"Додано товар",
                            description=item['name']
                        )
                history.save()

                stock.append(product.id)

        except Exception as ex:
            
            print("Помилка при оновленню товара: ", ex)

    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            history = History.objects.create(
                            name=f"Видалино товар",
                            description=product.name
                        )
            history.save()

            print('Видалино: ', product.name)
    


def get_rozpashni_shafi_komfortmebli():
    path = File.objects.get(id=30).files
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    img_cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1

        if r != 1:

            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][13].value).split(';')
            img_list.insert(0, str(sheet[r][12].value)) 

            des = (
                        f"<ul>\n"
                        f"  <li>Ширина, мм: <br> {sheet[r][4].value}</li>\n"
                        f"  <li>Тип фасаду: {sheet[r][5].value}</li>\n"
                        f"  <li>Колір корпусу: {sheet[r][6].value}</li>\n"
                        f"  <li>Стиль: {sheet[r][7].value}</li>\n"
                        f"  <li>Серія: {sheet[r][8].value}</li>\n"
                        f"  <li>Колір фасаду: {sheet[r][9].value}</li>\n"
                        f"  <li>Внутрішнє наповнення: {sheet[r][10].value}</li>\n"
                        f"</ul>" 
                    )


            cards.append({
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
            })

            size_cards.append({
                'id': prod_id,
                'w':sheet[r][4].value,
                'h':'',
                'd':'',
                'price':price
            })

            for img in img_list:
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_cards.append({
                        'id': prod_id,
                        'img': img_name,
                        'url': img
                    })

                except Exception as ex:
                    print(f'///-------{ex}---------///')


            print(prod_id,'-->', prom, ':', name, ':', price)
            print('-------------------------')

    
    #Додаємо записи в базу
    stock = []

    for item in cards:
        category = Category.objects.get(id=3)
        manufacturer = 1
        external_category = 'get_rozpashni_shafi_komfortmebli'

        change_category(manufacturer, 'shafi', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]
                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                print('old', product[0].name)

                history = History.objects.create(
                            name=f"Оновлено товар",
                            description=product[0].name
                        )
                history.save()

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)
                product.subcategory.add(Subcategory.objects.filter(id=3).first())

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                product.save()

                print('new', item['name'])

                history = History.objects.create(
                            name=f"Додано товар",
                            description=item['name']
                        )
                history.save()

                stock.append(product.id)

        except Exception as ex:
            
            print("Помилка при оновленню товара: ", ex)

    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            history = History.objects.create(
                            name=f"Видалино товар",
                            description=product.name
                        )
            history.save()

            print('Видалино: ', product.name)


def get_link_matrolux():
    img_cards = []
    cards = []
    size_cards = []

    print('Start ...')
    req = requests.get(File.objects.get(id=4).url, headers=HEADERS)  
    src = req.text

    print(src)

    print('Get src ---///')
    soup = BeautifulSoup(src, 'xml')
    item = soup.find_all('entry')

    for i in item:
        try: 
            title = i.find('title').get_text() if i.find('title') else None
            product_type = i.find('product_type').get_text() if i.find('product_type') else None
            product_id = i.find('id').get_text() if i.find('id') else None
            price = str(i.find('price').get_text()).replace('.00 UAH', '') if i.find('price') else None
            description = str(i.find('description').get_text()).replace('характеристики новых шкафов: ', '') if i.find('description') else ' '
            additional_image_link = i.find('additional_image_link').get_text() if i.find('additional_image_link') else None
            image_link = i.find('image_link').get_text() if i.find('image_link') else None
            img_list = [image_link, additional_image_link]
            old_price = None
            sale = False

            if i.find('sale_price'):
                old_price = price
                sale = True
                price = str(i.find('sale_price').get_text()).replace('.00 UAH', '') if i.find('sale_price') else None
            
            
            
            if  product_type == "Шафи" or product_type == "Шафи-купе":
                if len(str(product_id).split('-')) == 1:
                    print(title)
                    print('-'*50)
                    subcategory=[]
                    for cat_num in range(50):
                        if cat_num == 0:
                            cat = i.find(f'category').get_text() if i.find(f'category') else None
                        
                        else:
                            cat = i.find(f'category{cat_num}').get_text() if i.find(f'category{cat_num}') else None

                        if cat == 'Шафи-купе':
                            subcategory.append(2)
                            print('KYPE')
                            break

                        if cat == 'Розпашні шафи':
                            subcategory.append(3)
                            print('SHAFI')
                            break

                        if cat == 'Шафа-пенал':
                            subcategory.append(4)
                            print('PENAL')
                            break


                    cards.append({
                        'prom': product_id,
                        'id': product_id,
                        'name': title,
                        'des': description,
                        'subcategory':subcategory,
                    })

                    size_cards.append({
                        'id': product_id,
                        'price':price,
                        'old_price':old_price,
                        'sale':sale
                    })


                    for img in img_list:
                        try:
                            img_name = img.split('/')[len(img.split('/')) - 1]
                            img_cards.append({
                                'id': product_id,
                                'img': img_name,
                                'url': img
                            })

                            

                        except Exception as ex:
                            print(f'///-------{ex}---------///')

        except Exception as ex:
            print('-'*50)
            print(ex)
            print('-'*50)
    
    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=3)
        manufacturer = 3
        external_category = 'shafi'

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.old_price = s['old_price']
                        curent_price.sale = s['sale']
                        curent_price.save()

                        index += 1

                print('old', product[0].name)

                history = History.objects.create(
                            name=f"Оновлено товар",
                            description=product[0].name
                        )
                history.save()

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                for sub in item['subcategory']:
                    subcategory = Subcategory.objects.get(id=sub)
                    product.subcategory.add(subcategory)

                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            old_price=s['old_price'],
                            is_main=main_price,
                            sale=s['sale']
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                            print(f"Завантажено: {i['url']}")
                        except Exception as ex:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )
                            print(ex)

                        main_img = False
                        images_product.save()

                

                print('new', item['name'])

                history = History.objects.create(
                            name=f"Додано товар",
                            description=item['name']
                        )
                history.save()

        except Exception as ex:
            
            print(ex)
            

def get_products_fenix():
    print(File.objects.get(id=5).files)
    id = 0
    cards = []
    size_cards = []
    path = File.objects.get(id=5).files

    book = openpyxl.load_workbook(filename=path)
    sheet_name ={'name': "Стандарт", 'link': 'http://f-mebel.ua/shkafyi-kupe/seriya-standart/'},\
                {'name': "Комфорт", 'link': 'http://f-mebel.ua/shkafyi-kupe/seriya-komfort/'},\
                {'name': "Ультра", 'link': 'http://f-mebel.ua/shkafyi-kupe/seriya-ultra/'},\
                {'name': "Люкс", 'link': 'http://f-mebel.ua/shkafyi-kupe/seriya-lyuks/'}

    for n in sheet_name:
        sheet = book[n['name']]
        row = 0
        name = sheet[1][3].value

        while True:
            row += 1
            try:
                int(sheet[row][0].value)
                id += 1
                price = round(sheet[row][5].value * sheet[row][3].value + sheet[row][4].value)
                width = sheet[row][0].value
                height = sheet[row][2].value
                depth = sheet[row][1].value
                name_product = name + " " + str(sheet[row][0].value) + '*' + str(sheet[row][2].value)+ '*' + str(sheet[row][1].value)

                cards.append({
                    'id': id,
                    'name': name_product,
                    'prom':name_product
                })

                size_cards.append({
                    'id': id,
                    'w': num_check(width),
                    'h': num_check(height),
                    'd': num_check(depth),
                    'price': num_check(price)
                })

                print(name_product)

            except:
                if sheet[row][0].value == "Додаткова  комплектація":
                    break
                else:
                    pass

    #Додаємо записи в базу
    stock = []

    for item in cards:
        category = Category.objects.get(id=3)
        manufacturer = 4
        external_category = 'get_products_fenix'

        change_category(manufacturer, 'shafi', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]
                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                print('old', product[0].name)

                history = History.objects.create(
                            name=f"Оновлено товар",
                            description=product[0].name
                        )
                history.save()

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],
                )

                product.category.add(category)

                for sub in item['subcategory']:
                    subcategory = Subcategory.objects.get(id=sub)
                    product.subcategory.add(subcategory)

                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            is_main=main_price,
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                        )

                        prace_product.save()
                        
                        main_price = False

                print('new', item['name'])

                history = History.objects.create(
                            name=f"Додано товар",
                            description=item['name']
                        )
                history.save()

                stock.append(product.id)

        except Exception as ex:
            
            print(ex)
    
    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            history = History.objects.create(
                            name=f"Видалино товар",
                            description=product.name
                        )
            history.save()

            print('Видалино: ', product.name)
            
    
def get_shafi_neman():
    path =File.objects.get(id=6).files
    print(path)
    book = openpyxl.load_workbook(filename=path)
    sheet_1 = book.worksheets[0]

    prod_id = 0
    cards = []
    size_cards = []
    img_cards = []

    for r in range(sheet_1.max_row):
        r += 1

        cell = str(sheet_1[r][3].value)
        match = re.search('Шафа', cell)
        match2 = re.search('шафа', cell)
        match3 = re.search('Стелаж', cell)
        match4 = re.search('стелаж', cell)
        match5 = re.search('Шкаф', cell)

        if match or match2 or match3 or match4 or match5:
            prod_id += 1
            prom = str(sheet_1[r][0].value)
            name = str(sheet_1[r][3].value)
            des = str(sheet_1[r][37].value)
            name_ru = str(sheet_1[r][5].value)
            des_ru = str(sheet_1[r][38].value)
            price = str(sheet_1[r][9].value)[:-3]
            img_list = str(sheet_1[r][15].value).split(';')
            img_list = [line.rstrip() for line in img_list]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'name_ru': name_ru,
                'des':des,
                'des_ru':des_ru,
                'type':3
            })

            size_cards.append({
                'id': prod_id,
                'w':'',
                'h':'',
                'd':'',
                'price':price
            })

            for i in img_list:
                try:
                    img = i
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_name = str(img_name).replace('%', '')
                    img_cards.append({
                                'id': prod_id,
                                'img': img_name,
                                'url': img
                            })
                except:
                    pass

            print('--------------------------')
            print(prom)
            print(name, price)


    for r in range(sheet_1.max_row):
        r += 1

        cell = str(sheet_1[r][5].value)
        match = re.search('Пенал', cell)

        if match:
            prod_id += 1
            prom = str(sheet_1[r][0].value)
            name = str(sheet_1[r][5].value)
            des = str(sheet_1[r][37].value)
            name_ru = str(sheet_1[r][6].value)
            des_ru = str(sheet_1[r][38].value)
            price = str(sheet_1[r][9].value)[:-3]
            img_list = str(sheet_1[r][15].value).split(';')
            img_list = [line.rstrip() for line in img_list]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'name_ru': name_ru,
                'des':des,
                'des_ru':des_ru,
                'type':4
            })

            size_cards.append({
                'id': prod_id,
                'w':'',
                'h':'',
                'd':'',
                'price':price
            })

            for i in img_list:
                try:
                    img = i
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_name = str(img_name).replace('%', '')
                    img_bytes = requests.get(img).content
                    with open("/home/ay507291/notacomforta.pl.ua/www/media/" + img_name, "wb") as f:
                        f.write(img_bytes)

                    img_cards.append({
                        'id': prod_id,
                        'img': img_name
                    })
                except:
                    pass

            print('--------------------------')
            print(prom)
            print(name, price)

    #Додаємо записи в базу
    '''for item in cards:
        category = Category.objects.get(id=3)
        subcategory = Subcategory.objects.get(id=item['type'])
        manufacturer = 10
        external_category = 'shafi'
        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                print('old', product[0].name)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)
                product.subcategory.add(subcategory)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:

                        images_product = ProductImage.objects.create(
                            product=product,
                            image="/media/product_images/" + str(i['img']),
                            is_main=main_img
                        )

                        img_bytes = requests.get(i['url'], headers=HEADERS).content
                        with open(product_images_path + i['img'], "wb") as f:
                            f.write(img_bytes)

                        main_img = False
                        images_product.save()

                product.save()

                print('new', item['name'])

        except Exception as ex:
            
            print(ex)'''
            

def get_shafi_mixmebli():
    img_cards = []
    cards = []
    size_cards = []
    url = File.objects.get(id=7).url
    response = requests.get(url, headers=HEADERS)
    categoryId = ['73',]


    print('response ', response.status_code)

    if response.status_code == 200:
        root = ET.fromstring(response.content)

        for offer in root.find('.//shop/offers').iter('offer'):
            if offer.find('categoryId').text in categoryId and offer.attrib.get('available') == 'true':
                offer_id = offer.get('id')
                print(f"Offer ID: {offer_id}")

                description = offer.find('description').text
                params_html = "<table border='1'>"

                width = ''
                depth = ''
                height =''

                for param in offer.iter('param'):
                    if param.attrib.get('name') == 'Ширина':
                        width = param.text

                    if param.attrib.get('name') == 'Глибина':
                        depth = param.text
                    
                    if param.attrib.get('name') == 'Висота':
                        height = param.text

                    params_html += f"<tr><td>{param.attrib.get('name')}</td><td>{param.text}</td></tr>"
                
                params_html += "</table><br>"

                cards.append({
                    'id': offer_id,
                    'prom':offer_id,
                    'name': offer.find('name').text,
                    'des': description + params_html,
                })

                size_cards.append({
                    'id': offer_id,
                    'w':width,
                    'h':height,
                    'd':depth,
                    'price':str(offer.find('price').text).replace('.00', '')
                })

                for i in offer.findall('picture'):
                    img = str(i.text)
                    try:
                        img_name = img.split('/')[len(img.split('/')) - 1]
                        img_cards.append({
                            'id': offer_id,
                            'img': img_name,
                            'url': img
                        })

                    except Exception as ex:
                        pass
                
                print(f"{offer.find('name').text} {str(offer.find('price').text).replace('.00', '')}")
                print(50 * '-')

    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=3)
        manufacturer = 27
        external_category = 'shafi'

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                print('old', product[0].name)

                history = History.objects.create(
                            name=f"Оновлено товар",
                            description=product[0].name
                        )
                history.save()

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],
                    description=item['des'],
                    
                )

                product.category.add(category)

                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False
                
                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                            print(f"Завантажено: {i['url']}")
                        except Exception as ex:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )
                            print(ex)

                        main_img = False
                        images_product.save()

                
                print('new', item['name'])

                history = History.objects.create(
                            name=f"Додано товар",
                            description=item['name']
                        )
                history.save()

        except Exception as ex:
            
            print(ex)

    
            
def get_shafi_svitmebliv():
    print('Шафа ...')
    path = File.objects.get(id=13).files
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[5]

    
    modul_cards = []
    product_cards = []
    size_cards = []
    prod_id = 0

    modul_name = 'Модульні шафи "Спейс"'
    prom = modul_name.replace(' ', '').lower()


    modul_cards.append({
                'id': 0,
                'prom':prom,
                'name': modul_name
            })

    for row in range(sheet.max_row):
        row += 1
        column = 1
        
        name = sheet[row][0].value
        price = sheet[row][column].value

        

        if name and type(price) == int:
            column += 1
            find_price = True

            product_cards.append({
                    'modul_id': 0,
                    'prom':name.replace(' ', '').lower(),
                    'id': prod_id,
                    'name': name,
                })

            size_cards.append({
                'id': prod_id,
                'price':price,
            })

            if sheet[row][column].value:

                

                
                while find_price:
                    row -= 1
                    add_price = sheet[row][column].value

                    if type(add_price) == int:
                        
                        while True:
                            
                            add_price = sheet[row][column].value

                            if type(add_price) == int:
                                
                                size_cards.append({
                                    'id': prod_id,
                                    'price':price+add_price,
                                })
                        
                            else:
                                find_price = False
                                break

                            column += 1


            prod_id += 1

    #Оновлення товара
    stock = []

    for m in modul_cards:
        manufacturer = 2
        external_category = 'get_shafi_svitmebliv'
        category = Category.objects.get(id=3)
        
        seria = Seria.objects.filter(external_id = m['prom'], manufacturer_id = manufacturer)

        if seria.exists():
            #Оновлюємо
            seria = seria.first()
            print(f"old: {seria.name}")

            history = History.objects.create(
                            name=f"Оновлено комплети товарів",
                            description=seria.name
                        )
            history.save()

        else:
            #Додаємо
            seria = Seria(
                name=m['name'],
                external_id=m['prom'],
                manufacturer_id=manufacturer
            )
            seria.save()
            print(f"new: {m['name']}")

            history = History.objects.create(
                            name=f"Додано комплети товарів",
                            description=m['name']
                        )
            history.save()

        for p in product_cards:
                if p['modul_id'] == m['id']:

                    product = Product.objects.filter(external_id = p['prom'],\
                        seria = seria, manufacturer_id = manufacturer,\
                            external_category = external_category)

                    
                    if product.exists():
                        #Оновлюємо
                        product = product.first()
                        price = ProductPrice.objects.filter(product = product)
                        index = 0

                        for s in size_cards:
                            if s['id'] == p['id']:

                                try:
                                    curent_price = price[index]
                                    curent_price.price = s['price']
                                    curent_price.save()

                                    index += 1
                                
                                except:
                                    prace_product = ProductPrice.objects.create(
                                        product=product.first(),
                                        price=s['price'],
                                        info=s['info'],
                                    )

                                    prace_product.save()

                        print(f"old: {  product.name}")

                        history = History.objects.create(
                            name=f"Оновлено товар",
                            description=product.name
                        )
                        history.save()

                        stock.append(product.id)

                    else:
                        #Додаємо

                        product = Product(
                            seria=seria,
                            external_id=p['prom'],
                            manufacturer_id=manufacturer,
                            external_seria=m['prom'],
                            external_category=external_category,
                            name=p['name'],
                        )   
                        product.save()
                        

                        main_price = True

                        for s in size_cards:
                            if s['id'] == p['id']:

                                prace_product = ProductPrice.objects.create(
                                    product=product,
                                    price=s['price'],
                                    is_main=main_price,
                                )

                                prace_product.save()
                                
                                main_price = False
                        
                        product.category.add(category)

                        print(f"new: {p['name']}")

                        history = History.objects.create(
                            name=f"Додано товар",
                            description=product.name
                        )
                        history.save()

                        stock.append(product.id)


        print('-'*50)

    
    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            history = History.objects.create(
                            name=f"Видалино товар",
                            description=product.name
                        )
            history.save()

            print('Видалино: ', product.name)

    
    Seria.objects.filter(products__isnull=True).delete()
        


def get_stelazhi_svitmebliv():
    print('Шафа розпашні ...')
    path = File.objects.get(id=13).files
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]

    
    cards = []
    size_cards = []
    prod_id = 0

    

    for r in range(sheet.max_row):
        r += 1
        for c in range(sheet.max_column):
            cell = str(sheet[r][c].value).split(' ')[0]
            if cell == 'Стелаж':
                name = str(sheet[r][c].value)
                price = int(sheet[r][c+4].value)
                prom = name.replace(' ', '').lower()
                
                cards.append({
                    'prom':prom, 
                    'id': prod_id,
                    'name': name,
                })

                size_cards.append({
                    'id': prod_id,
                    'price':price
                })

                prod_id += 1


            if cell == "Стелажі":
                while True:
                    r += 1

                    try:
                        name = str(sheet[r][c].value)
                        price = int(sheet[r][c+4].value)
                        prom = name.replace(' ', '').lower()

                        cards.append({
                            'prom':prom, 
                            'id': prod_id,
                            'name': name,
                        })

                        size_cards.append({
                            'id': prod_id,
                            'price':price
                        })

                        prod_id += 1

                    except:
                        break

    #Додаємо записи в базу
    stock = []

    for item in cards:
        category = Category.objects.get(id=3)
        subcategory = Subcategory.objects.get(id=4)
        manufacturer = 2
        external_category = 'get_stelazhi_svitmebliv'

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]
                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                print('old', product[0].name)

                history = History.objects.create(
                            name=f"Оновлено товар",
                            description=product[0].name
                        )
                history.save()

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],
                )

                product.category.add(category)
                product.subcategory.add(subcategory)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                product.save()

                print('new', item['name'])

                history = History.objects.create(
                            name=f"Додано товар",
                            description=item['name']
                        )
                history.save()

                stock.append(product.id)

        except Exception as ex:
            
            print("Помилка при оновленню товара: ", ex)

    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            history = History.objects.create(
                            name=f"Видалино товар",
                            description=product.name
                        )
            history.save()

            print('Видалино: ', product.name)