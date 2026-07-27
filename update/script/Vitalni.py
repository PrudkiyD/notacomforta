from catalog.models import Product, ProductImage, ProductPrice, Category, Seria
from .Tools import HEADERS, change_category, change_category_modul, file_path, product_images_path, num_check
from update.models import File
from bs4 import BeautifulSoup
import requests
import time
import openpyxl
import re
import logging

logger = logging.getLogger(__name__)

def get_vitalni_products_bmk():
    path = File.objects.get(id=23).files
    logger.info(path)

    in_stock = []
    external_category = 'vitalni_products_bmk'

    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    for i in range(sheet.max_row):
        i +=1
        cell = str(sheet[i][1].value)
        equal = re.search('ВІТАЛЬНІ', cell)
        if equal:
            while True:
                i +=1
                try:
                    name = sheet[i][2].value
                    price = int(sheet[i][7].value)
                    prom = str(sheet[i][3].value)

                    logger.info('-'*50)

                    #Оновлюємо дані
                    category = Category.objects.get(id=5)
                    manufacturer_id = 7

                    in_stock.append({
                        'id':prom,
                        'manufacturer_id':manufacturer_id,
                        'external_category': 'modul',
                    })


                    #Оновлюємо дані про серію
                    serias = Seria.objects.filter(external_id=prom,\
                                                  manufacturer_id=manufacturer_id)
                    
                    if serias.exists():
                        #Оновлюємо
                        seria = serias.first()

                        logger.info(f"old modul: {seria.name}")

                        

                    else:
                        #Додаємо
                        seria = Seria.objects.create(
                            manufacturer_id=manufacturer_id,
                            name=name,
                            external_id=prom
                        )
                        seria.save()

                        logger.info(f'new modul: {name}')

                        
                    #---------------------------------------------------
                    

                    #Оновлюємо дані про товар
                    products = Product.objects.filter(external_id=prom,\
                                                      manufacturer_id=manufacturer_id,\
                                                        external_category='modul',\
                                                            external_seria=prom)

                    if products.exists():
                        #Оновлюємо
                        product = products.first()

                        product_price = product.prices.filter(is_main=True).first()
                        product_price.price = price
                        product_price.save()

                        logger.info(f"old: {product.name}")

                        
                    
                    else:
                        #Додаємо
                        product = Product.objects.create(
                            seria=seria,
                            manufacturer_id=manufacturer_id,
                            name=name,
                            external_id=prom,
                            external_category='modul',
                            external_seria=prom
                        )
                        product.category.add(category)

                        product.save()

                        product_price = ProductPrice.objects.create(
                            product=product,
                            is_main=True,
                            price=price
                        )

                        logger.info(f'new: {name}')

                        
                    #---------------------------------------------------
                except:
                    row = i
                    break

    for r in range(sheet.max_row):
        r += row
        cell = str(sheet[r][1].value)

        if cell =='СТОЛИ':
                    break

        elif str(sheet[r+1][1].value) == '1':
            name = cell
            prom = cell.lower()
            prom = prom.replace(' ', '')

            logger.info('-'*50)

            #Оновлюємо дані
            category = Category.objects.get(id=5)
            manufacturer_id = 7

            in_stock.append({
                'id':prom,
                'manufacturer_id':manufacturer_id,
                'external_category': 'modul',
            })


            #Оновлюємо дані про серію
            serias = Seria.objects.filter(external_id=prom,\
                                            manufacturer_id=manufacturer_id)
            
            if serias.exists():
                #Оновлюємо
                seria = serias.first()

                logger.info(f"old modul: {seria.name}")

                

            else:
                #Додаємо
                seria = Seria.objects.create(
                    manufacturer_id=manufacturer_id,
                    name=name,
                    external_id=prom
                )
                seria.save()

                logger.info(f'new modul: {name}')

                
            #---------------------------------------------------
            

            #Оновлюємо дані про товар
            products = Product.objects.filter(external_id=prom,\
                                                manufacturer_id=manufacturer_id,\
                                                external_category='modul',\
                                                    external_seria=prom)

            if products.exists():
                #Оновлюємо
                product = products.first()

                product_price = product.prices.filter(is_main=True).first()
                product_price.price = price
                product_price.save()

                logger.info(f"old: {product.name}")

                
            
            else:
                #Додаємо
                product = Product.objects.create(
                    seria=seria,
                    manufacturer_id=manufacturer_id,
                    name=name,
                    external_id=prom,
                    external_category='modul',
                    external_seria=prom
                )
                product.category.add(category)

                product.save()

                product_price = ProductPrice.objects.create(
                    product=product,
                    is_main=True,
                    price=price
                )

                logger.info(f'new: {name}')

                
            #---------------------------------------------------

            logger.info(name, prom)

            while True:
                r += 1
                try:
                    name = sheet[r][2].value
                    price = int(sheet[r][7].value)
                    prom = str(sheet[r][3].value)
                    
                    #Оновлюємо дані про товар
                    products = Product.objects.filter(external_id=prom,\
                                                        manufacturer_id=manufacturer_id,\
                                                        external_category='modul',\
                                                            external_seria=prom)

                    if products.exists():
                        #Оновлюємо
                        product = products.first()

                        product_price = product.prices.filter(is_main=True).first()
                        product_price.price = price
                        product_price.save()

                        logger.info(f"old: {product.name}")

                        
                    
                    else:
                        #Додаємо
                        product = Product.objects.create(
                            seria=seria,
                            manufacturer_id=manufacturer_id,
                            name=name,
                            external_id=prom,
                            external_category='modul',
                            external_seria=prom
                        )
                        product.category.add(category)

                        product.save()

                        product_price = ProductPrice.objects.create(
                            product=product,
                            is_main=True,
                            price=price
                        )

                        logger.info(f'new: {name}')

                        
                    #---------------------------------------------------

                    logger.info(name, prom, price)
                except:
                    break


def get_vitalni_products_gerbor():
    path = File.objects.get(id=22).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    for i in range(sheet.max_row):
        i +=1
        cell = str(sheet[i][2].value)
        equal = re.search('Вітальні', cell)
        if equal:
            logger.info("Розділ вітальні знайдено ...")
            while True:
                i +=1
                try:
                    name = sheet[i][2].value
                    price = int(sheet[i][7].value)
                    prom = str(sheet[i][3].value)

                    logger.info('-'*50)

                    #Оновлюємо дані
                    category = Category.objects.get(id=5)
                    manufacturer_id = 8


                    #Оновлюємо дані про серію
                    serias = Seria.objects.filter(external_id=prom,\
                                                  manufacturer_id=manufacturer_id)
                    
                    if serias.exists():
                        #Оновлюємо
                        seria = serias.first()

                        logger.info(f"old modul: {seria.name}")

                        

                    else:
                        #Додаємо
                        seria = Seria.objects.create(
                            manufacturer_id=manufacturer_id,
                            name=name,
                            external_id=prom
                        )
                        seria.save()

                        logger.info(f'new modul: {name}')

                        
                    #---------------------------------------------------
                    

                    #Оновлюємо дані про товар
                    products = Product.objects.filter(external_id=prom,\
                                                      manufacturer_id=manufacturer_id,\
                                                        external_category='get_vitalni_products_gerbor',\
                                                            external_seria=prom)

                    if products.exists():
                        #Оновлюємо
                        product = products.first()
                        product.external_category = 'get_vitalni_products_gerbor'
                        product.save()

                        product_price = product.prices.filter(is_main=True).first()
                        product_price.price = price
                        product_price.save()

                        logger.info(f"old: {product.name}")

                        
                    
                    else:
                        #Додаємо
                        product = Product.objects.create(
                            seria=seria,
                            manufacturer_id=manufacturer_id,
                            name=name,
                            external_id=prom,
                            external_category='get_vitalni_products_gerbor',
                            external_seria=prom
                        )
                        product.category.add(category)

                        product.save()

                        product_price = ProductPrice.objects.create(
                            product=product,
                            is_main=True,
                            price=price
                        )

                        logger.info(f'new: {name}')

                        
                    #---------------------------------------------------
                    

                except Exception as ex:
                    logger.info(ex)
                    break

        
def get_vitalni_products_svitmebliv():
    logger.info('Вітальні ...')
    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0

    path = File.objects.get(id=13).files
    
    
    logger.info(f"Вітальні: {path}")
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    find = False
    for i in range(sheet.max_row):
        i += 1
        try:
            name = sheet[i][0].value
            price = int(sheet[i][4].value)

            prom = name.replace(' ', '').lower()
            modul_cards.append({
                'id': modul_id,
                'prom':prom,
                'name': name
            })

            product_cards.append({
                'modul_id': modul_id,
                'id': prod_id,
                'prom':prom,
                'name': name,
                'w': '',
                'des': '',
                'price': price
            })

            modul_id += 1
            prod_id += 1

            find = True

            logger.info('>',name)
        except:
            if find:
                break
            else:
                pass
    

    
    logger.info(f"Вітальня: {path}")
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    for c in range(sheet.max_column):
        for r in range(sheet.max_row):
            r += 1
            
            cell = str(sheet[r][c].value)
            match = re.search("Вітальня „", cell)
            if match:
                name = cell[8:]
                column = c + 4
                row = r
                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })

                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'w': '',
                    'des': '',
                    'price': 0
                })

                logger.info('>',name)

                while True:
                    row += 1

                    try:
                        
                        name = str(sheet[row][c].value)
                        price = int(sheet[row][column].value)

                        prod_id += 1
                        prom = name.replace(' ', '').lower()

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': '',
                            'des': '',
                            'price': price
                        })

                        logger.info('>',name)

                    except:
                        break

                modul_id += 1
                prod_id += 1
            else:
                pass
    
    
    logger.info(f"Модульна система: {path}")
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[1]
    for c in range(sheet.max_column):
        for r in range(sheet.max_row):
            r += 1
            
            cell = str(sheet[r][c].value)
            match = re.search("модульна система", cell)
            if match:
                name = cell[16:]
                column = c + 4
                row = r
                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })

                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'w': '',
                    'des': '',
                    'price': 0
                })

                logger.info('>',name)

                while True:
                    row += 1

                    try:
                        
                        name = str(sheet[row][c].value)
                        price = int(sheet[row][column].value)

                        prod_id += 1
                        prom = name.replace(' ', '').lower()

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': '',
                            'des': '',
                            'price': price
                        })
                        logger.info(price)

                    except Exception as ex:
                        logger.info(ex)
                        break

                modul_id += 1
                prod_id += 1
            else:
                pass
    

    #Оновлення товара
    stock = []

    for m in modul_cards:
        manufacturer = 2
        external_category = 'get_vitalni_products_svitmebliv'
        category = Category.objects.get(id=5)
        
        seria = Seria.objects.filter(external_id = m['prom'], manufacturer_id = manufacturer)

        if seria.exists():
            #Оновлюємо
            seria = seria.first()
            logger.info(f"old: {seria.name}")

            

        else:
            #Додаємо
            seria = Seria(
                name=m['name'],
                external_id=m['prom'],
                manufacturer_id=manufacturer
            )
            seria.save()
            logger.info(f"new: {m['name']}")

            

        for p in product_cards:
                if p['modul_id'] == m['id']:

                    change_category_modul(manufacturer, 'modul', p['prom'], external_category, seria)

                    product = Product.objects.filter(external_id = p['prom'],\
                        seria = seria, manufacturer_id = manufacturer,\
                            external_category = external_category)

                    
                    if product.exists():
                        #Оновлюємо
                        product = product.first()
                        price = ProductPrice.objects.filter(product = product)

                        if price.exists():
                            price = price.first()

                            price.price = p['price']
                            price.save()

                        logger.info(f"old: {  product.name}")

                        

                        stock.append(product.id)

                    else:
                        #Додаємо

                        product = Product(
                            seria=seria,
                            external_id=p['prom'],
                            manufacturer_id=manufacturer,
                            external_seria=m['prom'],
                            external_category=external_category,
                            name=p['name'],
                        )   
                        product.save()
                        

                        price = ProductPrice(
                            product=product,
                            is_main=True,
                            price=p['price'],
                        )

                        price.save()
                        
                        product.category.add(category)

                        logger.info(f"new: {p['name']}")

                        

                        stock.append(product.id)


        logger.info('-'*50)

    
    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info('Видалино: ', product.name)

    
    Seria.objects.filter(products__isnull=True).delete()


def get_vitalni_products_comfortmebli():
    path = File.objects.get(id=31).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    
    modul_cards = []
    product_cards = []
    img_cards = []
    modul_id = 0
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            modul_id += 1

            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            width = str(sheet[r][5].value)

            img_list =str(sheet[r][14].value).split(';')
            img_list.insert(0, str(sheet[r][13].value)) 

            des = str(sheet[r][15].value)


            product_cards.append({
                'modul_id':modul_id,
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
                'width':width,
                'price':price
            })

            modul_cards.append({
                'prom':prom,
                'id':modul_id,
                'name':name,
            })

            logger.info(prom, '-->', name, '->', price)

            for img in img_list:
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_cards.append({
                        'id': prod_id,
                        'img': img_name,
                        'url': img
                    })

                except Exception as ex:
                    logger.info(f'///-------{ex}---------///')

            logger.info('-------------------------')


    #Оновлення товара
    stock = []

    for m in modul_cards:
        manufacturer = 1
        external_category = 'get_vitalni_products_comfortmebli'
        category = Category.objects.get(id=5)
        
        seria = Seria.objects.filter(external_id = m['prom'], manufacturer_id = manufacturer)

        if seria.exists():
            #Оновлюємо
            seria = seria.first()
            logger.info(f"old: {seria.name}")

            

        else:
            #Додаємо
            seria = Seria(
                name=m['name'],
                external_id=m['prom'],
                manufacturer_id=manufacturer
            )
            seria.save()
            logger.info(f"new: {m['name']}")

            

        for p in product_cards:
                if p['modul_id'] == m['id']:
                    product = Product.objects.filter(external_id = p['prom'],\
                        seria = seria, manufacturer_id = manufacturer,\
                            external_category = external_category)

                    
                    if product.exists():
                        #Оновлюємо
                        product = product.first()
                        price = ProductPrice.objects.filter(product = product)

                        if price.exists():
                            price = price.first()

                            price.price = p['price']
                            price.save()

                        logger.info(f"old: {  product.name}")

                        

                        stock.append(product.id)

                    else:
                        #Додаємо

                        product = Product(
                            published=True,
                            seria=seria,
                            external_id=p['prom'],
                            manufacturer_id=manufacturer,
                            external_seria=m['prom'],
                            external_category=external_category,
                            name=p['name'],
                            description=p['des']
                        )

                        product.save()
                        

                        price = ProductPrice(
                            product=product,
                            is_main=True,
                            price=p['price'],
                            width=p['width'],
                        )

                        price.save()
                        main_img = True

                        for i in img_cards:
                            if i['id'] == p['id']:
                                logger.info(i['url'])
                                try:
                                    img_bytes = requests.get(i['url'], headers=HEADERS).content
                                    with open(product_images_path + i['img'], "wb") as f:
                                        f.write(img_bytes)

                                    images_product = ProductImage.objects.create(
                                        product=product,
                                        image=str(i['img']),
                                        is_main=main_img
                                    )
                                except Exception as ex:
                                    logger.info(ex)
                                    images_product = ProductImage.objects.create(
                                        product=product,
                                        image=str(i['url']),
                                        is_main=main_img
                                    )

                                main_img = False
                                images_product.save()


                        
                        product.category.add(category)

                        logger.info(f"new: {p['name']}")

                        

                        stock.append(product.id)


        logger.info('-'*50)

    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info('Видалино: ', product.name)

    Seria.objects.filter(products__isnull=True).delete()