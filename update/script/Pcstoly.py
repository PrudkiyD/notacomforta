from catalog.models import Product, ProductImage, ProductPrice, Category, Subcategory
from .Tools import HEADERS, change_category, change_category_modul, file_path, product_images_path, num_check
import xml.etree.ElementTree as ET
from update.models import File
from bs4 import BeautifulSoup
import requests
import openpyxl
import time
import re
import logging

logger = logging.getLogger(__name__)

def get_pcstoly_comfortmebli():
    path =File.objects.get(id=36).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    img_cards = []
    size_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][10].value).split(';')
            img_list.insert(0, str(sheet[r][8].value)) 

            des = str(sheet[r][9].value)


            cards.append({
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
            })

            size_cards.append({
                'id': prod_id,
                'w':str(sheet[r][5].value),
                'h':str(sheet[r][4].value),
                'd':str(sheet[r][6].value),
                'price':price
            })

            logger.info(prom, '-->', name, '->', price)


            for img in img_list:
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_cards.append({
                        'id': prod_id,
                        'img': img_name,
                        'url': img
                    })

                except Exception as ex:
                    logger.info(f'///-------{ex}---------///')

    #Додаємо записи в базу

    stock=[]

    for item in cards:
        category = Category.objects.get(id=13)
        manufacturer = 1
        external_category = 'get_pcstoly_comfortmebli'

        change_category(manufacturer, 'pcstoly', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            is_main=main_price,
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                            logger.info(f"Завантажено: {i['url']}")
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                

                logger.info('new', item['name'])

                

                stock.append(product.id)

        except Exception as ex:
            
            logger.info(ex)



def get_pcstoly_matrolux():
    img_cards = []
    cards = []
    size_cards = []

    logger.info('Start ...')
    req = requests.get(File.objects.get(id=4).url, headers=HEADERS)  
    src = req.text
    logger.info('Get src ---///')
    soup = BeautifulSoup(src, 'xml')
    item = soup.find_all('entry')

    for i in item:
        try: 
            title = i.find('title').get_text() if i.find('title') else None
            product_type = i.find('product_type').get_text() if i.find('product_type') else None
            product_id = i.find('id').get_text() if i.find('id') else None
            price = str(i.find('price').get_text()).replace('.00 UAH', '') if i.find('price') else None
            description = str(i.find('description').get_text()).replace('характеристики новых шкафов: ', '') if i.find('description') else ' '
            additional_image_link = i.find('additional_image_link').get_text() if i.find('additional_image_link') else None
            image_link = i.find('image_link').get_text() if i.find('image_link') else None
            img_list = [additional_image_link, image_link]

            sale = False
            old_price = None

            if i.find('sale_price'):
                old_price = price
                sale = True
                price = str(i.find('sale_price').get_text()).replace('.00 UAH', '') if i.find('sale_price') else None
            
            
            
            if  product_type == "Корпусні меблі":
                if len(str(product_id).split('-')) == 1:
                
                    update = False

                    for cat_num in range(50):
                        if cat_num == 0:
                            cat = i.find(f'category').get_text() if i.find(f'category') else None
                        
                        else:
                            cat = i.find(f'category{cat_num}').get_text() if i.find(f'category{cat_num}') else None


                        if cat == "Комп'ютерні " or cat == "Офісні" or cat == "Пісьмові ":
                            update = True


                    if update:
                        logger.info(title)
                        logger.info('-'*50)

                        cards.append({
                            'prom':product_id,
                            'id': product_id,
                            'name': title,
                            'des':description,
                            
                        })

                        size_cards.append({
                            'id': product_id,
                            'old_price':old_price,
                            'price':price,
                            'sale':sale
                            
                        })

                        
                        for img in img_list:
                            try:
                                img_name = img.split('/')[len(img.split('/')) - 1]
                                img_cards.append({
                                    'id': product_id,
                                    'img': img_name,
                                    'url': img
                                })

                                

                            except Exception as ex:
                                logger.info(f'///-------{ex}---------///')
                        
                        

                    

        except Exception as ex:
            logger.info('-'*50)
            logger.info(ex)
            logger.info('-'*50)

    
    #Додаємо записи в базу
    for item in cards:
        category = Category.objects.get(id=13)
        manufacturer = 3
        external_category = 'pcstoly'

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.old_price = s['old_price']
                        curent_price.sale = s['sale']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                product.save()

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            old_price=s['old_price'],
                            is_main=main_price,
                            sale=s['sale']
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                            logger.info(f"Завантажено: {i['url']}")
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                

                logger.info('new', item['name'])

                

        except Exception as ex:
            
            logger.info(ex)
    

def get_pcstoly_neman():
    path =File.objects.get(id=24).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    cards = []
    size_cards = []
    img_cards = []
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        size = []
        cell = str(sheet[r][3].value)
        match = re.search('письмовий', cell)
        match2 = re.search('Письмовий', cell)

        if match or match2:
            prod_id += 1
            prom = str(sheet[r][0].value)
            name = str(sheet[r][3].value)
            des = str(sheet[r][45].value)
            name_ru = str(sheet[r][5].value)
            des_ru = str(sheet[r][47].value)
            size = [None,None,None]
            img_list = str(sheet[r][17].value).split(';')
            img_list = [line.rstrip() for line in img_list]
            price = str(sheet[r][11].value)[:-3]

            cards.append({
                'id': prod_id,
                'prom':prom,
                'name': name,
                'des': des,
                'name_ru': name_ru,
                'des_ru': des_ru,
            })

            logger.info(prom, name)
            #logger.info(size)
            #logger.info(price)
            
            try:
                size_cards.append({
                    'id': prod_id,
                    'option': None,
                    'gear': None,
                    'w': size[1],
                    'h': None,
                    'd': size[0],
                    'price': price,
                })
            except:
                size_cards.append({
                    'id': prod_id,
                    'option': None,
                    'gear': None,
                    'w': size[0],
                    'h': None,
                    'd': None,
                    'price': price,
                })

            for img in img_list:                
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_cards.append({
                        'id': prod_id,
                        'img': img_name,
                        'url': img
                    })

                    #logger.info(img_name)

                except Exception as ex:
                    logger.info(f'///{ex}')
            
            logger.info('-'*50)

    #Оновлюємо ціни
    #Додаємо записи в базу

    stock=[]

    for item in cards:
        category = Category.objects.get(id=13)
        manufacturer = 10
        external_category = 'get_pcstoly_neman'

        change_category(manufacturer, 'stoly', item['prom'], external_category)

        try:
        #Оновлюємо ціну
            product = Product.objects.filter(
                external_id=item['prom'],
                manufacturer_id=manufacturer,
                external_category=external_category
            )

            if product.exists():
                price = ProductPrice.objects.filter(product=product.first())
                index = 0

                for s in size_cards:
                    if s['id'] == item['id']:
                        curent_price = price[index]

                        curent_price.price = s['price']
                        curent_price.save()

                        index += 1

                logger.info('old', product[0].name)

                

                stock.append(product[0].id)

            #Додаємо новий товар
            else:
                product = Product.objects.create(
                    published=True,  
                    external_id=item['prom'],  
                    external_category=external_category,
                    manufacturer_id=manufacturer,  
                    name=item['name'],  
                    description=item['des'],
                )

                product.category.add(category)

                main_price = True

                for s in size_cards:
                    if s['id'] == item['id']:

                        prace_product = ProductPrice.objects.create(
                            product=product,
                            price=s['price'],
                            width=s['w'],
                            height=s['h'],
                            depth=s['d'],
                            is_main=main_price,
                            setup=s['option']
                        )

                        prace_product.save()
                        
                        main_price = False

                main_img = True

                for i in img_cards:
                    if i['id'] == item['id']:
                        try:
                            img_bytes = requests.get(i['url'], headers=HEADERS).content
                            with open(product_images_path + i['img'], "wb") as f:
                                f.write(img_bytes)

                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['img']),
                                is_main=main_img
                            )
                        except:
                            images_product = ProductImage.objects.create(
                                product=product,
                                image=str(i['url']),
                                is_main=main_img
                            )

                        main_img = False
                        images_product.save()

                product.save()

                logger.info('new', item['name'])

                

                stock.append(product.id)

        except Exception as ex:
            
            logger.info(ex)

    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info('Видалино: ', product.name)