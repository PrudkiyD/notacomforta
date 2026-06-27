from catalog.models import Product, ProductImage, ProductPrice, Category, Seria
from .Tools import HEADERS, change_category, change_category_modul, file_path, product_images_path, num_check
from update.models import File
import requests
import openpyxl
import re
import logging

logger = logging.getLogger(__name__)

def get_dytyachi_products_svitmebliv():
    logger.info('Дитячі ...')
    modul_cards = []
    product_cards = []
    modul_id = 0
    prod_id = 0

    path = File.objects.get(id=13).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    
    
    sheet = book.worksheets[0]
    for r in range(sheet.max_row):
        r += 1

        if r == 100:
            break

        for c in range(sheet.max_column):
            cell = str(sheet[r][c].value)
            match_1 = re.search("дитяча „", cell)
            match_2 = re.search('дитяча', cell)
            if match_1 or match_2:

                modul_id += 1
                prod_id += 1

                name = cell.replace('*','')
                column = c + 4
                row = r
                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })


                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'w': None,
                    'des': None,
                    'price': 0
                })

                while True:
                    row += 1
                    try:
                        name = str(sheet[row][c].value)
                        price = int(sheet[row][column].value)
                        prod_id += 1
                        prom = name.replace(' ', '').lower()

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': None,
                            'des': None,
                            'price': price
                        })

                    except:
                        break
            else:
                pass

    sheet = book.worksheets[1]
    for r in range(sheet.max_row):
        r += 1

        if r == 100:
            break

        for c in range(sheet.max_column):
            cell = str(sheet[r][c].value)
            match_1 = re.search("дитяча „", cell)
            match_2 = re.search('дитяча', cell)
            if match_1 or match_2:

                modul_id += 1
                prod_id += 1

                name = cell.replace('*','')
                column = c + 4
                row = r
                prom = name.replace(' ', '').lower()

                modul_cards.append({
                    'id': modul_id,
                    'prom':prom,
                    'name': name
                })


                product_cards.append({
                    'modul_id': modul_id,
                    'id': prod_id,
                    'prom':prom,
                    'name': name,
                    'w': None,
                    'des': None,
                    'price': 0
                })

                while True:
                    row += 1
                    try:
                        name = str(sheet[row][c].value)
                        price = int(sheet[row][column].value)
                        prod_id += 1
                        prom = name.replace(' ', '').lower()

                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': None,
                            'des': None,
                            'price': price
                        })
                        

                    except:
                        break
            else:
                pass
    
    sheet = book.worksheets[1]
    for r in range(sheet.max_row):
        r += 1

        if r == 100:
            break

        for c in range(sheet.max_column):

            cell = str(sheet[r][c].value)


            match_1 = re.search("Дитячі", cell)
            if match_1:
                row = r
                column = c+4

                while True:
                    try:
                        row += 1

                        name = str(sheet[row][c].value)
                        prom = name.replace(' ', '').lower()
                        price = int(sheet[row][column].value)

                        modul_id += 1
                        prod_id += 1

                        modul_cards.append({
                            'id': modul_id,
                            'prom':prom,
                            'name': name
                        })


                        product_cards.append({
                            'modul_id': modul_id,
                            'id': prod_id,
                            'prom':prom,
                            'name': name,
                            'w': None,
                            'des': None,
                            'price': price
                        })

                    except:
                        break


            else:
                pass

    
    #Оновлення товара

    stock = []

    for m in modul_cards:
        manufacturer = 2
        external_category = 'get_dytyachi_products_svitmebliv'
        category = Category.objects.get(id=9)
        
        seria = Seria.objects.filter(external_id = m['prom'], manufacturer_id = manufacturer)

        if seria.exists():
            #Оновлюємо
            seria = seria.first()
            logger.info(f"old: {seria.name}")

            

        else:
            #Додаємо
            seria = Seria(
                name=m['name'],
                external_id=m['prom'],
                manufacturer_id=manufacturer
            )
            seria.save()
            logger.info(f"new: {m['name']}")

            

        for p in product_cards:
                if p['modul_id'] == m['id']:

                    change_category_modul(manufacturer, 'modul', p['prom'], external_category, seria)

                    product = Product.objects.filter(external_id = p['prom'],\
                        seria = seria, manufacturer_id = manufacturer,\
                            external_category = external_category)

                    
                    if product.exists():
                        #Оновлюємо
                        product = product.first()
                        price = ProductPrice.objects.filter(product = product)

                        if price.exists():
                            price = price.first()

                            price.price = p['price']
                            price.save()

                        logger.info(f"old: {  product.name}")

                        

                        stock.append(product.id)

                    else:
                        #Додаємо

                        product = Product(
                            seria=seria,
                            external_id=p['prom'],
                            manufacturer_id=manufacturer,
                            external_seria=m['prom'],
                            external_category=external_category,
                            name=p['name'],
                        )
                        product.save()
                        

                        price = ProductPrice(
                            product=product,
                            is_main=True,
                            price=p['price'],
                        )
                        price.save()

                        
                        product.category.add(category)

                        logger.info(f"new: {p['name']}")

                        

                        stock.append(product.id)

        logger.info('-'*50)

    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info('Видалино: ', product.name)

    
    Seria.objects.filter(products__isnull=True).delete()



def get_dytyachi_products_comfortmebli():
    path = File.objects.get(id=34).files
    logger.info(path)
    book = openpyxl.load_workbook(filename=path)
    sheet = book.worksheets[0]
    
    modul_cards = []
    product_cards = []
    img_cards = []
    modul_id = 0
    prod_id = 0

    for r in range(sheet.max_row):
        r += 1
        cell = str(sheet[r][1].value)

        if r != 1:

            prod_id += 1
            modul_id += 1

            prom = str(sheet[r][0].value)
            name = str(sheet[r][1].value)
            price = str(sheet[r][2].value)

            img_list =str(sheet[r][12].value).split(';')
            img_list.insert(0, str(sheet[r][11].value)) 

            des = str(sheet[r][5].value)


            product_cards.append({
                'modul_id':modul_id,
                'prom':prom, 
                'id': prod_id,
                'name': name,
                'des': des,
                'price':price
            })

            modul_cards.append({
                'prom':prom,
                'id':modul_id,
                'name':name,
            })

            logger.info(prom, '-->', name, '->', price)

            for img in img_list:
                try:
                    img_name = img.split('/')[len(img.split('/')) - 1]
                    img_cards.append({
                        'id': prod_id,
                        'img': img_name,
                        'url': img
                    })

                except Exception as ex:
                    logger.info(f'///-------{ex}---------///')

            logger.info('-------------------------')


    #Оновлення товара
    stock = []

    for m in modul_cards:
        manufacturer = 1
        external_category = 'get_dytyachi_products_comfortmebli'
        category = Category.objects.get(id=9)
        
        seria = Seria.objects.filter(external_id = m['prom'], manufacturer_id = manufacturer)

        if seria.exists():
            #Оновлюємо
            seria = seria.first()
            logger.info(f"old: {seria.name}")

            

        else:
            #Додаємо
            seria = Seria(
                name=m['name'],
                external_id=m['prom'],
                manufacturer_id=manufacturer
            )
            seria.save()
            logger.info(f"new: {m['name']}")

            

        for p in product_cards:
                if p['modul_id'] == m['id']:
                    
                    change_category_modul(manufacturer, 'modul', p['prom'], external_category, seria)

                    product = Product.objects.filter(external_id = p['prom'],\
                        seria = seria, manufacturer_id = manufacturer,\
                            external_category = external_category)

                    
                    if product.exists():
                        #Оновлюємо
                        product = product.first()
                        price = ProductPrice.objects.filter(product = product)

                        if price.exists():
                            price = price.first()

                            price.price = p['price']
                            price.save()

                        logger.info(f"old: {  product.name}")

                        

                        stock.append(product.id)

                    else:
                        #Додаємо

                        product = Product(
                            published=True,
                            seria=seria,
                            external_id=p['prom'],
                            manufacturer_id=manufacturer,
                            external_seria=m['prom'],
                            external_category=external_category,
                            name=p['name'],
                            description=p['des']
                        )

                        product.save()
                        

                        price = ProductPrice(
                            product=product,
                            is_main=True,
                            price=p['price'],
                        )

                        price.save()
                        main_img = True

                        for i in img_cards:
                            if i['id'] == p['id']:
                                logger.info(i['url'])
                                try:
                                    img_bytes = requests.get(i['url'], headers=HEADERS).content
                                    with open(product_images_path + i['img'], "wb") as f:
                                        f.write(img_bytes)

                                    images_product = ProductImage.objects.create(
                                        product=product,
                                        image=str(i['img']),
                                        is_main=main_img
                                    )
                                except Exception as ex:
                                    logger.info(ex)
                                    images_product = ProductImage.objects.create(
                                        product=product,
                                        image=str(i['url']),
                                        is_main=main_img
                                    )

                                main_img = False
                                images_product.save()


                        
                        product.category.add(category)

                        logger.info(f"new: {p['name']}")

                        

                        stock.append(product.id)


        logger.info('-'*50)

    #Видаляємо товар якого немає в наявності
    products = Product.objects.filter(external_category=external_category)

    for product in products:

        if product.id not in stock:
            product.delete()

            

            logger.info('Видалино: ', product.name)

    
    Seria.objects.filter(products__isnull=True).delete()


